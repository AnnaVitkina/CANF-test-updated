"""
Compare and Find (CANF) - Match Shipments with Rate Card

This script:
1. Gets rate card from part4_rate_card_processing.py
2. Gets ETOF and LC dataframes from vocabular.py output (partly_df/vocabulary_mapping.xlsx)
3. Uses LC dataframe if present, otherwise uses ETOF dataframe
4. Matches shipments with Rate Card entries and identifies discrepancies
5. Validates business rules (e.g., Origin City -> Country + Postal Code)
"""

import pandas as pd
import os
import re

# Import business rules functions
from part4_rate_card_processing import (
    get_business_rules_lookup,
    process_business_rules,
    transform_business_rules_to_conditions,
    find_business_rule_columns
)

def normalize_value(value):
    """Converts a value to lowercase string, removes spaces and underscores, and handles NaN.
    Preserves leading zeros for postal codes and similar values."""
    if pd.isna(value):
        return None

    # Convert to string first
    str_value = str(value).strip()
    
    # Check if value starts with '0' and has more digits - likely a postal code or code with leading zeros
    # Don't convert to number to preserve leading zeros (e.g., "04123" should stay "04123", not become "4123")
    if str_value.startswith('0') and len(str_value) > 1 and str_value.lstrip('0').isdigit():
        # Preserve as string to keep leading zeros
        pass
    else:
        # Attempt to convert to a number if it looks like one, then convert to int if possible
        try:
            # Convert to float for numeric conversion
            num_val = float(str_value)
            if num_val == int(num_val):  # Check if it's an integer number (e.g., 7719.0)
                str_value = str(int(num_val))
            else:  # Keep as float if it has decimal (e.g., 123.45)
                str_value = str(num_val)
        except (ValueError, TypeError):
            # Not a number, keep original value
            pass

    # Apply lowercasing and cleaning
    return str_value.lower().replace(" ", "").replace("_", "")


def normalize_column_name(col_name):
    """Normalize column names for comparison (lowercase, remove spaces/underscores)."""
    if col_name is None:
        return None
    return str(col_name).lower().replace(" ", "").replace("_", "")


def load_business_rules_for_matching(rate_card_file_path):
    """
    Load business rules and create lookup structures for matching.
    
    Args:
        rate_card_file_path: Path to rate card file (relative to input folder)
    
    Returns:
        dict: Business rules lookup with:
            - 'rule_to_country': {rule_name: country_code}
            - 'rule_to_postal_codes': {rule_name: [postal_codes]}
            - 'business_rule_columns': set of column names containing business rules
            - 'column_to_rules': {column_name: [rule_names found in it]}
    """
    print(f"\n   [DEBUG] load_business_rules_for_matching called with: {rate_card_file_path}")
    try:
        result = get_business_rules_lookup(rate_card_file_path)
        print(f"   [DEBUG] get_business_rules_lookup returned:")
        print(f"      - result is None: {result is None}")
        if result:
            print(f"      - keys: {list(result.keys())}")
            print(f"      - rule_to_country count: {len(result.get('rule_to_country', {}))}")
            print(f"      - rule_to_postal_codes count: {len(result.get('rule_to_postal_codes', {}))}")
            print(f"      - business_rule_columns: {result.get('business_rule_columns', 'NOT SET')}")
        return result
    except Exception as e:
        print(f"   [WARNING] Could not load business rules: {e}")
        import traceback
        traceback.print_exc()
        return {
            'rule_to_country': {},
            'rule_to_postal_codes': {},
            'business_rule_columns': set(),
            'column_to_rules': {}
        }


def find_matching_business_rule_by_geo(etof_row, business_rule_col, business_rules_lookup):
    """
    Find a matching business rule based on country and postal code from ETOF row.
    Used when the business rule column value is NaN/empty - we check all rules for this column.
    
    Handles both:
    - Postal Code Zones: validates country + postal code prefix
    - Country Regions: validates country only
    
    Args:
        etof_row: Row from ETOF/LC dataframe
        business_rule_col: Name of the business rule column (e.g., 'Origin Postal Code Zone', 'Origin Country Region')
        business_rules_lookup: Business rules lookup dictionary
    
    Returns:
        tuple: (found_rule, validated_columns, message, failure_details)
            - found_rule: Name of the matching rule, or None
            - validated_columns: List of columns that were validated by this rule
            - message: Explanation message
            - failure_details: dict with details about why rules failed (for discrepancy reporting)
    """
    print(f"\n   {'='*60}")
    print(f"   [BUSINESS RULE GEO SEARCH] Searching for matching rule...")
    print(f"   {'='*60}")
    print(f"   [DEBUG] Business rule column: '{business_rule_col}'")
    
    failure_details = {
        'country_col': None,
        'postal_col': None,
        'actual_country': None,
        'actual_postal': None,
        'failed_rules': []  # List of (rule_name, failure_reason, expected_values)
    }
    
    if not business_rules_lookup:
        return None, [], "No business rules lookup", failure_details
    
    # Get rules that apply to this column
    column_to_rules = business_rules_lookup.get('column_to_rules', {})
    rule_to_country = business_rules_lookup.get('rule_to_country', {})
    rule_to_postal = business_rules_lookup.get('rule_to_postal_codes', {})
    
    # Find rules for this column (case-insensitive)
    rules_for_column = []
    for col, rules in column_to_rules.items():
        if normalize_column_name(col) == normalize_column_name(business_rule_col):
            rules_for_column = rules
            print(f"   [DEBUG] Found {len(rules)} rules for column '{col}'")
            break
    
    if not rules_for_column:
        print(f"   [DEBUG] No rules found for column '{business_rule_col}'")
        return None, [], f"No rules defined for column '{business_rule_col}'", failure_details
    
    # Determine if this is origin or destination
    col_lower = str(business_rule_col).lower()
    is_origin = 'origin' in col_lower or 'ship' in col_lower or 'from' in col_lower
    is_destination = 'destination' in col_lower or 'cust' in col_lower or 'to' in col_lower
    
    # Determine if this is a Country Region rule (only validates country) or Postal Code Zone (validates country + postal)
    is_country_region = 'country' in col_lower and 'region' in col_lower
    is_postal_code_zone = 'postal' in col_lower or 'zone' in col_lower
    
    print(f"   [DEBUG] Column type: {'ORIGIN' if is_origin else 'DESTINATION' if is_destination else 'UNKNOWN'}")
    print(f"   [DEBUG] Rule type: {'COUNTRY REGION' if is_country_region else 'POSTAL CODE ZONE' if is_postal_code_zone else 'UNKNOWN'}")
    
    if is_origin:
        country_col_variations = ['Origin Country', 'origin country', 'OriginCountry', 'ORIGIN_COUNTRY']
        postal_col_variations = ['Origin Postal Code', 'origin postal code', 'OriginPostalCode', 'ORIGIN_POSTAL_CODE']
    elif is_destination:
        country_col_variations = ['Destination Country', 'destination country', 'DestinationCountry', 'DESTINATION_COUNTRY']
        postal_col_variations = ['Destination Postal Code', 'destination postal code', 'DestinationPostalCode', 'DESTINATION_POSTAL_CODE']
    else:
        return None, [], f"Cannot determine origin/destination from column '{business_rule_col}'", failure_details
    
    # Get actual country and postal from ETOF row
    actual_country = None
    country_col_found = None
    for col_var in country_col_variations:
        if col_var in etof_row.index:
            actual_country = etof_row.get(col_var)
            country_col_found = col_var
            break
    
    actual_postal = None
    postal_col_found = None
    # Only look for postal if not a country-only region rule
    if not is_country_region:
        for col_var in postal_col_variations:
            if col_var in etof_row.index:
                actual_postal = etof_row.get(col_var)
                postal_col_found = col_var
                break
    
    failure_details['country_col'] = country_col_found
    failure_details['postal_col'] = postal_col_found
    failure_details['actual_country'] = actual_country
    failure_details['actual_postal'] = actual_postal
    failure_details['is_country_region'] = is_country_region
    
    print(f"   [DEBUG] ETOF country ({country_col_found}): '{actual_country}'")
    if not is_country_region:
        print(f"   [DEBUG] ETOF postal ({postal_col_found}): '{actual_postal}'")
    
    # Try each rule and see if the ETOF row matches
    # Track failures with same country (these are most relevant for discrepancy reporting)
    same_country_failures = []
    
    for rule_name in rules_for_column:
        print(f"\n   [DEBUG] Checking rule: '{rule_name}'")
        
        expected_country = rule_to_country.get(rule_name)
        expected_postal_codes = rule_to_postal.get(rule_name, [])
        
        print(f"   [DEBUG]   Expected country: {expected_country}")
        print(f"   [DEBUG]   Expected postal codes: {expected_postal_codes}")
        
        # Check country match
        # Handle multiple countries (comma-separated) like "BE, FR, LU, MC, NL"
        country_match = False
        country_failure_reason = None
        if expected_country and actual_country and not pd.isna(actual_country):
            actual_country_norm = str(actual_country).strip().upper()
            # Split expected countries by comma and normalize each
            expected_countries_list = [c.strip().upper() for c in str(expected_country).split(',') if c.strip()]
            country_match = actual_country_norm in expected_countries_list
            if not country_match:
                country_failure_reason = f"Country mismatch: '{actual_country_norm}' not in {expected_countries_list}"
            print(f"   [DEBUG]   Country match: {country_match} ('{actual_country_norm}' in {expected_countries_list})")
        elif not expected_country:
            country_match = True  # No country requirement
            print(f"   [DEBUG]   Country match: True (no country requirement)")
        else:
            country_failure_reason = "Country missing in ETOF"
            print(f"   [DEBUG]   Country match: False (missing actual country)")
        
        if not country_match:
            failure_details['failed_rules'].append({
                'rule_name': rule_name,
                'failure_type': 'country',
                'reason': country_failure_reason,
                'expected_country': expected_country,
                'expected_postal': expected_postal_codes
            })
            continue
        
        # Check postal code match (skip for Country Region rules - they only validate country)
        postal_match = False
        postal_failure_reason = None
        if is_country_region:
            # Country Region rules only validate country, not postal code
            postal_match = True
            print(f"   [DEBUG]   Postal match: True (Country Region rule - postal validation skipped)")
        elif expected_postal_codes and actual_postal and not pd.isna(actual_postal):
            actual_postal_norm = str(actual_postal).strip().lower()
            for expected_pc in expected_postal_codes:
                expected_pc_norm = str(expected_pc).strip().lower()
                if actual_postal_norm.startswith(expected_pc_norm):
                    postal_match = True
                    print(f"   [DEBUG]   Postal match: True ('{actual_postal_norm}' starts with '{expected_pc_norm}')")
                    break
            if not postal_match:
                postal_failure_reason = f"Postal '{actual_postal}' doesn't start with any of {expected_postal_codes}"
                print(f"   [DEBUG]   Postal match: False ('{actual_postal_norm}' doesn't start with any of {expected_postal_codes})")
        elif not expected_postal_codes:
            postal_match = True  # No postal requirement
            print(f"   [DEBUG]   Postal match: True (no postal requirement)")
        else:
            postal_failure_reason = "Postal code missing in ETOF"
            print(f"   [DEBUG]   Postal match: False (missing actual postal)")
        
        if country_match and postal_match:
            # Found a matching rule!
            validated_columns = [business_rule_col]
            if country_col_found:
                validated_columns.append(country_col_found)
            # Only add postal column for postal code zone rules, not country region rules
            if postal_col_found and not is_country_region:
                validated_columns.append(postal_col_found)
            
            print(f"\n   [DEBUG] ✓✓✓ FOUND MATCHING RULE: '{rule_name}' ✓✓✓")
            print(f"   [DEBUG] Validated columns: {validated_columns}")
            
            if is_country_region:
                return rule_name, validated_columns, f"Rule '{rule_name}' matches (Country={expected_country})", failure_details
            else:
                return rule_name, validated_columns, f"Rule '{rule_name}' matches (Country={expected_country}, Postal={expected_postal_codes})", failure_details
        
        # Country matched but postal didn't - this is the most relevant failure
        if country_match and not postal_match:
            same_country_failures.append({
                'rule_name': rule_name,
                'failure_type': 'postal',
                'reason': postal_failure_reason,
                'expected_country': expected_country,
                'expected_postal': expected_postal_codes
            })
            failure_details['failed_rules'].append({
                'rule_name': rule_name,
                'failure_type': 'postal',
                'reason': postal_failure_reason,
                'expected_country': expected_country,
                'expected_postal': expected_postal_codes
            })
    
    # Store same-country failures separately (these are most relevant)
    failure_details['same_country_failures'] = same_country_failures
    
    print(f"\n   [DEBUG] ✗✗✗ NO MATCHING RULE FOUND ✗✗✗")
    print(f"   [DEBUG] Same-country failures: {len(same_country_failures)}")
    for f in same_country_failures:
        print(f"   [DEBUG]   - Rule '{f['rule_name']}': {f['reason']}")
    
    return None, [], f"No matching rule found for column '{business_rule_col}'", failure_details


def validate_business_rule(etof_row, business_rule_col, rule_value, business_rules_lookup):
    """
    Validate if a business rule is correctly applied based on country and postal code.
    
    Args:
        etof_row: Row from ETOF/LC dataframe
        business_rule_col: Name of the business rule column (e.g., 'Origin City')
        rule_value: Value in the business rule column (e.g., 'Zhengzhou')
        business_rules_lookup: Business rules lookup dictionary
    
    Returns:
        tuple: (is_valid, validated_columns, message)
            - is_valid: True if rule is correctly applied
            - validated_columns: List of columns that were validated by this rule
            - message: Explanation message
    """
    print(f"\n   {'='*60}")
    print(f"   [BUSINESS RULE VALIDATION] Starting validation...")
    print(f"   {'='*60}")
    print(f"   [DEBUG] Input parameters:")
    print(f"      - business_rule_col: '{business_rule_col}'")
    print(f"      - rule_value: '{rule_value}'")
    print(f"      - business_rules_lookup keys: {list(business_rules_lookup.keys()) if business_rules_lookup else 'None'}")
    
    if not business_rules_lookup:
        print(f"   [DEBUG] EARLY EXIT: No lookup provided")
        return False, [], "No business rules lookup", {}
    
    # If rule_value is NaN/empty, try to find a matching rule by checking geo columns
    if rule_value is None or pd.isna(rule_value) or str(rule_value).strip().lower() in ['', 'nan', 'none']:
        print(f"   [DEBUG] Rule value is NaN/empty, searching for matching rule by geo...")
        found_rule, validated_cols, message, failure_details = find_matching_business_rule_by_geo(etof_row, business_rule_col, business_rules_lookup)
        return found_rule is not None, validated_cols, message, failure_details
    
    rule_value_str = str(rule_value).strip()
    rule_to_country = business_rules_lookup.get('rule_to_country', {})
    rule_to_postal = business_rules_lookup.get('rule_to_postal_codes', {})
    
    print(f"   [DEBUG] Available rules with country: {list(rule_to_country.keys())[:10]}{'...' if len(rule_to_country) > 10 else ''}")
    print(f"   [DEBUG] Available rules with postal codes: {list(rule_to_postal.keys())[:10]}{'...' if len(rule_to_postal) > 10 else ''}")
    
    # Try to find the rule (exact match or with suffix like "(Origin)")
    rule_name = None
    print(f"\n   [DEBUG] Searching for rule matching value '{rule_value_str}'...")
    
    for name in rule_to_country.keys():
        name_normalized = str(name).strip().lower()
        rule_value_normalized = rule_value_str.lower()
        
        # Check exact match or if rule_value is contained in the name
        if name_normalized == rule_value_normalized or rule_value_normalized in name_normalized:
            rule_name = name
            print(f"   [DEBUG] FOUND MATCH in rule_to_country: '{name}' matches '{rule_value_str}'")
            break
    
    if not rule_name:
        print(f"   [DEBUG] No match found in rule_to_country, trying rule_to_postal...")
        # Try in postal codes keys as well
        for name in rule_to_postal.keys():
            name_normalized = str(name).strip().lower()
            rule_value_normalized = rule_value_str.lower()
            if name_normalized == rule_value_normalized or rule_value_normalized in name_normalized:
                rule_name = name
                print(f"   [DEBUG] FOUND MATCH in rule_to_postal: '{name}' matches '{rule_value_str}'")
                break
    
    if not rule_name:
        print(f"   [DEBUG] NO RULE FOUND for value '{rule_value_str}'")
        return False, [], f"No business rule found for '{rule_value_str}'"
    
    # Get expected country and postal codes from the rule
    expected_country = rule_to_country.get(rule_name)
    expected_postal_codes = rule_to_postal.get(rule_name, [])
    
    print(f"\n   [BUSINESS RULE] Found rule '{rule_name}' for value '{rule_value_str}'")
    print(f"   [BUSINESS RULE]   Expected country: {expected_country}")
    print(f"   [BUSINESS RULE]   Expected postal codes: {expected_postal_codes}")
    
    # Determine if this is origin or destination based on column name
    col_lower = str(business_rule_col).lower()
    is_origin = 'origin' in col_lower or 'ship' in col_lower or 'from' in col_lower
    is_destination = 'destination' in col_lower or 'cust' in col_lower or 'to' in col_lower
    
    # Determine if this is a Country Region rule (only validates country) or Postal Code Zone (validates country + postal)
    is_country_region = 'country' in col_lower and 'region' in col_lower
    is_postal_code_zone = 'postal' in col_lower or 'zone' in col_lower
    
    print(f"\n   [DEBUG] Column classification:")
    print(f"      - Column name (lower): '{col_lower}'")
    print(f"      - is_origin: {is_origin}")
    print(f"      - is_destination: {is_destination}")
    print(f"      - is_country_region: {is_country_region}")
    print(f"      - is_postal_code_zone: {is_postal_code_zone}")
    
    # Find the corresponding country and postal code columns
    validated_columns = [business_rule_col]
    validation_passed = True
    
    if is_origin:
        country_col_variations = ['Origin Country', 'origin country', 'OriginCountry', 'ORIGIN_COUNTRY', 
                                   'Ship Country', 'SHIP_COUNTRY']
        postal_col_variations = ['Origin Postal Code', 'origin postal code', 'OriginPostalCode', 
                                 'ORIGIN_POSTAL_CODE', 'Ship Postal', 'SHIP_POST']
        print(f"   [DEBUG] Using ORIGIN column variations")
    elif is_destination:
        country_col_variations = ['Destination Country', 'destination country', 'DestinationCountry', 
                                  'DESTINATION_COUNTRY', 'Cust Country', 'CUST_COUNTRY']
        postal_col_variations = ['Destination Postal Code', 'destination postal code', 'DestinationPostalCode',
                                 'DESTINATION_POSTAL_CODE', 'Cust Postal', 'CUST_POST']
        print(f"   [DEBUG] Using DESTINATION column variations")
    else:
        print(f"   [DEBUG] CANNOT DETERMINE origin/destination from column '{business_rule_col}'")
        return False, [], f"Cannot determine origin/destination from column '{business_rule_col}'", {}
    
    print(f"   [DEBUG] Country column variations to search: {country_col_variations}")
    print(f"   [DEBUG] Postal column variations to search: {postal_col_variations}")
    print(f"   [DEBUG] Available columns in ETOF row: {list(etof_row.index)[:20]}{'...' if len(etof_row.index) > 20 else ''}")
    
    # Find actual country column in the row
    actual_country = None
    country_col_found = None
    for col_var in country_col_variations:
        if col_var in etof_row.index:
            actual_country = etof_row.get(col_var)
            country_col_found = col_var
            print(f"   [DEBUG] Found country column '{col_var}' with value '{actual_country}'")
            break
    
    if not country_col_found:
        print(f"   [DEBUG] NO COUNTRY COLUMN FOUND in ETOF row!")
    
    # Find actual postal code column in the row (only for Postal Code Zone rules, NOT Country Region)
    actual_postal = None
    postal_col_found = None
    if not is_country_region:
        for col_var in postal_col_variations:
            if col_var in etof_row.index:
                actual_postal = etof_row.get(col_var)
                postal_col_found = col_var
                print(f"   [DEBUG] Found postal column '{col_var}' with value '{actual_postal}'")
                break
        
        if not postal_col_found:
            print(f"   [DEBUG] NO POSTAL COLUMN FOUND in ETOF row!")
    else:
        print(f"   [DEBUG] Country Region rule - postal column lookup SKIPPED")
    
    print(f"\n   [BUSINESS RULE] VALIDATION SUMMARY:")
    print(f"   [BUSINESS RULE]   Rule name: '{rule_name}'")
    print(f"   [BUSINESS RULE]   Rule type: {'COUNTRY REGION' if is_country_region else 'POSTAL CODE ZONE'}")
    print(f"   [BUSINESS RULE]   Actual country ({country_col_found}): '{actual_country}'")
    if not is_country_region:
        print(f"   [BUSINESS RULE]   Actual postal ({postal_col_found}): '{actual_postal}'")
        print(f"   [BUSINESS RULE]   Expected postal codes: {expected_postal_codes}")
    print(f"   [BUSINESS RULE]   Expected country: '{expected_country}'")
    
    # Validate country
    # Handle multiple countries (comma-separated) like "BE, FR, LU, MC, NL"
    print(f"\n   [DEBUG] VALIDATING COUNTRY...")
    if expected_country:
        if actual_country is not None and not pd.isna(actual_country):
            actual_country_norm = str(actual_country).strip().upper()
            # Split expected countries by comma and normalize each
            expected_countries_list = [c.strip().upper() for c in str(expected_country).split(',') if c.strip()]
            print(f"   [DEBUG]   Comparing: '{actual_country_norm}' in {expected_countries_list}")
            if actual_country_norm in expected_countries_list:
                if country_col_found:
                    validated_columns.append(country_col_found)
                print(f"   [BUSINESS RULE]   ✓ Country MATCH: {actual_country_norm} in {expected_countries_list}")
            else:
                validation_passed = False
                print(f"   [BUSINESS RULE]   ✗ Country MISMATCH: {actual_country_norm} not in {expected_countries_list}")
        else:
            validation_passed = False
            print(f"   [BUSINESS RULE]   ✗ Country MISSING in ETOF/LC (actual_country={actual_country}, isna={pd.isna(actual_country) if actual_country is not None else 'N/A'})")
    else:
        print(f"   [DEBUG]   No expected country to validate (skipping)")
    
    # Validate postal code (if expected) - skip for Country Region rules
    print(f"\n   [DEBUG] VALIDATING POSTAL CODE...")
    if is_country_region:
        print(f"   [DEBUG]   Country Region rule - postal validation SKIPPED (only country matters)")
    elif expected_postal_codes:
        if actual_postal is not None and not pd.isna(actual_postal):
            actual_postal_norm = str(actual_postal).strip().lower()
            print(f"   [DEBUG]   Actual postal (normalized): '{actual_postal_norm}'")
            # Check if actual postal starts with or matches any expected postal code
            postal_match = False
            for expected_pc in expected_postal_codes:
                expected_pc_norm = str(expected_pc).strip().lower()
                print(f"   [DEBUG]   Checking if '{actual_postal_norm}' starts with or equals '{expected_pc_norm}'...")
                if actual_postal_norm.startswith(expected_pc_norm) or actual_postal_norm == expected_pc_norm:
                    postal_match = True
                    print(f"   [DEBUG]   MATCH FOUND!")
                    break
            
            if postal_match:
                if postal_col_found:
                    validated_columns.append(postal_col_found)
                print(f"   [BUSINESS RULE]   ✓ Postal MATCH: '{actual_postal_norm}' matches one of {expected_postal_codes}")
            else:
                validation_passed = False
                print(f"   [BUSINESS RULE]   ✗ Postal MISMATCH: '{actual_postal_norm}' not in {expected_postal_codes}")
        else:
            # Postal code is missing but expected - might be optional
            print(f"   [BUSINESS RULE]   ⚠ Postal MISSING in ETOF/LC (actual_postal={actual_postal}, isna={pd.isna(actual_postal) if actual_postal is not None else 'N/A'})")
            print(f"   [DEBUG]   Postal code validation skipped (may be optional)")
    else:
        print(f"   [DEBUG]   No expected postal codes to validate (skipping)")
    
    print(f"\n   {'='*60}")
    print(f"   [BUSINESS RULE] FINAL RESULT:")
    print(f"   {'='*60}")
    print(f"   [DEBUG]   validation_passed: {validation_passed}")
    print(f"   [DEBUG]   validated_columns: {validated_columns}")
    
    if validation_passed:
        if is_country_region:
            msg = f"Business rule '{rule_name}' validated: Country={expected_country} (Country Region rule)"
        else:
            msg = f"Business rule '{rule_name}' validated: Country={expected_country}, Postal={expected_postal_codes}"
        print(f"   [BUSINESS RULE]   ✓✓✓ VALIDATION PASSED ✓✓✓")
        print(f"   [BUSINESS RULE]   Columns to skip from discrepancy checking: {validated_columns}")
        print(f"   {'='*60}\n")
        return True, validated_columns, msg, {}
    else:
        print(f"   [BUSINESS RULE]   ✗✗✗ VALIDATION FAILED ✗✗✗")
        print(f"   {'='*60}\n")
        # Build failure details for when rule was specified but didn't match
        failure_details = {
            'country_col': country_col_found,
            'postal_col': postal_col_found,
            'actual_country': actual_country,
            'actual_postal': actual_postal,
            'failed_rules': [{
                'rule_name': rule_name,
                'failure_type': 'validation',
                'expected_country': expected_country,
                'expected_postal': expected_postal_codes
            }]
        }
        return False, [], f"Business rule '{rule_name}' validation failed", failure_details


# Note: extract_country_code is already applied in part1_etof_file_processing.py
# No need to duplicate it here as vocabular.py uses processed dataframes


def load_conditions():
    """Load conditional rules from Filtered_Rate_Card_with_Conditions.xlsx."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    conditions_file = os.path.join(script_dir, "Filtered_Rate_Card_with_Conditions.xlsx")
    
    if not os.path.exists(conditions_file):
        print(f"Warning: {conditions_file} not found. Conditions will not be validated.")
        return {}
    
    try:
        df_conditions = pd.read_excel(conditions_file, sheet_name='Conditions')
        
        # Parse conditions into a dictionary: {column_name: [list of condition rules]}
        conditions_dict = {}
        current_column = None
        
        for _, row in df_conditions.iterrows():
            column = row.get('Column', '')
            condition_rule = row.get('Condition Rule', '')
            
            if pd.notna(column) and str(column).strip() and str(column).strip() != 'nan':
                current_column = str(column).strip()
                if current_column not in conditions_dict:
                    conditions_dict[current_column] = []
            
            if pd.notna(condition_rule) and str(condition_rule).strip() and current_column:
                condition_text = str(condition_rule).strip()
                # Skip header lines like "Conditional rules:"
                if condition_text.lower() not in ['conditional rules:', 'conditional rules']:
                    conditions_dict[current_column].append(condition_text)
        
        print(f"\nLoaded conditions for {len(conditions_dict)} columns")
        return conditions_dict
    except Exception as e:
        print(f"Warning: Could not load conditions: {e}")
        return {}


def parse_condition(condition_text, rate_card_value):
    """Parse a condition rule and extract the value it applies to.
    
    Example: "NAC: RATE_TYPE is empty in any item and does not contain FAK in any item"
    Returns: ('NAC', condition_logic)
    """
    if not condition_text or pd.isna(condition_text):
        return None, None
    
    condition_text = str(condition_text).strip()
    
    # Check if condition starts with a value followed by colon (e.g., "NAC: ...")
    if ':' in condition_text:
        parts = condition_text.split(':', 1)
        condition_value = parts[0].strip()
        condition_logic = parts[1].strip() if len(parts) > 1 else ''
        
        return condition_value, condition_logic
    
    return None, condition_text


def value_satisfies_condition(resmed_value, rate_card_value, condition_text, debug=False):
    """Check if a ResMed value satisfies the condition for a given rate card value.
    
    Args:
        resmed_value: The value from ResMed dataframe
        rate_card_value: The value from Rate Card (e.g., 'NAC')
        condition_text: The condition rule text
        debug: Whether to print debug information
    
    Returns:
        True if the value satisfies the condition, False otherwise
    
    Example:
        condition_text = "NAC: RATE_TYPE is empty in any item and does not contain FAK in any item"
        rate_card_value = "NAC"
        resmed_value = nan (empty)
        Returns: True (because empty satisfies "is empty")
    """
    if debug:
        print(f"            [CONDITION EVAL] Evaluating condition...")
        print(f"               - Condition text: {condition_text[:100] if condition_text else 'None'}...")
        print(f"               - Rate Card value: '{rate_card_value}'")
        print(f"               - Shipment value: '{resmed_value}'")
    
    if not condition_text or pd.isna(condition_text):
        if debug:
            print(f"               - Result: FALSE (condition is empty/None)")
        return False
    
    condition_text = str(condition_text).strip()
    condition_lower = condition_text.lower()
    rate_card_val_str = str(rate_card_value).lower() if pd.notna(rate_card_value) else ''
    
    # Check if condition is for this rate card value (format: "1. NAC: ..." or "NAC: ...")
    if ':' in condition_text:
        # Handle numbered conditions like "1. NAC:" or "1.NAC:" or just "NAC:"
        # Remove leading number and dot if present (e.g., "1. " or "1.")
        condition_text_cleaned = re.sub(r'^\d+\.\s*', '', condition_text)
        condition_parts = condition_text_cleaned.split(':', 1)
        condition_value = condition_parts[0].strip()
        condition_logic = condition_parts[1].strip() if len(condition_parts) > 1 else ''
        
        if debug:
            print(f"               - Condition value: '{condition_value}', Logic: '{condition_logic[:50]}...'")
        
        # Check if this condition applies to the rate card value
        if rate_card_val_str and condition_value.lower() != rate_card_val_str:
            if debug:
                print(f"               - Result: FALSE (condition value '{condition_value}' != rate card value '{rate_card_val_str}')")
            return False
        
        condition_text = condition_logic  # Use only the logic part
        condition_lower = condition_text.lower()
    
    # Check if ResMed value is empty/NaN
    is_empty = pd.isna(resmed_value) or str(resmed_value).strip() == '' or str(resmed_value).lower() in ['nan', 'none', 'null', '']
    resmed_val_str = str(resmed_value).lower() if pd.notna(resmed_value) else ''
    
    if debug:
        print(f"               - Shipment value is empty: {is_empty}")
        print(f"               - Shipment value (normalized): '{resmed_val_str}'")
    
    # Parse condition logic
    # Example: "RATE_TYPE is empty in any item and does not contain FAK in any item"
    
    # Check "is empty" condition
    if 'is empty' in condition_lower or 'is empty in any item' in condition_lower:
        if debug:
            print(f"               - Checking 'is empty' condition...")
        if is_empty:
            # Value is empty - check if there are additional conditions
            # If condition has "and does not contain", empty values satisfy this (empty doesn't contain anything)
            if 'does not contain' in condition_lower or 'and' in condition_lower:
                # For "and" conditions, all must be satisfied
                # Empty value satisfies "is empty" and "does not contain X" (empty doesn't contain anything)
                if debug:
                    print(f"               - Result: TRUE (value is empty AND 'does not contain' also satisfied)")
                return True
            if debug:
                print(f"               - Result: TRUE (value is empty)")
            return True
        else:
            if debug:
                print(f"               - 'is empty' check: FAILED (value is not empty)")
    
    # Check "does not contain" condition
    if 'does not contain' in condition_lower:
        if debug:
            print(f"               - Checking 'does not contain' condition...")
        if is_empty:
            if debug:
                print(f"               - Result: TRUE (empty values don't contain anything)")
            return True  # Empty values don't contain anything
        
        # Extract what it should not contain
        parts = condition_lower.split('does not contain')
        if len(parts) > 1:
            forbidden_part = parts[1].split('in any item')[0].strip()
            # Handle comma-separated values (e.g., "EY,ETIHAD,ETIHAD AIRWAYS")
            forbidden_values = [v.strip() for v in forbidden_part.split(',')]
            if debug:
                print(f"               - Forbidden values: {forbidden_values}")
            # Check if ResMed value contains any forbidden value
            for forbidden in forbidden_values:
                if forbidden and forbidden in resmed_val_str:
                    if debug:
                        print(f"               - Result: FALSE (value contains forbidden '{forbidden}')")
                    return False  # Contains forbidden value - condition not satisfied
            if debug:
                print(f"               - Result: TRUE (value does not contain any forbidden values)")
            return True  # Doesn't contain any forbidden value
    
    # Check "does not equal" condition
    if 'does not equal' in condition_lower or 'does not equal to' in condition_lower:
        if debug:
            print(f"               - Checking 'does not equal' condition...")
        if is_empty:
            if debug:
                print(f"               - Result: TRUE (empty values don't equal anything)")
            return True  # Empty values don't equal anything
        
        parts = condition_lower.split('does not equal')
        if len(parts) > 1:
            forbidden_part = parts[1].split('in any item')[0].strip()
            # Handle comma-separated values
            forbidden_values = [v.strip() for v in forbidden_part.split(',')]
            if debug:
                print(f"               - Forbidden values: {forbidden_values}")
            # Check if ResMed value equals any forbidden value
            for forbidden in forbidden_values:
                if forbidden and resmed_val_str == forbidden:
                    if debug:
                        print(f"               - Result: FALSE (value equals forbidden '{forbidden}')")
                    return False  # Equals forbidden value - condition not satisfied
            if debug:
                print(f"               - Result: TRUE (value does not equal any forbidden values)")
            return True  # Doesn't equal any forbidden value
    
    # Check "contains" condition (positive match)
    if 'contains' in condition_lower and 'does not contain' not in condition_lower:
        if debug:
            print(f"               - Checking 'contains' condition...")
        if is_empty:
            if debug:
                print(f"               - Result: FALSE (empty values don't contain anything)")
            return False  # Empty values don't contain anything
        
        parts = condition_lower.split('contains')
        if len(parts) > 1:
            required_part = parts[1].split('in any item')[0].strip()
            # Handle comma-separated values
            required_values = [v.strip() for v in required_part.split(',')]
            if debug:
                print(f"               - Required values: {required_values}")
            # Check if ResMed value contains any required value
            for required in required_values:
                if required and required in resmed_val_str:
                    if debug:
                        print(f"               - Result: TRUE (value contains required '{required}')")
                    return True  # Contains required value
            if debug:
                print(f"               - Result: FALSE (value does not contain any required values)")
            return False  # Doesn't contain any required value
    
    # Check "equals" or "equal to" condition
    if 'equal to' in condition_lower or ('equals' in condition_lower and 'does not equal' not in condition_lower):
        if debug:
            print(f"               - Checking 'equals/equal to' condition...")
        if is_empty:
            if debug:
                print(f"               - Result: FALSE (empty values don't equal anything)")
            return False  # Empty values don't equal anything
        
        if 'equal to' in condition_lower:
            parts = condition_lower.split('equal to')
        else:
            parts = condition_lower.split('equals')
        if len(parts) > 1:
            required_part = parts[1].split('in any item')[0].strip()
            # Handle comma-separated values
            required_values = [v.strip() for v in required_part.split(',')]
            if debug:
                print(f"               - Required values: {required_values}")
            # Check if ResMed value equals any required value
            for required in required_values:
                if required and resmed_val_str == required:
                    if debug:
                        print(f"               - Result: TRUE (value equals required '{required}')")
                    return True  # Equals required value
            if debug:
                print(f"               - Result: FALSE (value does not equal any required values)")
            return False  # Doesn't equal any required value
    
    if debug:
        print(f"               - Result: FALSE (no matching condition type found)")
    return False


def find_condition_for_value(rate_card_value, column_name, conditions_dict):
    """Find the condition text for a rate card value (even if not satisfied).
    
    This helps show what codes the user should enter when there's a discrepancy.
    For example, if rate card has "Long Beach" and condition says "Long Beach: equals LGB",
    we return the condition so we can tell the user to enter "LGB" instead.
    
    Returns:
        str or None: The condition text, or None if not found
    """
    # Try to find column in conditions_dict (case-insensitive)
    column_key = None
    for key in conditions_dict.keys():
        if normalize_column_name(key) == normalize_column_name(column_name):
            column_key = key
            break
    
    if column_key is None:
        return None
    
    conditions = conditions_dict[column_key]
    rate_card_val_str = str(rate_card_value).lower() if pd.notna(rate_card_value) else ''
    
    if not rate_card_val_str:
        return None
    
    # Handle both string and list formats for conditions
    if isinstance(conditions, str):
        conditions_list = [line.strip() for line in conditions.split('\n') if line.strip()]
    elif isinstance(conditions, list):
        conditions_list = conditions
    else:
        conditions_list = [str(conditions)]
    
    for condition_text in conditions_list:
        condition_lower = str(condition_text).lower()
        
        # Skip header lines
        if 'conditional rules' in condition_lower and ':' not in condition_text:
            continue
        
        # Check if condition is for this rate card value
        # Pattern: (optional number + dot + space) + rate_card_value + colon
        pattern = rf'(?:\d+\.\s*)?{re.escape(rate_card_val_str)}:'
        if re.search(pattern, condition_lower):
            return condition_text
    
    return None


def check_value_against_conditions(resmed_value, rate_card_value, column_name, conditions_dict, debug=False):
    """Check if ResMed value satisfies any condition for the rate card value.
    
    Returns:
        (is_valid, matching_condition) tuple
    """
    # Try to find column in conditions_dict (case-insensitive)
    column_key = None
    for key in conditions_dict.keys():
        if normalize_column_name(key) == normalize_column_name(column_name):
            column_key = key
            break
    
    if column_key is None:
        if debug:
            print(f"      [CONDITION DEBUG] Column '{column_name}' not found in conditions_dict")
        return False, None
    
    conditions = conditions_dict[column_key]
    rate_card_val_str = str(rate_card_value).lower() if pd.notna(rate_card_value) else ''
    
    if debug:
        print(f"      [CONDITION DEBUG] Checking conditions for column '{column_name}':")
        print(f"         - Shipment value: '{resmed_value}'")
        print(f"         - Rate Card value: '{rate_card_value}'")
        print(f"         - Conditions found: {conditions}")
    
    # Handle both string and list formats for conditions
    # rate_card_processing.py returns conditions as a string (from cell comments)
    if isinstance(conditions, str):
        # Split string by newlines to get individual condition lines
        conditions_list = [line.strip() for line in conditions.split('\n') if line.strip()]
    elif isinstance(conditions, list):
        conditions_list = conditions
    else:
        # If it's neither string nor list, try to convert
        conditions_list = [str(conditions)]
    
    if debug:
        print(f"         - Parsed conditions list ({len(conditions_list)} items): {conditions_list[:3]}{'...' if len(conditions_list) > 3 else ''}")
    
    for condition_text in conditions_list:
        # Check if this condition applies to the rate card value
        # Format: "1. NAC: RATE_TYPE is empty..." or "NAC: RATE_TYPE is empty..."
        condition_lower = str(condition_text).lower()
        
        # Skip header lines
        if 'conditional rules' in condition_lower and ':' not in condition_text:
            continue
        
        # Check if condition is for this rate card value
        # Look for pattern like "NAC:" or "1. NAC:" or "1.NAC:"
        if rate_card_val_str:
            # Handle numbered conditions like "1. NAC:" or "1.NAC:"
            # Pattern: (optional number + dot + space) + rate_card_value + colon
            # Match patterns like "1. nac:", "1.nac:", "nac:"
            pattern = rf'(?:\d+\.\s*)?{re.escape(rate_card_val_str)}:'
            if re.search(pattern, condition_lower):
                # This condition applies to this rate card value
                if debug:
                    print(f"         - Condition matches rate card value '{rate_card_val_str}': {condition_text[:80]}...")
                is_valid = value_satisfies_condition(resmed_value, rate_card_value, condition_text, debug=debug)
                if debug:
                    print(f"         - Condition satisfied: {is_valid}")
                if is_valid:
                    return True, condition_text
    
    if debug:
        print(f"         - No matching condition found for rate card value '{rate_card_val_str}'")
    return False, None


def load_standardized_dataframes():
    """Load standardized dataframes from shipments.py output."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, "Standardized_Data.xlsx")
    
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found.")
        print("Please run shipments.py first to generate standardized data.")
        return None, None
    
    try:
        df_resmed = pd.read_excel(excel_file, sheet_name='ResMed (Standardized)')
        df_rate_card = pd.read_excel(excel_file, sheet_name='Rate Card (Standardized)')
        
        print(f"Loaded ResMed (Standardized): {df_resmed.shape}")
        print(f"Loaded Rate Card (Standardized): {df_rate_card.shape}")
        
        return df_resmed, df_rate_card
    except Exception as e:
        print(f"Error loading standardized data: {e}")
        return None, None


def find_common_columns(df_resmed, df_rate_card):
    """Find common columns between the two dataframes."""
    resmed_cols = set(df_resmed.columns)
    rate_card_cols = set(df_rate_card.columns)
    
    # Debug: Show all columns being compared
    print(f"\n[DEBUG] Columns comparison:")
    print(f"   ETOF/LC columns ({len(resmed_cols)}):")
    for col in sorted(resmed_cols):
        print(f"      - '{col}'")
    
    print(f"\n   Rate Card columns ({len(rate_card_cols)}):")
    for col in sorted(rate_card_cols):
        print(f"      - '{col}'")
    
    # Find common columns (exact match)
    common_cols = sorted(list(resmed_cols & rate_card_cols))
    
    print(f"\nFound {len(common_cols)} common columns for matching (exact match):")
    for col in common_cols:
        print(f"  - {col}")
    
    # If no common columns found, try normalized matching
    if not common_cols:
        print(f"\n[DEBUG] No exact matches found. Trying normalized column matching...")
        
        # Normalize column names for comparison
        resmed_normalized = {normalize_column_name(col): col for col in resmed_cols}
        rate_card_normalized = {normalize_column_name(col): col for col in rate_card_cols}
        
        print(f"\n   ETOF/LC normalized columns:")
        for norm, orig in sorted(resmed_normalized.items()):
            print(f"      '{norm}' <- '{orig}'")
        
        print(f"\n   Rate Card normalized columns:")
        for norm, orig in sorted(rate_card_normalized.items()):
            print(f"      '{norm}' <- '{orig}'")
        
        # Find common normalized columns
        common_normalized = set(resmed_normalized.keys()) & set(rate_card_normalized.keys())
        
        print(f"\n   Common normalized columns ({len(common_normalized)}):")
        for norm in sorted(common_normalized):
            etof_orig = resmed_normalized[norm]
            rc_orig = rate_card_normalized[norm]
            print(f"      '{norm}': ETOF='{etof_orig}', RC='{rc_orig}'")
        
        # Use the ETOF/LC column names for the common columns
        common_cols = sorted([resmed_normalized[norm] for norm in common_normalized])
    
    return common_cols


def analyze_discrepancy_patterns(all_discrepancies, conditions_dict=None):
    """
    Analyze discrepancies to find common patterns.
    
    Args:
        all_discrepancies: List of discrepancy dictionaries, each with 'column', 'etofs_value', 'rate_card_value'
        conditions_dict: Dictionary of conditional rules (for extracting codes from conditions)
    
    Returns:
        tuple: (has_common_pattern, pattern_comment, minor_discrepancies)
            - has_common_pattern: True if there's a common pattern, False if all different
            - pattern_comment: The summarized comment based on the pattern
            - minor_discrepancies: List of discrepancies for "Also" section (or empty list)
    """
    if not all_discrepancies:
        return False, "Please recheck the shipment details", []
    
    # Group discrepancies by column name
    column_counts = {}
    column_discrepancies = {}
    
    for disc in all_discrepancies:
        col = disc.get('column', 'Unknown')
        if col not in column_counts:
            column_counts[col] = 0
            column_discrepancies[col] = []
        column_counts[col] += 1
        column_discrepancies[col].append(disc)
    
    total_discrepancies = len(all_discrepancies)
    unique_columns = len(column_counts)
    
    # If all discrepancies are for the same column - clear pattern
    if unique_columns == 1:
        column_name = list(column_counts.keys())[0]
        return True, f"{column_name}: Shipment value needs to be changed", []
    
    # Check if one column dominates (has majority of discrepancies, at least 70%)
    for col, count in column_counts.items():
        if count / total_discrepancies >= 0.7:
            # Collect minor discrepancies (from other columns)
            minor_discs = []
            for other_col, discs in column_discrepancies.items():
                if other_col != col:
                    # Get unique discrepancies for this column (first one as representative)
                    if discs:
                        minor_discs.append(discs[0])
            return True, f"{col}: Shipment value needs to be changed", minor_discs
    
    # Check if a few columns (2-3) cover most discrepancies (80%+)
    sorted_columns = sorted(column_counts.items(), key=lambda x: x[1], reverse=True)
    top_columns = []
    covered_count = 0
    
    for col, count in sorted_columns[:3]:  # Check top 3 columns
        top_columns.append(col)
        covered_count += count
        if covered_count / total_discrepancies >= 0.8:
            break
    
    if len(top_columns) <= 3 and covered_count / total_discrepancies >= 0.8:
        # Format: "Column1, Column2: Shipment values need to be changed"
        columns_str = ", ".join(top_columns)
        return True, f"{columns_str}: Shipment values need to be changed", []
    
    # No clear pattern - all different
    return False, "Please recheck the shipment details", []


def match_shipments_with_rate_card(df_etofs, df_filtered_rate_card, common_columns, conditions_dict=None, debug_conditions=True, rate_card_file_path=None, business_rules_lookup=None):
    """Match ResMed shipments with Rate Card entries and identify discrepancies.
    
    Args:
        df_etofs: Shipment dataframe (ETOF or LC) from vocabular.py
        df_filtered_rate_card: Rate Card standardized dataframe from rate_card_processing.py
        common_columns: List of common column names
        conditions_dict: Dictionary of conditional rules from rate_card_processing.py
        debug_conditions: Enable debug output for condition checking (default: True)
        rate_card_file_path: Path to rate card file for loading business rules
        business_rules_lookup: Pre-loaded business rules lookup (optional, will load if not provided)
    """
    print(f"\n[DEBUG] match_shipments_with_rate_card called with debug_conditions={debug_conditions}")
    print(f"[DEBUG] rate_card_file_path parameter: {rate_card_file_path}")
    print(f"[DEBUG] business_rules_lookup parameter is None: {business_rules_lookup is None}")
    
    # Load business rules if not provided
    if business_rules_lookup is None and rate_card_file_path:
        print(f"\n[DEBUG] Loading business rules from {rate_card_file_path}...")
        business_rules_lookup = load_business_rules_for_matching(rate_card_file_path)
        print(f"[DEBUG] After loading, business_rules_lookup is None: {business_rules_lookup is None}")
    elif business_rules_lookup is None:
        print(f"[DEBUG] WARNING: business_rules_lookup is None and rate_card_file_path is also None/empty!")
        print(f"[DEBUG] Business rule validation will NOT be available!")
    
    if business_rules_lookup:
        print(f"\n[DEBUG] Business Rules Available in matching function:")
        print(f"   - Rule to country mappings: {len(business_rules_lookup.get('rule_to_country', {}))}")
        print(f"   - Rule to postal codes mappings: {len(business_rules_lookup.get('rule_to_postal_codes', {}))}")
        print(f"   - Business rule columns: {business_rules_lookup.get('business_rule_columns', set())}")
        if not business_rules_lookup.get('business_rule_columns'):
            print(f"   [WARNING] business_rule_columns is EMPTY - no columns will be validated!")
    else:
        print(f"\n[DEBUG] WARNING: No business_rules_lookup available after all attempts!")
    
    if conditions_dict is None:
        conditions_dict = {}
    
    print(f"[DEBUG] conditions_dict received: {len(conditions_dict)} columns")
    
    # Debug: Print loaded conditions
    if debug_conditions:
        print("\n" + "="*80)
        print("[DEBUG] CONDITIONS LOADED FOR MATCHING")
        print("="*80)
        if conditions_dict:
            for col_name, conditions in conditions_dict.items():
                print(f"\n   Column: '{col_name}'")
                if isinstance(conditions, str):
                    cond_lines = [line.strip() for line in conditions.split('\n') if line.strip()]
                    for i, line in enumerate(cond_lines[:5]):  # Show first 5 lines
                        print(f"      {i+1}. {line[:100]}{'...' if len(line) > 100 else ''}")
                    if len(cond_lines) > 5:
                        print(f"      ... and {len(cond_lines) - 5} more conditions")
                else:
                    print(f"      {conditions}")
        else:
            print("   No conditions loaded.")
        print("="*80 + "\n")
    
    # Create a copy to preserve all original columns
    df_etofs = df_etofs.copy()
    
    # Note: Country code extraction is already done in part1_etof_file_processing.py
    # No need to apply it again here as vocabular.py uses processed dataframes
    
    # Create mappings from normalized column names back to original column names
    etofs_original_to_normalized = {col: normalize_column_name(col) for col in df_etofs.columns}
    rate_card_original_to_normalized = {col: normalize_column_name(col) for col in df_filtered_rate_card.columns}
    
    etofs_normalized_to_original = {v: k for k, v in etofs_original_to_normalized.items()}
    rate_card_normalized_to_original = {v: k for k, v in rate_card_original_to_normalized.items()}
    
    # Get normalized versions of common columns
    common_columns_normalized = [normalize_column_name(col) for col in common_columns]
    
    # Get the original column names for the common normalized columns
    common_etofs_cols_original = [etofs_normalized_to_original[col_norm] 
                                   for col_norm in common_columns_normalized 
                                   if col_norm in etofs_normalized_to_original]
    common_rate_card_cols_original = [rate_card_normalized_to_original[col_norm] 
                                      for col_norm in common_columns_normalized 
                                      if col_norm in rate_card_normalized_to_original]
    
    print(f"\nMatching based on {len(common_columns_normalized)} common attributes:")
    print(common_columns_normalized)
    
    # Pre-calculate unique normalized values from df_filtered_rate_card for efficiency
    unique_rc_orig_countries_norm = set()
    unique_rc_dest_countries_norm = set()
    unique_rc_orig_dest_combinations = set()  # Store (origin, dest) tuples
    
    print("Rate Card columns:")
    print(df_filtered_rate_card.columns)

    # Find origin country column in rate card (handle variations)
    rc_origin_col = None
    rc_origin_variations = ['Origin Country', 'origin country', 'OriginCountry', 'origincountrycode', 
                            'ORIGIN COUNTRY CODE', 'Origin Country Code', 'ORIGIN_COUNTRY', 'ORIGIN_COUNTRY_CODE']
    for col in rc_origin_variations:
        if col in df_filtered_rate_card.columns:
            rc_origin_col = col
            break
    # Also check by normalized comparison if not found by exact match
    if rc_origin_col is None:
        for col in df_filtered_rate_card.columns:
            if normalize_column_name(col) in ['origincountry', 'origincountrycode']:
                rc_origin_col = col
                break
    
    # Find destination country column in rate card (handle variations)
    rc_dest_col = None
    rc_dest_variations = ['Destination Country', 'destination country', 'DestinationCountry', 'destinationcountrycode',
                          'DESTINATION COUNTRY CODE', 'Destination Country Code', 'DESTINATION_COUNTRY', 'DESTINATION_COUNTRY_CODE']
    for col in rc_dest_variations:
        if col in df_filtered_rate_card.columns:
            rc_dest_col = col
            break
    # Also check by normalized comparison if not found by exact match
    if rc_dest_col is None:
        for col in df_filtered_rate_card.columns:
            if normalize_column_name(col) in ['destinationcountry', 'destinationcountrycode']:
                rc_dest_col = col
                break
    
    if rc_origin_col:
        print(f"   Found Rate Card Origin Country column: '{rc_origin_col}'")
        unique_rc_orig_countries_norm = set(df_filtered_rate_card[rc_origin_col].apply(normalize_value).dropna())
    if rc_dest_col:
        print(f"   Found Rate Card Destination Country column: '{rc_dest_col}'")
        unique_rc_dest_countries_norm = set(df_filtered_rate_card[rc_dest_col].apply(normalize_value).dropna())
    
    # Build country code mappings from conditions (e.g., "Singapore: equals SG" -> {"sg": "singapore"})
    # This allows matching shipment country codes (like "SG") to rate card countries (like "Singapore")
    origin_code_to_country = {}  # Maps normalized code -> normalized country name
    dest_code_to_country = {}
    
    if rc_origin_col and rc_origin_col in conditions_dict:
        origin_conditions = conditions_dict[rc_origin_col]
        print(f"   [DEBUG] Origin Country conditions found: {origin_conditions[:100]}...")
        # Parse conditions like "Singapore: equals SG" or "1. Singapore: equals SG,SGP"
        import re
        for line in str(origin_conditions).split('\n'):
            line = line.strip()
            if ':' in line and ('equals' in line.lower() or 'equal to' in line.lower()):
                # Extract country name and code(s)
                # Pattern: "1. Singapore: equals SG,SGP" or "Singapore: equals SG"
                match = re.match(r'(?:\d+\.\s*)?([^:]+):\s*(?:equals?|equal to)\s*(.+)', line, re.IGNORECASE)
                if match:
                    country_name = normalize_value(match.group(1).strip())
                    codes = [normalize_value(c.strip()) for c in match.group(2).split(',')]
                    for code in codes:
                        if code:
                            origin_code_to_country[code] = country_name
                            print(f"      Origin code mapping: '{code}' -> '{country_name}'")
    
    if rc_dest_col and rc_dest_col in conditions_dict:
        dest_conditions = conditions_dict[rc_dest_col]
        print(f"   [DEBUG] Destination Country conditions found: {dest_conditions[:100]}...")
        import re
        for line in str(dest_conditions).split('\n'):
            line = line.strip()
            if ':' in line and ('equals' in line.lower() or 'equal to' in line.lower()):
                match = re.match(r'(?:\d+\.\s*)?([^:]+):\s*(?:equals?|equal to)\s*(.+)', line, re.IGNORECASE)
                if match:
                    country_name = normalize_value(match.group(1).strip())
                    codes = [normalize_value(c.strip()) for c in match.group(2).split(',')]
                    for code in codes:
                        if code:
                            dest_code_to_country[code] = country_name
                            print(f"      Destination code mapping: '{code}' -> '{country_name}'")
    
    # Helper function to check if a shipment country value matches rate card countries
    def country_matches_rate_card(shipment_country_norm, rc_countries_set, code_to_country_map):
        """Check if shipment country matches any rate card country (directly or via conditions)."""
        if not shipment_country_norm:
            return False
        # Direct match
        if shipment_country_norm in rc_countries_set:
            return True
        # Match via condition code mapping (e.g., "sg" -> "singapore")
        if shipment_country_norm in code_to_country_map:
            mapped_country = code_to_country_map[shipment_country_norm]
            if mapped_country in rc_countries_set:
                return True
        return False
    
    # Create set of (origin, destination) combinations from rate card
    if rc_origin_col and rc_dest_col:
        for _, rc_row in df_filtered_rate_card.iterrows():
            orig = normalize_value(rc_row.get(rc_origin_col))
            dest = normalize_value(rc_row.get(rc_dest_col))
            if orig and dest:
                unique_rc_orig_dest_combinations.add((orig, dest))
    
    # Initialize a new 'comment' column in df_etofs
    #df_etofs['Comments'] = ''
    
    # Check if "Carrier agreement" column exists in rate card
    carrier_agreement_col = None
    for col in df_filtered_rate_card.columns:
        if col.lower().replace(' ', '').replace('_', '') == 'carrieragreement':
            carrier_agreement_col = col
            print(f"\n   [INFO] Found 'Carrier agreement' column in rate card: '{col}'")
            break
    
    # Initialize "Rate Card" column in df_etofs if Carrier agreement exists
    if carrier_agreement_col:
        df_etofs['Rate Card'] = None
    
    # Iterate through each row of df_etofs
    for index_etofs, row_etofs in df_etofs.iterrows():
        comments_for_current_etofs_row = []
        carrier_agreement_value = None  # Track the best match's Carrier agreement
        
        # Initialize columns validated by business rules for this row
        # This will be populated by business rule validation before country validation
        columns_validated_by_business_rules = set()
        business_rule_validation_messages = []
        
        # ===== PRE-STEP: Quick Business Rules Validation to identify columns to skip =====
        # This runs BEFORE country validation so we know which country columns to skip
        # Track which columns passed and which failed (only report failures if column has no passing rule)
        business_rule_columns_passed = set()  # Columns where at least one rule passed
        business_rule_columns_failed = {}  # Column -> list of failure messages (only used if no rule passes)
        
        if business_rules_lookup and business_rules_lookup.get('business_rule_columns'):
            business_rule_cols = business_rules_lookup.get('business_rule_columns', set())
            
            for br_col in business_rule_cols:
                # Find the column in the ETOF row (handle variations)
                br_col_value = None
                br_col_found = None
                
                # Try exact match first
                if br_col in row_etofs.index:
                    br_col_value = row_etofs.get(br_col)
                    br_col_found = br_col
                else:
                    # Try case-insensitive match
                    br_col_norm = normalize_column_name(br_col)
                    for col in row_etofs.index:
                        if normalize_column_name(col) == br_col_norm:
                            br_col_value = row_etofs.get(col)
                            br_col_found = col
                            break
                
                # Call validation whether the value is present or NaN
                is_valid, validated_cols, message, failure_details = validate_business_rule(
                    row_etofs, br_col_found if br_col_found else br_col, br_col_value, business_rules_lookup
                )
                
                br_col_name = br_col_found if br_col_found else br_col
                br_col_norm = normalize_column_name(br_col_name)
                
                if is_valid:
                    # Add validated columns to skip set (normalize for comparison)
                    for vc in validated_cols:
                        normalized_vc = normalize_column_name(vc)
                        columns_validated_by_business_rules.add(normalized_vc)
                    business_rule_validation_messages.append(f"[Business Rule Applied] {message}")
                    business_rule_columns_passed.add(br_col_norm)
                else:
                    # Business rule validation FAILED - store for later (only report if no rule passes for this column)
                    if br_col_norm not in business_rule_columns_failed:
                        business_rule_columns_failed[br_col_norm] = []
                    
                    if failure_details:
                        # Extract ALL failure details (both country mismatches and postal mismatches)
                        failed_rules = failure_details.get('failed_rules', [])
                        same_country_failures = failure_details.get('same_country_failures', [])
                        actual_country = failure_details.get('actual_country', 'N/A')
                        actual_postal = failure_details.get('actual_postal', 'N/A')
                        is_country_region = failure_details.get('is_country_region', False)
                        
                        # First, check same_country_failures (most specific - country matched but postal didn't)
                        if same_country_failures:
                            for fail in same_country_failures:
                                rule_name = fail.get('rule_name', 'Unknown')
                                expected_postal = fail.get('expected_postal', [])
                                failure_msg = f"Business rule '{rule_name}' failed: Postal code '{actual_postal}' does not start with any of {expected_postal}"
                                if failure_msg not in business_rule_columns_failed[br_col_norm]:
                                    business_rule_columns_failed[br_col_norm].append(failure_msg)
                        elif failed_rules:
                            # Check for country mismatches
                            for fail in failed_rules:
                                rule_name = fail.get('rule_name', 'Unknown')
                                failure_type = fail.get('failure_type', 'unknown')
                                expected_country = fail.get('expected_country', 'N/A')
                                expected_postal = fail.get('expected_postal', [])
                                
                                if failure_type == 'country':
                                    if is_country_region:
                                        failure_msg = f"Business rule '{rule_name}' failed: Country '{actual_country}' does not match expected '{expected_country}'"
                                    else:
                                        failure_msg = f"Business rule '{rule_name}' failed: Country '{actual_country}' does not match expected '{expected_country}' (expected postal starting with {expected_postal})"
                                elif failure_type == 'postal':
                                    failure_msg = f"Business rule '{rule_name}' failed: Postal code '{actual_postal}' does not start with any of {expected_postal}"
                                else:
                                    failure_msg = f"Business rule '{rule_name}' validation failed for column '{br_col_name}'"
                                
                                if failure_msg not in business_rule_columns_failed[br_col_norm]:
                                    business_rule_columns_failed[br_col_norm].append(failure_msg)
        
        if columns_validated_by_business_rules:
            print(f"\n   [PRE-STEP Row {index_etofs}] Business rules validated columns: {columns_validated_by_business_rules}")
        
        # ===== STEP 1: Check Origin and Destination Countries =====
        # ONLY perform country validation if country columns exist in BOTH rate card AND ETOF/LC
        
        # Check if rate card has country columns
        rate_card_has_country_cols = rc_origin_col is not None or rc_dest_col is not None
        
        if rate_card_has_country_cols:
            # Find origin and destination country columns (handle variations)
            shipment_orig_country_norm = None
            shipment_dest_country_norm = None
            etof_orig_col_found = None
            etof_dest_col_found = None
            
            for col in ['Origin Country', 'origin country', 'OriginCountry', 'origincountrycode', 'ORIGIN COUNTRY CODE']:
                if col in row_etofs:
                    shipment_orig_country_norm = normalize_value(row_etofs[col])
                    etof_orig_col_found = col
                    break
            
            for col in ['Destination Country', 'destination country', 'DestinationCountry', 'destinationcountrycode', 'DESTINATION COUNTRY CODE']:
                if col in row_etofs:
                    shipment_dest_country_norm = normalize_value(row_etofs[col])
                    etof_dest_col_found = col
                    break
            
            # Only perform country validation if ETOF/LC also has country columns
            etof_has_country_data = etof_orig_col_found is not None or etof_dest_col_found is not None
            
            # Check if origin/destination country columns were validated by business rules
            # (this will be checked again after business rule validation)
            origin_validated_by_br = normalize_column_name('Origin Country') in columns_validated_by_business_rules
            dest_validated_by_br = normalize_column_name('Destination Country') in columns_validated_by_business_rules
            
            if etof_has_country_data:
                if debug_conditions:
                    print(f"\n   [DEBUG] Country validation for row {index_etofs}:")
                    print(f"      Rate card origin col: {rc_origin_col}")
                    print(f"      Rate card dest col: {rc_dest_col}")
                    print(f"      ETOF origin col: {etof_orig_col_found} = '{shipment_orig_country_norm}'")
                    print(f"      ETOF dest col: {etof_dest_col_found} = '{shipment_dest_country_norm}'")
                    print(f"      Origin validated by business rule: {origin_validated_by_br}")
                    print(f"      Destination validated by business rule: {dest_validated_by_br}")
                
                # Skip country validation if validated by business rules
                if origin_validated_by_br and dest_validated_by_br:
                    if debug_conditions:
                        print(f"      [SKIP] Both countries validated by business rules - skipping country validation")
                elif origin_validated_by_br:
                    # Only check destination
                    if rc_dest_col and shipment_dest_country_norm is None:
                        comments_for_current_etofs_row.append("Destination country is missing")
                    elif shipment_dest_country_norm:
                        dest_matches = country_matches_rate_card(shipment_dest_country_norm, unique_rc_dest_countries_norm, dest_code_to_country)
                        if not dest_matches:
                            dest_val = row_etofs.get('Destination Country', row_etofs.get('destination country', 'N/A'))
                            comments_for_current_etofs_row.append(f"Destination country '{dest_val}' is missing")
                    if debug_conditions:
                        print(f"      [SKIP] Origin validated by business rule - only checking destination")
                elif dest_validated_by_br:
                    # Only check origin
                    if rc_origin_col and shipment_orig_country_norm is None:
                        comments_for_current_etofs_row.append("Origin country is missing")
                    elif shipment_orig_country_norm:
                        orig_matches = country_matches_rate_card(shipment_orig_country_norm, unique_rc_orig_countries_norm, origin_code_to_country)
                        if not orig_matches:
                            orig_val = row_etofs.get('Origin Country', row_etofs.get('origin country', 'N/A'))
                            comments_for_current_etofs_row.append(f"Origin country '{orig_val}' is missing")
                    if debug_conditions:
                        print(f"      [SKIP] Destination validated by business rule - only checking origin")
                else:
                    # Neither validated by business rule - do full validation
                    # Check if origin country is missing (only if rate card expects it)
                    if rc_origin_col and shipment_orig_country_norm is None:
                        comments_for_current_etofs_row.append("Origin country is missing")
                    # Check if destination country is missing (only if rate card expects it)
                    elif rc_dest_col and shipment_dest_country_norm is None:
                        comments_for_current_etofs_row.append("Destination country is missing")
                    # If both present (or not required), check if they exist in rate card
                    elif shipment_orig_country_norm and shipment_dest_country_norm:
                        # Check if origin matches (directly or via condition code mapping)
                        orig_matches = country_matches_rate_card(shipment_orig_country_norm, unique_rc_orig_countries_norm, origin_code_to_country)
                        dest_matches = country_matches_rate_card(shipment_dest_country_norm, unique_rc_dest_countries_norm, dest_code_to_country)
                        
                        if debug_conditions:
                            print(f"      Shipment Origin: '{shipment_orig_country_norm}' -> matches: {orig_matches}")
                            print(f"      Shipment Dest: '{shipment_dest_country_norm}' -> matches: {dest_matches}")
                            if shipment_orig_country_norm in origin_code_to_country:
                                print(f"      Origin code '{shipment_orig_country_norm}' maps to '{origin_code_to_country[shipment_orig_country_norm]}'")
                            if shipment_dest_country_norm in dest_code_to_country:
                                print(f"      Dest code '{shipment_dest_country_norm}' maps to '{dest_code_to_country[shipment_dest_country_norm]}'")
                        
                        if not orig_matches and not dest_matches:
                            comments_for_current_etofs_row.append("Origin-Destination are missing")
                        elif not orig_matches:
                            orig_val = row_etofs.get('Origin Country', row_etofs.get('origin country', 'N/A'))
                            comments_for_current_etofs_row.append(f"Origin country '{orig_val}' is missing")
                        elif not dest_matches:
                            dest_val = row_etofs.get('Destination Country', row_etofs.get('destination country', 'N/A'))
                            comments_for_current_etofs_row.append(f"Destination country '{dest_val}' is missing")
                    else:
                        # Both countries exist individually (via direct match or condition), check combination
                        # Map shipment codes to rate card country names for combination check
                        orig_for_combo = origin_code_to_country.get(shipment_orig_country_norm, shipment_orig_country_norm)
                        dest_for_combo = dest_code_to_country.get(shipment_dest_country_norm, shipment_dest_country_norm)
                        combination = (orig_for_combo, dest_for_combo)
                        
                        if debug_conditions:
                            print(f"      Checking combination: {combination}")
                            print(f"      Available combinations (first 5): {list(unique_rc_orig_dest_combinations)[:5]}")
                        
                        if combination not in unique_rc_orig_dest_combinations:
                            comments_for_current_etofs_row.append("Origin-Destination country combination is missing")
            else:
                if debug_conditions:
                    print(f"\n   [DEBUG] Row {index_etofs}: Skipping country validation - ETOF/LC has no country columns")
        else:
            if debug_conditions:
                print(f"\n   [DEBUG] Row {index_etofs}: Skipping country validation - Rate card has no country columns")
        
        # If country validation failed, skip matching and go to next row
        if comments_for_current_etofs_row:
            comment_text = '\n'.join(comments_for_current_etofs_row)
            df_etofs.loc[index_etofs, 'comment'] = comment_text
            print(f"   [COMMENT] Row {index_etofs}: Country validation failed")
            print(f"   [COMMENT]   -> '{comment_text}'")
            continue
        
        # ===== STEP 2: Check date within validity (working fine, keep as is) =====
        # Find Valid from/Valid to columns (once per row)
        valid_from_col = None
        valid_to_col = None
        for col in df_filtered_rate_card.columns:
            col_lower = str(col).lower()
            if 'valid' in col_lower and 'from' in col_lower:
                valid_from_col = col
            elif 'valid' in col_lower and 'to' in col_lower:
                valid_to_col = col
        
        # Find date column
        date_col = None
        date_value = None
        for col in ['SHIP_DATE', 'ship_date', 'Ship Date', 'ship date', 'SHIP DATE',
                   'Loading date', 'Loading Date', 'loading date', 'LOADING DATE']:
            if col in row_etofs:
                date_value = row_etofs[col]
                if pd.notna(date_value):
                    date_col = col
                    break
        
        # Check date validity before proceeding with matching
        date_invalid = False
        if date_col and date_value and valid_from_col and valid_to_col:
            # We'll check date validity for each match later, but first check if we should skip matching
            # For now, we'll proceed with matching and check dates during discrepancy identification
            pass
        
        # ===== STEP 3: Find rows with most common values =====
        # Only proceed with matching if countries are valid (no comments added yet)
        # Prepare normalized values for the current ETOFS row
        etofs_normalized_values = {
            col_norm: normalize_value(row_etofs[common_etofs_cols_original[i]])
            for i, col_norm in enumerate(common_columns_normalized)
            if i < len(common_etofs_cols_original) and common_etofs_cols_original[i] in row_etofs
        }
        
        max_matches = -1
        best_matching_rate_card_rows = []
        
        # Iterate through each row of df_filtered_rate_card
        for index_rate_card, row_rate_card in df_filtered_rate_card.iterrows():
            current_matches = 0
            
            # Prepare normalized values for the current Rate Card row
            rate_card_normalized_values = {
                col_norm: normalize_value(row_rate_card[common_rate_card_cols_original[i]])
                for i, col_norm in enumerate(common_columns_normalized)
                if i < len(common_rate_card_cols_original) and common_rate_card_cols_original[i] in row_rate_card
            }
            
            # Compare normalized values
            for col_norm in common_columns_normalized:
                if col_norm in etofs_normalized_values and col_norm in rate_card_normalized_values:
                    etofs_val = etofs_normalized_values[col_norm]
                    rc_val = rate_card_normalized_values[col_norm]
                    
                    # Special handling for postal code columns - "starts with" matching
                    is_postal_code_column = 'post' in col_norm.lower() or 'ship_post' in col_norm.lower() or 'cust_post' in col_norm.lower()
                    
                    if is_postal_code_column and etofs_val and rc_val:
                        # For postal codes: count as match if shipment starts with rate card value
                        if str(etofs_val).startswith(str(rc_val)):
                            current_matches += 1
                    elif etofs_val == rc_val:
                        current_matches += 1
            
            # Update best matches
            if current_matches > max_matches:
                max_matches = current_matches
                best_matching_rate_card_rows = [{'rate_card_row': row_rate_card.to_dict(), 'discrepancies': []}]
            elif current_matches == max_matches and current_matches > 0:  # Only append if there's at least one match
                best_matching_rate_card_rows.append({'rate_card_row': row_rate_card.to_dict(), 'discrepancies': []})
        
        # Track if we have too many matches (will be analyzed after discrepancies are collected)
        too_many_matches = len(best_matching_rate_card_rows) > 4
        
        # Capture Carrier agreement value from the best match (first match)
        if carrier_agreement_col and len(best_matching_rate_card_rows) > 0:
            first_best_match = best_matching_rate_card_rows[0]['rate_card_row']
            carrier_agreement_value = first_best_match.get(carrier_agreement_col)
            if carrier_agreement_value:
                df_etofs.loc[index_etofs, 'Rate Card'] = carrier_agreement_value
                print(f"   [RATE CARD] Row {index_etofs}: Carrier agreement = '{carrier_agreement_value}'")
        
        # Only proceed with date validation and discrepancy checking if we have matches
        if len(best_matching_rate_card_rows) == 0:
            # No matches found - add comment and skip to next row
            if not comments_for_current_etofs_row:
                comments_for_current_etofs_row.append("No matching rate card entries found")
            comment_text = '\n'.join(comments_for_current_etofs_row)
            df_etofs.loc[index_etofs, 'comment'] = comment_text
            print(f"   [COMMENT] Row {index_etofs}: No matches found")
            print(f"   [COMMENT]   -> '{comment_text}'")
            continue
        
        # ===== STEP 2: Check date within validity (working fine, keep as is) =====
        # Find Valid from/Valid to columns (once per row)
        valid_from_col = None
        valid_to_col = None
        for col in df_filtered_rate_card.columns:
            col_lower = str(col).lower()
            if 'valid' in col_lower and 'from' in col_lower:
                valid_from_col = col
            elif 'valid' in col_lower and 'to' in col_lower:
                valid_to_col = col
        
        # Find date column
        date_col = None
        date_value = None
        for col in ['SHIP_DATE', 'ship_date', 'Ship Date', 'ship date', 'SHIP DATE',
                   'Loading date', 'Loading Date', 'loading date', 'LOADING DATE']:
            if col in row_etofs:
                date_value = row_etofs[col]
                if pd.notna(date_value):
                    date_col = col
                    break
        
        # Debug: Print date column detection results
        print(f"\n   [DEBUG Row {index_etofs}] Date validation setup:")
        print(f"      - Date column found: '{date_col}' with value: '{date_value}' (type: {type(date_value).__name__})")
        print(f"      - Valid from column: '{valid_from_col}'")
        print(f"      - Valid to column: '{valid_to_col}'")
        
        # Check date validity for all best matches first (before discrepancy checking)
        # FILTER OUT matches that are outside the validity date range
        if date_col and date_value and valid_from_col and valid_to_col:
            valid_matches = []  # Only keep matches within validity range
            invalid_matches = []  # Track removed matches for logging
            print(f"      - Checking {len(best_matching_rate_card_rows)} best matching rate card rows for date validity...")
            
            for match_idx, best_match_info in enumerate(best_matching_rate_card_rows):
                rate_card_row_dict = best_match_info['rate_card_row']
                valid_from = rate_card_row_dict.get(valid_from_col)
                valid_to = rate_card_row_dict.get(valid_to_col)
                
                print(f"      - Match {match_idx + 1}: Valid from='{valid_from}' (type: {type(valid_from).__name__}), Valid to='{valid_to}' (type: {type(valid_to).__name__})")
                
                is_date_valid = True  # Assume valid unless proven otherwise
                
                if pd.notna(valid_from) and pd.notna(valid_to):
                    try:
                        # Try to parse date_value - handle YYYYMMDD format (e.g., 20250905)
                        date_str = str(date_value).strip()
                        if date_str.isdigit() and len(date_str) == 8:
                            # YYYYMMDD format
                            date_dt = pd.to_datetime(date_str, format='%Y%m%d', errors='coerce')
                            print(f"        Detected YYYYMMDD format: '{date_str}' -> {date_dt}")
                        else:
                            date_dt = pd.to_datetime(date_value, errors='coerce')
                        
                        # Parse valid_from - handle DD.MM.YYYY format (e.g., 01.07.2025 = July 1, 2025)
                        valid_from_str = str(valid_from).strip()
                        if '.' in valid_from_str:
                            # DD.MM.YYYY format (European date format)
                            valid_from_dt = pd.to_datetime(valid_from_str, format='%d.%m.%Y', errors='coerce')
                            print(f"        Detected DD.MM.YYYY format for valid_from: '{valid_from_str}' -> {valid_from_dt}")
                        elif valid_from_str.isdigit() and len(valid_from_str) == 8:
                            # DDMMYYYY format (no separators)
                            valid_from_dt = pd.to_datetime(valid_from_str, format='%d%m%Y', errors='coerce')
                            print(f"        Detected DDMMYYYY format for valid_from: '{valid_from_str}' -> {valid_from_dt}")
                        else:
                            valid_from_dt = pd.to_datetime(valid_from, dayfirst=True, errors='coerce')
                        
                        # Parse valid_to - handle DD.MM.YYYY format (e.g., 31.12.2025 = December 31, 2025)
                        valid_to_str = str(valid_to).strip()
                        if '.' in valid_to_str:
                            # DD.MM.YYYY format (European date format)
                            valid_to_dt = pd.to_datetime(valid_to_str, format='%d.%m.%Y', errors='coerce')
                            print(f"        Detected DD.MM.YYYY format for valid_to: '{valid_to_str}' -> {valid_to_dt}")
                        elif valid_to_str.isdigit() and len(valid_to_str) == 8:
                            # DDMMYYYY format (no separators)
                            valid_to_dt = pd.to_datetime(valid_to_str, format='%d%m%Y', errors='coerce')
                            print(f"        Detected DDMMYYYY format for valid_to: '{valid_to_str}' -> {valid_to_dt}")
                        else:
                            valid_to_dt = pd.to_datetime(valid_to, dayfirst=True, errors='coerce')
                        
                        print(f"        Parsed: date_dt={date_dt}, valid_from_dt={valid_from_dt}, valid_to_dt={valid_to_dt}")
                        
                        if pd.notna(date_dt) and pd.notna(valid_from_dt) and pd.notna(valid_to_dt):
                            if date_dt < valid_from_dt or date_dt > valid_to_dt:
                                is_date_valid = False
                                print(f"        Result: DATE INVALID (outside range) - EXCLUDING this match")
                            else:
                                print(f"        Result: Date is within valid range - KEEPING this match")
                        else:
                            print(f"        Result: Could not parse one or more dates (NaT detected) - keeping match")
                    except Exception as e:
                        print(f"        Result: EXCEPTION during date parsing: {e} - keeping match")
                else:
                    print(f"        Result: Skipped (valid_from or valid_to is NaN) - keeping match")
                
                # Add to appropriate list
                if is_date_valid:
                    valid_matches.append(best_match_info)
                else:
                    invalid_matches.append(best_match_info)
            
            # Update best_matching_rate_card_rows to only include valid matches
            original_count = len(best_matching_rate_card_rows)
            best_matching_rate_card_rows = valid_matches
            
            print(f"      - Filtered: {len(valid_matches)} valid matches, {len(invalid_matches)} removed (outside date range)")
            
            # Update Carrier agreement value if first match changed
            if carrier_agreement_col and len(best_matching_rate_card_rows) > 0:
                first_best_match = best_matching_rate_card_rows[0]['rate_card_row']
                new_carrier_agreement = first_best_match.get(carrier_agreement_col)
                if new_carrier_agreement and new_carrier_agreement != carrier_agreement_value:
                    carrier_agreement_value = new_carrier_agreement
                    df_etofs.loc[index_etofs, 'Rate Card'] = carrier_agreement_value
                    print(f"   [RATE CARD] Row {index_etofs}: Updated Carrier agreement = '{carrier_agreement_value}' (after date filtering)")
            
            # If ALL matches were filtered out (date invalid for all), add comment and skip
            if len(best_matching_rate_card_rows) == 0 and original_count > 0:
                print(f"      - All matches have invalid dates, skipping discrepancy checking")
                comments_for_current_etofs_row.append(f"Date '{date_value}' is outside valid date range for all matching rate card entries")
                comment_text = '\n'.join(comments_for_current_etofs_row)
                df_etofs.loc[index_etofs, 'comment'] = comment_text
                print(f"   [COMMENT] Row {index_etofs}: Date validation failed")
                print(f"   [COMMENT]   -> '{comment_text}'")
                continue
        else:
            print(f"      - Skipping date validation (missing: date_col={date_col is not None}, date_value={date_value is not None}, valid_from_col={valid_from_col is not None}, valid_to_col={valid_to_col is not None})")
        
        # ===== STEP 3: Business rules validation status (already done in PRE-STEP) =====
        print(f"\n   {'#'*70}")
        print(f"   # STEP 3: BUSINESS RULES VALIDATION STATUS")
        print(f"   {'#'*70}")
        
        # Note: Business rules were already validated in PRE-STEP before country validation
        # columns_validated_by_business_rules and business_rule_validation_messages are already populated
        print(f"   [DEBUG] Business rules already validated in PRE-STEP")
        print(f"   [DEBUG] Columns validated by business rules: {columns_validated_by_business_rules}")
        print(f"   [DEBUG] Validation messages: {len(business_rule_validation_messages)}")
        
        if columns_validated_by_business_rules:
            print(f"   [DEBUG Row {index_etofs}] COLUMNS TO SKIP (validated by business rules): {columns_validated_by_business_rules}")
        
        # Add ALL business rule validation messages (both success AND failure)
        if business_rule_validation_messages:
            for br_msg in business_rule_validation_messages:
                comments_for_current_etofs_row.append(br_msg)
                print(f"   [COMMENT] Adding business rule message: {br_msg}")
        
        # ===== STEP 4: Identify discrepancies for best matching rows =====
        print(f"\n   [DEBUG Row {index_etofs}] Identifying discrepancies for {len(best_matching_rate_card_rows)} best matching rows...")
        for match_idx, best_match_info in enumerate(best_matching_rate_card_rows):
            rate_card_row_dict = best_match_info['rate_card_row']
            discrepancies = []
            
            # Date validation (check for each match, but don't skip if only some are invalid)
            if date_col and date_value and valid_from_col and valid_to_col:
                valid_from = rate_card_row_dict.get(valid_from_col)
                valid_to = rate_card_row_dict.get(valid_to_col)
                
                print(f"      [DEBUG Match {match_idx + 1}] Date discrepancy check:")
                print(f"         - Ship date: '{date_value}' (type: {type(date_value).__name__})")
                print(f"         - Valid from: '{valid_from}' (type: {type(valid_from).__name__})")
                print(f"         - Valid to: '{valid_to}' (type: {type(valid_to).__name__})")
                
                if pd.notna(valid_from) and pd.notna(valid_to):
                    try:
                        # Convert dates to pandas datetime
                        # Handle YYYYMMDD format (e.g., 20250905)
                        date_str = str(date_value).strip()
                        if date_str.isdigit() and len(date_str) == 8:
                            # YYYYMMDD format
                            date_dt = pd.to_datetime(date_str, format='%Y%m%d', errors='coerce')
                            print(f"         - Detected YYYYMMDD format: '{date_str}' -> {date_dt}")
                        else:
                            date_dt = pd.to_datetime(date_value, errors='coerce')
                        
                        # Parse valid_from - handle DD.MM.YYYY format (e.g., 01.07.2025 = July 1, 2025)
                        valid_from_str = str(valid_from).strip()
                        if '.' in valid_from_str:
                            # DD.MM.YYYY format (European date format)
                            valid_from_dt = pd.to_datetime(valid_from_str, format='%d.%m.%Y', errors='coerce')
                            print(f"         - Detected DD.MM.YYYY format for valid_from: '{valid_from_str}' -> {valid_from_dt}")
                        elif valid_from_str.isdigit() and len(valid_from_str) == 8:
                            # DDMMYYYY format (no separators)
                            valid_from_dt = pd.to_datetime(valid_from_str, format='%d%m%Y', errors='coerce')
                            print(f"         - Detected DDMMYYYY format for valid_from: '{valid_from_str}' -> {valid_from_dt}")
                        else:
                            valid_from_dt = pd.to_datetime(valid_from, dayfirst=True, errors='coerce')
                        
                        # Parse valid_to - handle DD.MM.YYYY format (e.g., 31.12.2025 = December 31, 2025)
                        valid_to_str = str(valid_to).strip()
                        if '.' in valid_to_str:
                            # DD.MM.YYYY format (European date format)
                            valid_to_dt = pd.to_datetime(valid_to_str, format='%d.%m.%Y', errors='coerce')
                            print(f"         - Detected DD.MM.YYYY format for valid_to: '{valid_to_str}' -> {valid_to_dt}")
                        elif valid_to_str.isdigit() and len(valid_to_str) == 8:
                            # DDMMYYYY format (no separators)
                            valid_to_dt = pd.to_datetime(valid_to_str, format='%d%m%Y', errors='coerce')
                            print(f"         - Detected DDMMYYYY format for valid_to: '{valid_to_str}' -> {valid_to_dt}")
                        else:
                            valid_to_dt = pd.to_datetime(valid_to, dayfirst=True, errors='coerce')
                        
                        print(f"         - Parsed date_dt: {date_dt} (NaT: {pd.isna(date_dt)})")
                        print(f"         - Parsed valid_from_dt: {valid_from_dt} (NaT: {pd.isna(valid_from_dt)})")
                        print(f"         - Parsed valid_to_dt: {valid_to_dt} (NaT: {pd.isna(valid_to_dt)})")
                        
                        # Check if all dates were successfully parsed
                        if pd.notna(date_dt) and pd.notna(valid_from_dt) and pd.notna(valid_to_dt):
                            # Check if date is within valid range
                            if date_dt < valid_from_dt or date_dt > valid_to_dt:
                                print(f"         - Result: DATE DISCREPANCY - outside valid range")
                                discrepancies.append({
                                    'column': date_col,
                                    'etofs_value': date_value,
                                    'rate_card_value': f"Valid from: {valid_from} to {valid_to}",
                                    'condition': None,
                                    'type': 'date_range'
                                })
                            else:
                                print(f"         - Result: Date is within valid range (no discrepancy)")
                        else:
                            print(f"         - Result: Could not parse dates (NaT detected), skipping validation")
                    except Exception as e:
                        # If date parsing fails, skip this validation
                        print(f"         - Result: EXCEPTION during date parsing: {e}")
                else:
                    print(f"         - Result: Skipped (valid_from or valid_to is NaN)")
            
            for i, col_norm in enumerate(common_columns_normalized):
                if i >= len(common_etofs_cols_original) or i >= len(common_rate_card_cols_original):
                    continue
                    
                etofs_original_col = common_etofs_cols_original[i]
                rate_card_original_col = common_rate_card_cols_original[i]
                
                # Skip columns that were validated by business rules
                etofs_col_norm = normalize_column_name(etofs_original_col)
                if etofs_col_norm in columns_validated_by_business_rules:
                    if debug_conditions:
                        print(f"\n      {'>'*40}")
                        print(f"      [SKIPPING] Column '{etofs_original_col}' (normalized: '{etofs_col_norm}')")
                        print(f"      [SKIPPING]   Reason: VALIDATED BY BUSINESS RULE")
                        print(f"      [SKIPPING]   Skip set contains: {columns_validated_by_business_rules}")
                        print(f"      {'>'*40}")
                    continue
                
                etofs_val = row_etofs.get(etofs_original_col)
                rate_card_val = rate_card_row_dict.get(rate_card_original_col)
                
                # Normalize values for consistent comparison for discrepancy reporting
                normalized_etofs_val = normalize_value(etofs_val)
                normalized_rate_card_val = normalize_value(rate_card_val)
                
                # Special handling for postal code columns - check "starts with" instead of exact match
                is_postal_code_column = 'post' in col_norm.lower() or 'ship_post' in col_norm.lower() or 'cust_post' in col_norm.lower() or 'postal' in col_norm.lower()
                                
                # Only report discrepancy if normalized values are different
                if normalized_etofs_val != normalized_rate_card_val:
                    if debug_conditions:
                        print(f"      [DISCREPANCY CHECK] Column '{etofs_original_col}': VALUES DIFFER")
                        print(f"         - Shipment value: '{etofs_val}' (normalized: '{normalized_etofs_val}')")
                        print(f"         - Rate Card value: '{rate_card_val}' (normalized: '{normalized_rate_card_val}')")
                    
                    # For postal codes: check "starts with" instead of exact match
                    if is_postal_code_column and normalized_etofs_val and normalized_rate_card_val:
                        # For postal codes: shipment value should START WITH rate card value
                        # Example: RC has "194", shipment has "19454" -> OK (starts with "194")
                        if str(normalized_etofs_val).startswith(str(normalized_rate_card_val)):
                            if debug_conditions:
                                print(f"         - Result: POSTAL CODE MATCH (starts with)")
                            continue  # No discrepancy - postal code matches (starts with)
                    
                    # Check if ResMed value satisfies the condition for this rate card value
                    is_valid, matching_condition = check_value_against_conditions(
                        etofs_val, rate_card_val, etofs_original_col, conditions_dict, debug=debug_conditions
                    )
                    
                    if is_valid:
                        # Value satisfies condition - don't report as discrepancy
                        # Example: Rate Card has "NAC", condition says "NAC: RATE_TYPE is empty", 
                        #          ResMed has "nan" (empty) -> This is valid, no discrepancy
                        if debug_conditions:
                            print(f"         - Result: CONDITION SATISFIED - No discrepancy")
                            print(f"         - Matching condition: {matching_condition[:80] if matching_condition else 'N/A'}...")
                        pass  # Skip this discrepancy
                    else:
                        # Check if this is a business rule column with NaN value - skip it
                        # (the actual discrepancy is in the geo columns, not the zone column)
                        is_business_rule_col = False
                        if business_rules_lookup:
                            br_cols = business_rules_lookup.get('business_rule_columns', set())
                            for bc in br_cols:
                                if normalize_column_name(bc) == etofs_col_norm:
                                    is_business_rule_col = True
                                    break
                        
                        if is_business_rule_col and (etofs_val is None or pd.isna(etofs_val) or str(etofs_val).strip().lower() in ['', 'nan', 'none']):
                            if debug_conditions:
                                print(f"         - Result: SKIPPING - Business rule column with NaN value (actual issue is in geo columns)")
                            continue  # Skip - the real discrepancy was already reported in business rule validation
                        
                        # Value doesn't match and doesn't satisfy condition - report discrepancy
                        if debug_conditions:
                            print(f"         - Result: DISCREPANCY FOUND")
                        
                        # Try to find the condition text for the rate card value (even if not satisfied)
                        # This helps us show what codes the user should enter
                        condition_for_rc_value = find_condition_for_value(
                            rate_card_val, etofs_original_col, conditions_dict
                        )
                        
                        if debug_conditions and condition_for_rc_value:
                            print(f"         - Found condition for rate card value: {condition_for_rc_value[:80]}...")
                        
                        discrepancies.append({
                            'column': etofs_original_col,
                            'etofs_value': etofs_val,
                            'rate_card_value': rate_card_val,
                            'condition': condition_for_rc_value
                        })
                else:
                    if debug_conditions:
                        print(f"      [MATCH] Column '{etofs_original_col}': Values match ('{normalized_etofs_val}')")
            best_match_info['discrepancies'] = discrepancies
        
        # Check if any single match has more than 5 fields to update (discrepancies)
        has_too_many_fields_to_update = any(len(match['discrepancies']) > 5 for match in best_matching_rate_card_rows)
        if has_too_many_fields_to_update:
            comments_for_current_etofs_row.append("Please recheck the shipment details. Too many shipment details to update.")
        
        # If too many matches (>4 lanes), use pattern analysis to find dominant issue
        if too_many_matches:
            # Collect ALL discrepancies from all matches for pattern analysis
            all_discrepancies = []
            for match_info in best_matching_rate_card_rows:
                all_discrepancies.extend(match_info['discrepancies'])
            
            # Only analyze patterns if there are actual discrepancies
            if all_discrepancies:
                # Analyze patterns in discrepancies
                has_common_pattern, pattern_comment, minor_discrepancies = analyze_discrepancy_patterns(all_discrepancies, conditions_dict)
                comments_for_current_etofs_row.append(pattern_comment)
            else:
                # No discrepancies but multiple matching lanes - this is good!
                has_common_pattern = False
                minor_discrepancies = []
                # Don't add any negative comment - everything matches
            
            # If there's a common pattern, show "Also" for minor discrepancies with specific values
            if has_common_pattern:
                # Add "Also" lines for minor discrepancies
                for minor_disc in minor_discrepancies:
                    target_value = minor_disc.get('rate_card_value', '')
                    condition_text = minor_disc.get('condition')
                    
                    # Try to extract code from condition
                    if condition_text:
                        equals_match = re.search(r':\s*equals?\s+([^\n]+)', str(condition_text), re.IGNORECASE)
                        if equals_match:
                            target_value = equals_match.group(1).strip()
                    
                    also_comment = f"Also: {minor_disc.get('column', 'Unknown')}: '{minor_disc.get('etofs_value', '')}' → '{target_value}'"
                    comments_for_current_etofs_row.append(also_comment)
                
                comments_for_current_etofs_row.append(f"({len(best_matching_rate_card_rows)} possible rate lanes can be applied with this change)")
        
        # Check if "please recheck the shipment details" is already in comments
        # If so, don't add discrepancy details - this will be the full comment
        has_recheck_comment = "Please recheck the shipment details" in '\n'.join(comments_for_current_etofs_row)
        
        # Add discrepancy details to comments only if "please recheck" is not present and not too many matches
        if not has_recheck_comment and not too_many_matches:
            for match_idx, best_match_info in enumerate(best_matching_rate_card_rows):
                discrepancies = best_match_info['discrepancies']
                if discrepancies:
                    rate_card_row_dict = best_match_info['rate_card_row']
                    lane_num = rate_card_row_dict.get('Lane #', rate_card_row_dict.get('Lane#', 'N/A'))
                    comments_for_current_etofs_row.append(f"Discrepancies for Match {match_idx+1}:")
                    for disc in discrepancies:
                        if disc.get('type') == 'date_range':
                            # Special formatting for date range discrepancies
                            comment = f"  {disc['column']}: Shipment value '{disc['etofs_value']}' is outside valid date range ({disc['rate_card_value']})"
                        else:
                            # Check if there's a condition with "equals" - extract the code from it
                            target_value = disc['rate_card_value']
                            condition_text = disc.get('condition')
                            
                            print(f"   [DEBUG DISCREPANCY] Column: '{disc['column']}'")
                            print(f"   [DEBUG DISCREPANCY]   etofs_value: '{disc['etofs_value']}'")
                            print(f"   [DEBUG DISCREPANCY]   rate_card_value: '{disc['rate_card_value']}'")
                            print(f"   [DEBUG DISCREPANCY]   condition: '{condition_text}'")
                            
                            if condition_text:
                                # Try to extract code from condition like "Long Beach: equals LGB" or "Long Beach: equals LGB,LAX"
                                import re
                                equals_match = re.search(r':\s*equals?\s+([^\n]+)', str(condition_text), re.IGNORECASE)
                                print(f"   [DEBUG DISCREPANCY]   equals_match: {equals_match}")
                                if equals_match:
                                    # Get the codes (might be comma-separated like "LGB,LAX")
                                    codes = equals_match.group(1).strip()
                                    # Use the codes instead of the rate card value
                                    target_value = codes
                                    print(f"   [DEBUG DISCREPANCY]   Extracted code: '{codes}'")
                            
                            print(f"   [DEBUG DISCREPANCY]   Final target_value: '{target_value}'")
                            comment = f" {disc['column']}: Shipment value '{disc['etofs_value']}' needs to be changed to '{target_value}'"
                        comments_for_current_etofs_row.append(comment)
        
        # ===== PASS 1: Add business rule failure messages for sides that have NO passing rule =====
        # First, determine which sides have passing rules
        origin_side_passed = any('origin' in col for col in business_rule_columns_passed)
        destination_side_passed = any('destination' in col for col in business_rule_columns_passed)
        
        # Track which failure messages belong to which side AND their rule type
        # Rule types: "country_region" (only country matters) vs "postal_code_zone" (country + postal matter)
        origin_failure_messages = []  # list of (message, rule_type)
        destination_failure_messages = []  # list of (message, rule_type)
        
        if business_rule_columns_failed:
            for br_col_norm, failure_msgs in business_rule_columns_failed.items():
                # Only add failure messages for columns that have NO passing rule
                if br_col_norm not in business_rule_columns_passed:
                    is_origin_rule = 'origin' in br_col_norm
                    is_destination_rule = 'destination' in br_col_norm
                    
                    # Determine rule type from column name
                    # "Country Region" rules only check country
                    # "Postal Code Zone" rules check country + postal
                    is_country_region_rule = 'countryregion' in br_col_norm or 'country region' in br_col_norm.replace('_', ' ')
                    is_postal_zone_rule = 'postalcode' in br_col_norm or 'postal' in br_col_norm
                    rule_type = 'country_region' if is_country_region_rule else 'postal_code_zone' if is_postal_zone_rule else 'unknown'
                    
                    # Skip if a rule on the same side already passed
                    if is_origin_rule and origin_side_passed:
                        continue
                    if is_destination_rule and destination_side_passed:
                        continue
                    
                    for failure_msg in failure_msgs:
                        full_msg = f"[Business Rule Failed] {failure_msg}"
                        comments_for_current_etofs_row.append(full_msg)
                        # Track which side and rule type this message belongs to
                        if is_origin_rule:
                            origin_failure_messages.append((full_msg, rule_type))
                        elif is_destination_rule:
                            destination_failure_messages.append((full_msg, rule_type))
        
        # ===== PASS 2: Remove ONLY "Country Region" rule failures when no country discrepancy =====
        # Rule types:
        # - "Country Region" rules (EXPORT_S, IMPORT_S): Only check country
        #   → Remove if no country discrepancy (direct values match like ES = ES)
        # - "Postal Code Zone" rules (DE Zone 26): Check country + postal
        #   → ALWAYS keep (both country and postal failures are relevant)
        if origin_failure_messages or destination_failure_messages:
            # Check for actual COUNTRY discrepancies (non-business-rule comments) per side
            has_origin_country_discrepancy = False
            has_destination_country_discrepancy = False
            
            for comment in comments_for_current_etofs_row:
                comment_lower = comment.lower()
                # Skip business rule messages
                if comment_lower.startswith('[business rule'):
                    continue
                # Check for origin-related discrepancy (country, postal, zone, region)
                if ('origin country' in comment_lower or 
                    'origincountry' in comment_lower or
                    'origin postal' in comment_lower or
                    ('origin' in comment_lower and ('zone' in comment_lower or 'region' in comment_lower))):
                    has_origin_country_discrepancy = True
                # Check for destination-related discrepancy
                if ('destination country' in comment_lower or 
                    'destinationcountry' in comment_lower or
                    'destination postal' in comment_lower or
                    ('destination' in comment_lower and ('zone' in comment_lower or 'region' in comment_lower))):
                    has_destination_country_discrepancy = True
            
            # Determine which failure messages to remove
            # Only remove "Country Region" rule failures when no country discrepancy
            # ALWAYS keep "Postal Code Zone" rule failures
            messages_to_remove = set()
            
            if not has_origin_country_discrepancy and origin_failure_messages:
                for msg, rule_type in origin_failure_messages:
                    # Only remove "Country Region" rule failures
                    # Keep "Postal Code Zone" failures - they're always relevant
                    if rule_type == 'country_region':
                        print(f"   [FILTER] Removing origin Country Region failure (no origin geo discrepancy)")
                        messages_to_remove.add(msg)
                    # Keep postal_code_zone and unknown rule types
            
            if not has_destination_country_discrepancy and destination_failure_messages:
                for msg, rule_type in destination_failure_messages:
                    # Only remove "Country Region" rule failures
                    if rule_type == 'country_region':
                        print(f"   [FILTER] Removing destination Country Region failure (no destination geo discrepancy)")
                        messages_to_remove.add(msg)
                    # Keep postal_code_zone and unknown rule types
            
            # Filter out the irrelevant messages
            if messages_to_remove:
                comments_for_current_etofs_row = [
                    c for c in comments_for_current_etofs_row if c not in messages_to_remove
                ]
        
        if comments_for_current_etofs_row:
            comment_text = '\n'.join(comments_for_current_etofs_row)
            df_etofs.loc[index_etofs, 'comment'] = comment_text
            print(f"   [COMMENT] Row {index_etofs}: {len(comments_for_current_etofs_row)} items")
            print(f"   [COMMENT]   -> '{comment_text[:100]}{'...' if len(comment_text) > 100 else ''}'")
        else:
            df_etofs.loc[index_etofs, 'comment'] = 'No discrepancies found'
            print(f"   [COMMENT] Row {index_etofs}: No discrepancies found")
    
    return df_etofs


def run_matching(rate_card_file_path=None):
    """
    Run the matching workflow to match shipments with rate card.
    
    Args:
        rate_card_file_path (str, optional): Path to rate card file relative to "input/" folder.
                                            If None, will try to find rate_card.xlsx or rate_card.xls in input folder.
    
    Returns:
        str: Path to the output file (Matched_Shipments_with.xlsx) if successful, None otherwise
    """
    import sys
    
    print("="*80)
    print("COMPARE AND FIND (CANF) - Match Shipments with Rate Card")
    print("="*80)
    
    # Step 1: Get rate card from part4_rate_card_processing.py
    print("\n1. Getting Rate Card from part4_rate_card_processing.py...")
    
    #rate_card_file_path = ["rate_iff_1.xlsx", "rate_iff_2.xlsx"]
    # If rate_card_file_path not provided, try to find it
    if rate_card_file_path is None:
        input_folder = "input"
        possible_names = ["rate_card.xlsx", "rate_ahaha.xls"]
        for name in possible_names:
            full_path = os.path.join(input_folder, name)
            if os.path.exists(full_path):
                rate_card_file_path = name
                print(f"   Auto-detected rate card file: {rate_card_file_path}")
                break
        
        if rate_card_file_path is None:
            print(f"   [ERROR] Rate card file not found. Tried: {possible_names}")
            return None
    
    try:
        from part4_rate_card_processing import process_rate_card
        
        df_rate_card, rate_card_columns, rate_card_conditions = process_rate_card(rate_card_file_path)
        
        print(f"   Rate Card loaded: {df_rate_card.shape[0]} rows x {df_rate_card.shape[1]} columns")
        print(f"   Rate Card columns: {len(rate_card_columns)}")
        print(f"   Conditions loaded: {len(rate_card_conditions)} columns with conditions")
        if rate_card_conditions:
            print(f"   Columns with conditions: {list(rate_card_conditions.keys())}")
            print(f"\n   [DEBUG] Conditions content:")
            for col_name, cond in rate_card_conditions.items():
                cond_preview = str(cond)[:100] + "..." if len(str(cond)) > 100 else str(cond)
                print(f"      - {col_name}: {cond_preview}")
        else:
            print(f"   [DEBUG] No conditions found in rate card!")
        
    except ImportError as e:
        print(f"   [ERROR] Could not import part4_rate_card_processing: {e}")
        print("   Please ensure part4_rate_card_processing.py is in the same directory.")
        return None
    except Exception as e:
        print(f"   [ERROR] Failed to process rate card: {e}")
        import traceback
        traceback.print_exc()
        return None
    
    # Step 1b: Load Business Rules for matching validation
    print("\n1b. Loading Business Rules for matching validation...")
    business_rules_lookup = None
    try:
        business_rules_lookup = load_business_rules_for_matching(rate_card_file_path)
        if business_rules_lookup:
            print(f"   Business Rules loaded successfully:")
            print(f"      - Rules with country: {len(business_rules_lookup.get('rule_to_country', {}))}")
            print(f"      - Rules with postal codes: {len(business_rules_lookup.get('rule_to_postal_codes', {}))}")
            print(f"      - Columns containing rules: {business_rules_lookup.get('business_rule_columns', set())}")
        else:
            print(f"   No business rules found (this is OK if the file doesn't have them)")
    except Exception as e:
        print(f"   [WARNING] Could not load business rules: {e}")
        print(f"   Matching will continue without business rule validation")
    
    # Step 2: Get dataframes from vocabular.py output (partly_df/vocabulary_mapping.xlsx)
    print("\n2. Loading Shipment dataframes from vocabular.py output...")
    
    # Step 2: Get dataframes from vocabular.py output (partly_df/vocabulary_mapping.xlsx)
    print("\n2. Loading Shipment dataframes from vocabular.py output...")
    
    # Vocabular output is stored in partly_df folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    vocabular_output_path = os.path.join(script_dir, "partly_df", "vocabulary_mapping.xlsx")
    
    if not os.path.exists(vocabular_output_path):
        print(f"   [ERROR] vocabulary_mapping.xlsx not found at: {vocabular_output_path}")
        print(f"   Please ensure vocabular.py has been run and the file exists in the partly_df folder.")
        return None
    
    print(f"   Found vocabulary_mapping.xlsx at: {vocabular_output_path}")

    
    try:
        # Read Excel file with all sheets
        excel_file = pd.ExcelFile(vocabular_output_path)
        sheet_names = excel_file.sheet_names
        print(f"   Found sheets in Excel file: {sheet_names}")
        
        etof_renamed = None
        lc_renamed = None
        origin_renamed = None
        
        # Read ETOF sheet if it exists
        if 'ETOF' in sheet_names:
            etof_renamed = pd.read_excel(vocabular_output_path, sheet_name='ETOF')
            print(f"   Loaded ETOF DataFrame: {etof_renamed.shape[0]} rows x {etof_renamed.shape[1]} columns")
        else:
            print(f"   [WARNING] ETOF sheet not found in vocabular_output.xlsx")
        
        # Read LC sheet if it exists
        if 'LC' in sheet_names:
            lc_renamed = pd.read_excel(vocabular_output_path, sheet_name='LC')
            print(f"   Loaded LC DataFrame: {lc_renamed.shape[0]} rows x {lc_renamed.shape[1]} columns")
        else:
            print(f"   [WARNING] LC sheet not found in vocabular_output.xlsx")
        
        # Read Origin sheet if it exists
        if 'Origin' in sheet_names:
            origin_renamed = pd.read_excel(vocabular_output_path, sheet_name='Origin')
            print(f"   Loaded Origin DataFrame: {origin_renamed.shape[0]} rows x {origin_renamed.shape[1]} columns")
        else:
            print(f"   [INFO] Origin sheet not found in vocabular_output.xlsx (optional)")
        
        if etof_renamed is None and lc_renamed is None:
            print(f"   [ERROR] No ETOF or LC dataframes found in vocabulary_mapping.xlsx")
            print(f"   Please ensure vocabular.py has been run and generated the Excel file with ETOF or LC sheets.")
            return None
    
    except FileNotFoundError:
        print(f"   [ERROR] File not found: {vocabular_output_path}")
        print(f"   Please run vocabular.py first to generate partly_df/vocabulary_mapping.xlsx")
        return None
    except Exception as e:
        print(f"   [ERROR] Failed to read vocabular_output.xlsx: {e}")
        import traceback
        traceback.print_exc()
        return None
    
    # Step 3: Select dataframe to process (LC if present, otherwise ETOF)
    print("\n3. Selecting shipment dataframe:")
    
    df_to_process = None
    shipment_type = None
    
    # Priority: LC first, then ETOF
    if lc_renamed is not None and not lc_renamed.empty:
        df_to_process = lc_renamed
        shipment_type = "LC"
        print(f"   Using LC DataFrame: {df_to_process.shape[0]} rows x {df_to_process.shape[1]} columns")
    elif etof_renamed is not None and not etof_renamed.empty:
        df_to_process = etof_renamed
        shipment_type = "ETOF"
        print(f"   LC not available, using ETOF DataFrame: {df_to_process.shape[0]} rows x {df_to_process.shape[1]} columns")
    else:
        print("\n   [ERROR] No LC or ETOF dataframes available to process!")
        print("   Please ensure vocabular.py has been run and generated LC or ETOF sheets.")
        sys.exit(1)
    
    # Step 4: Filter to only rows with values in ETOF # column
    print(f"\n4. Filtering rows with values in ETOF # column...")
    
    # Find ETOF # column (handle variations)
    etof_col = None
    etof_col_variations = ['ETOF #', 'ETOF#', 'etof #', 'etof#', 'ETOF', 'etof']
    
    for col in df_to_process.columns:
        col_normalized = str(col).strip()
        for variation in etof_col_variations:
            if col_normalized.lower() == variation.lower() or col_normalized.lower().replace(' ', '') == variation.lower().replace(' ', ''):
                etof_col = col
                break
        if etof_col:
            break
    
    if etof_col:
        print(f"   Found ETOF column: '{etof_col}'")
        initial_row_count = len(df_to_process)
        
        # Filter to keep only rows where ETOF # has a value (not null, not empty, not NaN)
        df_to_process = df_to_process[df_to_process[etof_col].notna()]
        df_to_process = df_to_process[df_to_process[etof_col].astype(str).str.strip() != '']
        df_to_process = df_to_process[df_to_process[etof_col].astype(str).str.lower() != 'nan']
        
        filtered_row_count = len(df_to_process)
        removed_rows = initial_row_count - filtered_row_count
        
        print(f"   Initial rows: {initial_row_count}")
        print(f"   Rows with ETOF # values: {filtered_row_count}")
        print(f"   Rows removed (no ETOF # value): {removed_rows}")
        
        if filtered_row_count == 0:
            print(f"\n   [ERROR] No rows remaining after filtering for ETOF # values!")
            print(f"   Please ensure the dataframe has rows with values in the ETOF # column.")
            return None
    else:
        print(f"   [WARNING] ETOF # column not found. Processing all rows.")
        print(f"   Searched for columns: {etof_col_variations}")
        print(f"   Available columns: {list(df_to_process.columns)}")
    
    # Step 5: Print input dataframe before matching
    print(f"\n5. Input {shipment_type} DataFrame before matching:")
    print(f"   Shape: {df_to_process.shape[0]} rows x {df_to_process.shape[1]} columns")
    print(f"   Columns: {list(df_to_process.columns)}")
    print(f"\n   First few rows:")
    print(df_to_process.head())
    
    # Step 6: Find common columns and match
    print(f"\n6. Finding common columns...")
    common_columns = find_common_columns(df_to_process, df_rate_card)
    
    if not common_columns:
        print("\nError: No common columns found between shipment and rate card dataframes.")
        return None
    
    # Step 7: Match shipments with rate card
    print("\n" + "="*80)
    print("MATCHING SHIPMENTS WITH RATE CARD")
    print("="*80)
    print("Note: Values will be validated against conditional rules before reporting discrepancies.")
    
    # Debug: Print conditions being passed to matching function
    print(f"\n[DEBUG] Passing to matching function:")
    print(f"   - rate_card_conditions type: {type(rate_card_conditions)}")
    print(f"   - rate_card_conditions length: {len(rate_card_conditions) if rate_card_conditions else 0}")
    if rate_card_conditions:
        for col, cond in list(rate_card_conditions.items())[:3]:
            print(f"   - '{col}': {str(cond)[:80]}...")
    
    print(f"\n[DEBUG] Business rules being passed:")
    print(f"   - rate_card_file_path: {rate_card_file_path}")
    print(f"   - business_rules_lookup is None: {business_rules_lookup is None}")
    if business_rules_lookup:
        print(f"   - business_rules_lookup keys: {list(business_rules_lookup.keys())}")
        print(f"   - business_rule_columns: {business_rules_lookup.get('business_rule_columns', 'NOT SET')}")
    
    df_result = match_shipments_with_rate_card(
        df_to_process, df_rate_card, common_columns, 
        conditions_dict=rate_card_conditions,
        debug_conditions=True,
        rate_card_file_path=rate_card_file_path,
        business_rules_lookup=business_rules_lookup
    )
    
    # Step 8: Reorder columns and save results
    print("\n8. Reordering columns and saving results...")
    # Handle Colab environment where __file__ is not defined
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        # In Colab or interactive environments, use current working directory
        script_dir = os.getcwd()
    
    # Use absolute path to ensure it works even after directory changes
    output_file = os.path.abspath(os.path.join(script_dir, "Matched_Shipments_with.xlsx"))
    print(f"   Output file will be saved to: {output_file}")
    
    # Reorder columns: LC #, ETOF #, Shipment ID, Delivery Number, Carrier, Ship date, then others
    def reorder_columns(df):
        """Reorder columns with priority: LC #, ETOF #, Shipment ID, Delivery Number, Carrier, Ship date, then others."""
        if df is None or df.empty:
            return df
        
        # Find priority columns (handle variations)
        priority_columns = []
        other_columns = []
        
        # Define priority column patterns
        lc_patterns = ['LC #', 'LC#', 'lc #', 'lc#']
        etof_patterns = ['ETOF #', 'ETOF#', 'etof #', 'etof#']
        shipment_id_patterns = ['Shipment ID', 'ShipmentID', 'shipment id', 'shipmentid', 
                               'SHIPMENT_ID', 'SHIPMENT ID', 'Shipment', 'shipment', 'SHIPMENT']
        delivery_patterns = ['Delivery Number', 'DeliveryNumber', 'delivery number', 'deliverynumber',
                           'DELIVERY_NUMBER', 'Delivery', 'delivery', 'DELIVERY']
        carrier_patterns = ['Carrier', 'carrier', 'CARRIER']
        ship_date_patterns = ['SHIP_DATE', 'ship_date', 'Ship Date', 'ship date', 'SHIP DATE',
                             'Loading date', 'Loading Date', 'loading date', 'LOADING DATE']
        
        # Find and collect priority columns
        lc_col = None
        etof_col = None
        shipment_id_col = None
        delivery_col = None
        carrier_col = None
        ship_date_col = None
        
        for col in df.columns:
            col_str = str(col).strip()
            col_lower = col_str.lower().replace(' ', '').replace('_', '')
            
            # Check for LC #
            if not lc_col:
                for pattern in lc_patterns:
                    if col_lower == pattern.lower().replace(' ', '') or col_str == pattern:
                        lc_col = col
                        break
            
            # Check for ETOF #
            if not etof_col:
                for pattern in etof_patterns:
                    if col_lower == pattern.lower().replace(' ', '') or col_str == pattern:
                        etof_col = col
                        break
            
            # Check for Shipment ID
            if not shipment_id_col:
                for pattern in shipment_id_patterns:
                    pattern_lower = pattern.lower().replace(' ', '').replace('_', '')
                    if col_lower == pattern_lower or col_str == pattern:
                        shipment_id_col = col
                        break
                # Also check if column contains "shipment" and "id"
                if not shipment_id_col and 'shipment' in col_lower and 'id' in col_lower:
                    shipment_id_col = col
            
            # Check for Delivery Number
            if not delivery_col:
                for pattern in delivery_patterns:
                    pattern_lower = pattern.lower().replace(' ', '').replace('_', '')
                    if col_lower == pattern_lower or col_str == pattern:
                        delivery_col = col
                        break
                # Also check if column contains "delivery" and "number"
                if not delivery_col and 'delivery' in col_lower and 'number' in col_lower:
                    delivery_col = col
            
            # Check for Carrier
            if not carrier_col:
                for pattern in carrier_patterns:
                    if col_lower == pattern.lower() or col_str == pattern:
                        carrier_col = col
                        break
            
            # Check for Ship date
            if not ship_date_col:
                for pattern in ship_date_patterns:
                    if col_lower == pattern.lower().replace(' ', '_') or col_str == pattern:
                        ship_date_col = col
                        break
        
        # Build ordered column list
        ordered_columns = []
        
        # Add priority columns in order: LC #, ETOF #, Shipment ID, Delivery Number, Carrier, Ship date
        if lc_col:
            ordered_columns.append(lc_col)
        if etof_col:
            ordered_columns.append(etof_col)
        if shipment_id_col:
            ordered_columns.append(shipment_id_col)
        if delivery_col:
            ordered_columns.append(delivery_col)
        if carrier_col:
            ordered_columns.append(carrier_col)
        if ship_date_col:
            ordered_columns.append(ship_date_col)
        
        # Add all other columns (excluding priority columns and comment)
        priority_set = {lc_col, etof_col, shipment_id_col, delivery_col, carrier_col, ship_date_col}
        for col in df.columns:
            if col not in priority_set and col != 'comment':
                ordered_columns.append(col)
        
        # Add comment column last
        if 'comment' in df.columns:
            ordered_columns.append('comment')
        
        # Reorder dataframe
        df_reordered = df[ordered_columns]
        
        print(f"   Column order: {ordered_columns[:7]}... (and {len(ordered_columns) - 7} more)")
        
        return df_reordered
    
    # Reorder result dataframe
    df_result_reordered = reorder_columns(df_result)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_result_reordered.to_excel(writer, sheet_name='Matched Shipments', index=False)
            
            # Apply formatting to the workbook
            workbook = writer.book
            
            # Format "Matched Shipments" sheet
            if 'Matched Shipments' in workbook.sheetnames:
                ws = workbook['Matched Shipments']
                
                # Style header row
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set width with some padding, but cap at 50
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Freeze header row
                ws.freeze_panes = 'A2'
                
                # Style comment column if it exists
                if 'comment' in df_result_reordered.columns:
                    comment_col_idx = list(df_result_reordered.columns).index('comment') + 1
                    comment_col_letter = get_column_letter(comment_col_idx)
                    
                    # Make comment column wider
                    ws.column_dimensions[comment_col_letter].width = 60
                    
                    # Wrap text in comment column
                    for row in ws.iter_rows(min_row=2, min_col=comment_col_idx, max_col=comment_col_idx):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
        
        print(f"\n[SUCCESS] Results saved to: {output_file}")
        print(f"  - Sheet: Matched Shipments with comments (formatted)")
        print(f"\nTotal rows processed: {len(df_result)}")
        print(f"Total columns: {len(df_result.columns)} (reordered: LC #, ETOF #, Carrier, Ship date, then others)")
        
        # Show summary
        rows_with_comments = df_result[df_result['comment'] != '']
        rows_no_discrepancies = df_result[df_result['comment'] == 'No discrepancies found for best match.']
        print(f"  - Rows with comments/discrepancies: {len(rows_with_comments)}")
        print(f"  - Rows with no discrepancies: {len(rows_no_discrepancies)}")
        
    except ImportError:
        # Fallback if openpyxl formatting is not available
        print("   [WARNING] openpyxl formatting not available, saving without formatting...")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_result_reordered.to_excel(writer, sheet_name='Matched Shipments', index=False)
        print(f"\n[SUCCESS] Results saved to: {output_file} (without formatting)")
    except PermissionError:
        print(f"\n[ERROR] Permission denied: Cannot write to {output_file}")
        print("   The file is likely open in Excel. Please close it and run again.")
        return None
    except Exception as e:
        print(f"\n[ERROR] Failed to save results: {e}")
        import traceback
        traceback.print_exc()
        return None
    
    print(f"\n✅ Matching complete! Results saved to: {output_file}")
    print("="*80)
    
    return output_file


if __name__ == "__main__":
    run_matching()

