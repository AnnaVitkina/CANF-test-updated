import pandas as pd
import openpyxl
import os


def process_rate_card(file_path):
    """
    Process a Rate Card Excel file from the input folder.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder (e.g., "rate_card.xlsx")
    
    Returns:
        tuple: (dataframe, list of column names, conditions dictionary)
            - dataframe: Processed pandas DataFrame (filtered to black font columns)
            - list: List of column names in the processed dataframe
            - dict: Dictionary of conditions where keys are column names and values are condition text
    """
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Read the Excel file
    df_rate_card = pd.read_excel(full_path, sheet_name="Rate card", skiprows=2)
    
    # Find first column index (where data actually starts)
    first_column_index = None
    if df_rate_card is not None:
        for i, col in enumerate(df_rate_card.columns):
            if "nan" not in str(df_rate_card.iloc[0, i]).lower():
                first_column_index = i
                break
    
    if first_column_index is not None:
        df_rate_card = df_rate_card.iloc[:, :first_column_index]
    
    # Drop rows where the first column is NaN
    if df_rate_card is not None:
        df_rate_card.dropna(subset=[df_rate_card.columns[0]], inplace=True)
    
    # Set column names from first row
    new_columns = df_rate_card.iloc[0].tolist()
    df_rate_card.columns = new_columns
    df_rate_card = df_rate_card.iloc[1:]
    
    # Load the workbook to extract conditions and check font colors
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    sheet = workbook["Rate card"]
    
    # Find the header row that contains "Currency"
    first_data_row_index = None
    currency_index = None
    
    for row_index in range(1, min(10, sheet.max_row + 1)):
        row = sheet[row_index]
        row_values = [cell.value for cell in row]
        if "Currency" in row_values:
            currency_index = row_values.index("Currency")
            first_data_row_index = row_index
            break
    
    black_font_values = []
    column_notes = {}  # Will store conditions/notes for each column
    
    if first_data_row_index is not None and currency_index is not None:
        # Access the data in this row
        first_data_row = sheet[first_data_row_index]
        first_data_values = [cell.value for cell in first_data_row]
        truncated_data_values = first_data_values[:currency_index]
        
        # Extract conditional rules/notes from COMMENTS in the header row cells
        header_row_index = first_data_row_index
        if header_row_index and header_row_index <= sheet.max_row:
            for i, col_name in enumerate(truncated_data_values, 1):
                if col_name:  # Only process non-empty column names
                    header_cell = sheet.cell(row=header_row_index, column=i)
                    
                    # Check for comments (where conditional rules are stored)
                    if header_cell.comment:
                        comment_text = header_cell.comment.text
                        if comment_text and comment_text.strip():
                            column_notes[col_name] = comment_text.strip()
                    
                    # Also check for cell value notes (in case some are stored as values in row 2)
                    if col_name not in column_notes:
                        notes_row_index = 2
                        if notes_row_index <= sheet.max_row:
                            note_cell = sheet.cell(row=notes_row_index, column=i)
                            if note_cell.value and str(note_cell.value).strip():
                                column_notes[col_name] = str(note_cell.value).strip()
        
        # Check font color to identify black font columns (required columns)
        # Track column indices with black font to handle duplicate column names
        black_font_indices = []  # Store (index, column_name) tuples
        
        for i, value in enumerate(truncated_data_values):
            if i < len(first_data_row):
                cell = first_data_row[i]
                font_color = "black"
                if cell.font and cell.font.color:
                    hex_color = cell.font.color.rgb
                    if hex_color is not None:
                        # Convert to string and handle different formats
                        hex_str = str(hex_color).upper()
                        # Remove 'FF' prefix if present (ARGB format)
                        if hex_str.startswith('FF') and len(hex_str) == 8:
                            hex_str = hex_str[2:]
                        
                        # Check if it's black
                        if hex_str == '000000' or hex_str == '00000000':
                            font_color = "black"
                        else:
                            # Check if it's a shade of grey (R, G, and B are close)
                            try:
                                if len(hex_str) == 6:
                                    r, g, b = int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
                                    # Check if it's a shade of grey (R, G, and B are close)
                                    if abs(r - g) < 10 and abs(g - b) < 10 and r > 0:  # Grey (not black, not white)
                                        font_color = "grey"
                                    else:
                                        font_color = "other non-black"  # For colors that are not black or grey
                            except (ValueError, IndexError):
                                pass
                
                if font_color == "black":
                    black_font_values.append(value)
                    black_font_indices.append(i)  # Track the index
    
    # Filter the DataFrame to keep only the columns with black font
    # Handle duplicate column names by using positional selection
    if df_rate_card is not None and black_font_indices:
        # Get the original column positions in the dataframe
        # The dataframe columns should correspond to truncated_data_values after setting columns from row
        
        # Check for duplicate column names and keep only black font versions
        seen_columns = {}  # {column_name: index_in_black_font_indices}
        indices_to_keep = []
        
        for idx in black_font_indices:
            col_name = truncated_data_values[idx] if idx < len(truncated_data_values) else None
            if col_name is not None:
                if col_name not in seen_columns:
                    # First occurrence of this column name with black font
                    seen_columns[col_name] = idx
                    indices_to_keep.append(idx)
                else:
                    # Duplicate column name - skip it (keep only the first black font occurrence)
                    print(f"   [DEBUG] Duplicate column '{col_name}' found at index {idx}, keeping first occurrence at index {seen_columns[col_name]}")
        
        # Select columns by position using iloc
        if indices_to_keep:
            # Map the indices to dataframe column positions
            # The dataframe was truncated to first_column_index columns
            valid_indices = [i for i in indices_to_keep if i < len(df_rate_card.columns)]
            if valid_indices:
                df_filtered_rate_card = df_rate_card.iloc[:, valid_indices]
                # Update black_font_values to match the filtered columns
                black_font_values = [truncated_data_values[i] for i in valid_indices if i < len(truncated_data_values)]
            else:
                df_filtered_rate_card = df_rate_card
        else:
            df_filtered_rate_card = df_rate_card
    else:
        df_filtered_rate_card = df_rate_card
    
    # Get list of column names
    column_names = df_filtered_rate_card.columns.tolist()
    
    # Create conditions dictionary (only for columns that exist in the filtered dataframe)
    # Use cleaned conditions for better matching
    conditions = {}
    for col_name in column_names:
        if col_name in column_notes:
            # Clean the condition text for better parsing
            raw_condition = column_notes[col_name]
            cleaned_condition = clean_condition_text(raw_condition)
            conditions[col_name] = cleaned_condition
            
            # Debug: Print raw vs cleaned conditions
            print(f"   [DEBUG] Condition for '{col_name}':")
            print(f"      Raw: {raw_condition[:80]}..." if len(raw_condition) > 80 else f"      Raw: {raw_condition}")
            print(f"      Cleaned: {cleaned_condition[:80]}..." if len(cleaned_condition) > 80 else f"      Cleaned: {cleaned_condition}")
    
    return df_filtered_rate_card, column_names, conditions


def process_business_rules(file_path):
    """
    Process the Business rules tab from a Rate Card Excel file.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder
    
    Returns:
        dict: Dictionary containing:
            - 'postal_code_zones': list of zone rules with name, country, postal_codes, exclude
            - 'country_regions': list of region rules with name, country, postal_codes, exclude
            - 'no_data_added': list of entries with no data
            - 'raw_rules': all parsed rules as a list of dicts
    """
    import re
    
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Load the workbook
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    
    # Check if "Business rules" sheet exists
    if "Business rules" not in workbook.sheetnames:
        print(f"   [WARNING] 'Business rules' sheet not found in {file_path}")
        return {
            'postal_code_zones': [],
            'country_regions': [],
            'no_data_added': [],
            'raw_rules': []
        }
    
    sheet = workbook["Business rules"]
    
    # DEBUG: Print sheet info
    print(f"\n{'='*60}")
    print(f"[DEBUG] BUSINESS RULES SHEET ANALYSIS")
    print(f"{'='*60}")
    print(f"   Sheet name: 'Business rules'")
    print(f"   Total rows in sheet: {sheet.max_row}")
    print(f"   Total columns: {sheet.max_column}")
    print(f"   Available sheets: {workbook.sheetnames}")
    
    # STEP 1: Read all rows and filter out empty ones (skip first 2 rows)
    print(f"\n   [DEBUG] Step 1: Reading and filtering rows (skipping first 2 rows)...")
    
    all_rows = []  # Will store (original_row_idx, row_values) tuples
    for row_idx in range(3, sheet.max_row + 1):
        row = sheet[row_idx]
        row_values = [cell.value for cell in row]
        
        # Check if row is empty
        is_empty = all(v is None or (isinstance(v, str) and v.strip() == '') for v in row_values)
        
        if not is_empty:
            all_rows.append((row_idx, row_values))
    
    print(f"   [DEBUG] Total non-empty rows found: {len(all_rows)} (out of {sheet.max_row - 2} after skipping first 2)")
    
    # DEBUG: Print first 20 non-empty rows to see structure
    print(f"\n   [DEBUG] First 20 non-empty rows content:")
    for i, (row_idx, row_values) in enumerate(all_rows[:20]):
        non_empty = [(col_i, v) for col_i, v in enumerate(row_values) if v is not None]
        print(f"      Row {row_idx}: {non_empty}")
    
    if len(all_rows) > 20:
        print(f"      ... and {len(all_rows) - 20} more rows")
    
    # Marker values to look for (case-insensitive)
    markers = ['postal code zones', 'country regions', 'no data added']
    
    # Result structure
    result = {
        'postal_code_zones': [],
        'country_regions': [],
        'no_data_added': [],
        'raw_rules': []
    }
    
    # Track sections and their header columns
    current_section = None
    header_columns = {}  # Maps column index to header name
    waiting_for_header = False  # Flag to indicate we found a marker and are waiting for header row
    
    print(f"\n   [DEBUG] Step 2: Searching for markers: {markers}")
    print(f"   [DEBUG] Structure: MARKER row -> HEADER row (below) -> DATA rows")
    
    # Process non-empty rows
    for i, (row_idx, row_values) in enumerate(all_rows):
        # Check if this row contains a marker (section header)
        row_text_lower = ' '.join(str(v).lower() for v in row_values if v is not None)
        
        found_marker = None
        for marker in markers:
            if marker in row_text_lower:
                found_marker = marker
                print(f"\n   [DEBUG] >>> MARKER FOUND: '{marker}' at row {row_idx}")
                break
        
        if found_marker:
            # This is a marker row - next non-empty row will be the header
            current_section = found_marker.replace(' ', '_')
            waiting_for_header = True
            header_columns = {}  # Reset header columns for new section
            print(f"   [DEBUG]     Section: '{current_section}'")
            print(f"   [DEBUG]     Waiting for header row...")
            continue
        
        # If we're waiting for header, this row should be the header
        if waiting_for_header:
            waiting_for_header = False
            header_columns = {}
            
            print(f"   [DEBUG]     Header row (row {row_idx}): {[v for v in row_values if v is not None]}")
            
            for col_idx, cell_value in enumerate(row_values):
                if cell_value:
                    header_name = str(cell_value).strip().lower()
                    # Normalize header names
                    if 'name' in header_name:
                        header_columns[col_idx] = 'name'
                    elif 'country' in header_name:
                        header_columns[col_idx] = 'country'
                    elif 'postal' in header_name or 'code' in header_name:
                        header_columns[col_idx] = 'postal_code'
                    elif 'exclude' in header_name:
                        header_columns[col_idx] = 'exclude'
                    else:
                        header_columns[col_idx] = header_name
            
            print(f"   [DEBUG]     Mapped header columns: {header_columns}")
            continue
        
        # If we're in a section and have header columns, parse the data row
        if current_section and header_columns:
            rule_data = {
                'section': current_section,
                'name': None,
                'country': None,
                'postal_code': None,
                'exclude': None
            }
            
            # Extract values based on header columns
            for col_idx, header_name in header_columns.items():
                if col_idx < len(row_values):
                    value = row_values[col_idx]
                    if value is not None:
                        rule_data[header_name] = str(value).strip() if value else None
            
            # Only add if we have at least a name or postal code
            if rule_data['name'] or rule_data['postal_code'] or rule_data['country']:
                result['raw_rules'].append(rule_data)
                print(f"   [DEBUG]     + Rule added: {rule_data}")
                
                # Add to appropriate section list
                if current_section == 'postal_code_zones':
                    result['postal_code_zones'].append(rule_data)
                elif current_section == 'country_regions':
                    result['country_regions'].append(rule_data)
                elif current_section == 'no_data_added':
                    result['no_data_added'].append(rule_data)
    
    print(f"\n{'='*60}")
    print(f"[DEBUG] BUSINESS RULES SUMMARY")
    print(f"{'='*60}")
    print(f"   - Postal Code Zones: {len(result['postal_code_zones'])} rules")
    print(f"   - Country Regions: {len(result['country_regions'])} rules")
    print(f"   - No Data Added: {len(result['no_data_added'])} entries")
    print(f"   - Total raw rules: {len(result['raw_rules'])}")
    
    if not result['raw_rules']:
        print(f"\n   [WARNING] No rules were found! Possible issues:")
        print(f"      1. Markers not found in expected format")
        print(f"      2. Headers not in row above markers")
        print(f"      3. Data structure different than expected")
    
    return result


def transform_business_rules_to_conditions(business_rules):
    """
    Transform parsed business rules into condition format.
    
    Args:
        business_rules (dict): Output from process_business_rules()
    
    Returns:
        dict: Dictionary mapping zone/region names to their conditions
              Format: {zone_name: {'country': 'XX', 'postal_codes': ['12', '34'], 'exclude': bool}}
              Note: For country_regions, postal_codes will be empty (only country is validated)
    """
    conditions = {}
    
    for rule in business_rules.get('raw_rules', []):
        name = rule.get('name')
        if not name:
            continue
        
        section = rule.get('section', '')
        
        # Parse postal codes (comma-separated, possibly with spaces)
        # For country_regions, we don't use postal codes - only country matters
        postal_code_str = rule.get('postal_code', '')
        postal_codes = []
        
        if section != 'country_regions' and postal_code_str:
            # Split by comma and clean up each code
            postal_codes = [code.strip() for code in str(postal_code_str).split(',') if code.strip()]
        
        # Determine if this is an exclusion rule
        exclude_value = rule.get('exclude')
        is_exclude = False
        if exclude_value:
            exclude_str = str(exclude_value).lower().strip()
            is_exclude = exclude_str in ['yes', 'true', '1', 'x', 'exclude']
        
        condition = {
            'section': section,
            'country': rule.get('country'),
            'postal_codes': postal_codes,
            'exclude': is_exclude,
            'raw_postal_code': postal_code_str if section != 'country_regions' else ''
        }
        
        conditions[name] = condition
    
    return conditions


def format_business_rule_condition(rule_name, condition):
    """
    Format a business rule condition into a readable string.
    
    Args:
        rule_name (str): Name of the rule/zone
        condition (dict): Condition dictionary from transform_business_rules_to_conditions
    
    Returns:
        str: Human-readable condition string
    """
    parts = []
    
    if condition.get('country'):
        parts.append(f"Country: {condition['country']}")
    
    if condition.get('postal_codes'):
        prefix_list = ', '.join(condition['postal_codes'][:5])
        if len(condition['postal_codes']) > 5:
            prefix_list += f", ... (+{len(condition['postal_codes']) - 5} more)"
        parts.append(f"Postal codes starting with: {prefix_list}")
    
    if condition.get('exclude'):
        parts.append("(EXCLUDE)")
    
    return ' | '.join(parts) if parts else 'No conditions'


def find_business_rule_columns(rate_card_df, business_rules_conditions):
    """
    Find which columns in the rate card contain business rule values.
    
    Args:
        rate_card_df (pd.DataFrame): The rate card dataframe
        business_rules_conditions (dict): Dictionary of business rule conditions with rule names as keys
    
    Returns:
        dict: Dictionary with:
            - 'rule_to_columns': {rule_name: [list of columns where found]}
            - 'column_to_rules': {column_name: [list of rules found in it]}
            - 'unique_columns': set of unique column names that contain any business rule
    """
    rule_names = list(business_rules_conditions.keys())
    
    result = {
        'rule_to_columns': {},  # Which columns contain each rule
        'column_to_rules': {},  # Which rules are in each column
        'unique_columns': set()
    }
    
    if rate_card_df is None or rate_card_df.empty or not rule_names:
        return result
    
    print(f"\n{'='*60}")
    print(f"[DEBUG] FINDING BUSINESS RULE COLUMNS IN RATE CARD")
    print(f"{'='*60}")
    print(f"   Searching for {len(rule_names)} rule names in {len(rate_card_df.columns)} columns...")
    
    # Create a set of rule names for faster lookup (case-insensitive)
    rule_names_lower = {str(name).lower(): name for name in rule_names}
    
    # For each column, check which rule names are present
    for col in rate_card_df.columns:
        try:
            # Get unique values in this column
            unique_values = rate_card_df[col].dropna().unique()
            
            # Check each unique value against rule names
            for val in unique_values:
                val_str = str(val).strip().lower()
                
                if val_str in rule_names_lower:
                    original_rule_name = rule_names_lower[val_str]
                    
                    # Track rule to columns mapping
                    if original_rule_name not in result['rule_to_columns']:
                        result['rule_to_columns'][original_rule_name] = []
                    if col not in result['rule_to_columns'][original_rule_name]:
                        result['rule_to_columns'][original_rule_name].append(col)
                    
                    # Track column to rules mapping
                    if col not in result['column_to_rules']:
                        result['column_to_rules'][col] = []
                    if original_rule_name not in result['column_to_rules'][col]:
                        result['column_to_rules'][col].append(original_rule_name)
                    
                    result['unique_columns'].add(col)
        except Exception as e:
            # Skip columns that can't be processed
            pass
    
    # Initialize empty lists for rules not found
    for rule_name in rule_names:
        if rule_name not in result['rule_to_columns']:
            result['rule_to_columns'][rule_name] = []
    
    # Print results
    print(f"\n   [RESULT] Unique columns containing business rules:")
    if result['unique_columns']:
        for col in sorted(result['unique_columns']):
            rules_in_col = result['column_to_rules'].get(col, [])
            print(f"      - '{col}': {len(rules_in_col)} rules found")
            # Show first few rules as examples
            if rules_in_col:
                examples = rules_in_col[:3]
                if len(rules_in_col) > 3:
                    print(f"         Examples: {examples} ... (+{len(rules_in_col) - 3} more)")
                else:
                    print(f"         Rules: {examples}")
    else:
        print(f"      No columns found containing business rule values")
    
    print(f"\n   [SUMMARY] {len(result['unique_columns'])} unique columns contain business rule values")
    
    return result


def get_business_rules_lookup(file_path):
    """
    Get a lookup dictionary from business rule names to their country and postal codes.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
    
    Returns:
        dict: Dictionary with:
            - 'rule_to_country': {rule_name: country_code}
            - 'rule_to_postal_codes': {rule_name: [list of postal codes]}
            - 'business_rule_columns': set of column names containing business rules
            - 'all_rules': list of all rule data with name, country, postal_codes
    """
    # Process business rules
    business_rules = process_business_rules(file_path)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    # Get rate card to find which columns contain business rules
    rate_card_df, rate_card_columns, _ = process_rate_card(file_path)
    business_rule_cols_info = find_business_rule_columns(rate_card_df, business_rules_conditions)
    
    result = {
        'rule_to_country': {},
        'rule_to_postal_codes': {},
        'business_rule_columns': business_rule_cols_info.get('unique_columns', set()),
        'column_to_rules': business_rule_cols_info.get('column_to_rules', {}),
        'all_rules': []
    }
    
    for rule_name, condition in business_rules_conditions.items():
        country = condition.get('country')
        postal_codes = condition.get('postal_codes', [])
        
        if country:
            result['rule_to_country'][rule_name] = country
        if postal_codes:
            result['rule_to_postal_codes'][rule_name] = postal_codes
        
        result['all_rules'].append({
            'name': rule_name,
            'country': country,
            'postal_codes': postal_codes,
            'section': condition.get('section'),
            'exclude': condition.get('exclude', False)
        })
    
    print(f"\n[DEBUG] Business Rules Lookup:")
    print(f"   - Rules with country: {len(result['rule_to_country'])}")
    print(f"   - Rules with postal codes: {len(result['rule_to_postal_codes'])}")
    print(f"   - Columns containing rules: {sorted(result['business_rule_columns'])}")
    
    return result


def get_required_geo_columns():
    """
    Get the list of required geographic columns that should be in the final output.
    These are derived from business rules and should be mapped from ETOF/LC files.
    
    Returns:
        list: List of required column names for origin/destination country and postal codes
    """
    return [
        'Origin Country',
        'Origin Postal Code', 
        'Destination Country',
        'Destination Postal Code'
    ]


def clean_condition_text(condition_text):
    """
    Clean up condition text for better readability.
    
    Transforms:
        "Conditional rules:
        1. 33321-6422: TOPOSTALCODE starts with 33321-6422,333216422"
    To:
        "1. 33321-6422: starts with 33321-6422,333216422"
    """
    import re
    
    if not condition_text:
        return condition_text
    
    # Remove "Conditional rules:" header (case insensitive)
    cleaned = re.sub(r'(?i)^conditional\s*rules\s*:\s*\n?', '', condition_text.strip())
    
    # Remove column name references like "TOPOSTALCODE ", "FROMPOSTALCODE ", etc.
    # Pattern: After the colon and value identifier, remove uppercase column names followed by space
    # Example: "33321-6422: TOPOSTALCODE starts with" -> "33321-6422: starts with"
    cleaned = re.sub(r':\s*[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r': \1', cleaned)
    
    # Also handle cases without numbered format
    cleaned = re.sub(r'^[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r'\1', cleaned, flags=re.MULTILINE)
    
    # Clean up extra whitespace and newlines
    lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
    cleaned = '\n'.join(lines)
    
    return cleaned


def save_rate_card_output(file_path, output_path=None):
    """
    Process rate card and save output to Excel file with data and conditions.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
        output_path (str): Optional output path. If None, saves to "Filtered_Rate_Card_with_Conditions.xlsx"
    
    Returns:
        str: Path to the saved Excel file
    """
    # Process the rate card
    rate_card_dataframe, rate_card_column_names, rate_card_conditions = process_rate_card(file_path)
    
    # Process business rules
    business_rules = process_business_rules(file_path)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    # Find which columns in rate card contain business rule values
    business_rule_columns = find_business_rule_columns(rate_card_dataframe, business_rules_conditions)
    
    # Set output path - save to partly_df folder (relative to script location)
    if output_path is None:
        # Get the directory where this script is located
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Ensure partly_df folder exists in the script's directory
        partly_df_folder = os.path.join(script_dir, "partly_df")
        if not os.path.exists(partly_df_folder):
            os.makedirs(partly_df_folder)
        output_path = os.path.join(partly_df_folder, "Filtered_Rate_Card_with_Conditions.xlsx")
    
    # Create conditions DataFrame with cleaned condition text
    conditions_data = []
    for col_name in rate_card_column_names:
        raw_condition = rate_card_conditions.get(col_name, "")
        cleaned_condition = clean_condition_text(raw_condition) if raw_condition else ""
        conditions_data.append({
            'Column': col_name,
            'Has Condition': 'Yes' if col_name in rate_card_conditions else 'No',
            'Condition Rule': cleaned_condition
        })
    
    df_conditions = pd.DataFrame(conditions_data)
    
    # Save to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Rate Card Data
        rate_card_dataframe.to_excel(writer, sheet_name='Rate Card Data', index=False)
        
        # Sheet 2: Conditions
        df_conditions.to_excel(writer, sheet_name='Conditions', index=False)
        
        # Sheet 3: Business Rules
        business_rules_data = []
        for rule_name, condition in business_rules_conditions.items():
            # Get the columns where this rule is found
            rule_columns = business_rule_columns['rule_to_columns'].get(rule_name, [])
            columns_str = ', '.join(rule_columns) if rule_columns else '(not found in data)'
            
            business_rules_data.append({
                'Rule Name': rule_name,
                'Section': condition.get('section', '').replace('_', ' ').title(),
                'Country': condition.get('country', ''),
                'Postal Codes': condition.get('raw_postal_code', ''),
                'Exclude': 'Yes' if condition.get('exclude') else 'No',
                'Rate Card Columns': columns_str,
                'Formatted Condition': format_business_rule_condition(rule_name, condition)
            })
        
        df_business_rules = pd.DataFrame(business_rules_data)
        if not df_business_rules.empty:
            df_business_rules.to_excel(writer, sheet_name='Business Rules', index=False)
        
        # Sheet 4: Summary
        unique_cols_list = sorted(business_rule_columns['unique_columns']) if business_rule_columns['unique_columns'] else ['(none)']
        
        summary_data = {
            'Metric': [
                'Total Rows',
                'Total Columns',
                'Columns with Conditions',
                'Columns without Conditions',
                'Business Rules - Postal Code Zones',
                'Business Rules - Country Regions',
                'Business Rules - No Data Added',
                'Columns Using Business Rules',
                'Business Rule Column Names',
                'Source File'
            ],
            'Value': [
                len(rate_card_dataframe),
                len(rate_card_column_names),
                len(rate_card_conditions),
                len(rate_card_column_names) - len(rate_card_conditions),
                len(business_rules.get('postal_code_zones', [])),
                len(business_rules.get('country_regions', [])),
                len(business_rules.get('no_data_added', [])),
                len(business_rule_columns['unique_columns']),
                ', '.join(unique_cols_list),
                file_path
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    print(f"\n✅ Rate Card output saved to: {output_path}")
    print(f"   - Sheet 'Rate Card Data': {len(rate_card_dataframe)} rows x {len(rate_card_column_names)} columns")
    print(f"   - Sheet 'Conditions': {len(rate_card_conditions)} columns with conditions")
    print(f"   - Sheet 'Business Rules': {len(business_rules_conditions)} rules")
    print(f"   - Sheet 'Summary': Overview statistics")
    
    return output_path


def extract_general_info(file_path):
    """
    Extract metadata from the "General info" tab of a rate card.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder
    
    Returns:
        dict: Dictionary containing:
            - 'carrier_agreement': Agreement number value
            - 'valid_from': Start date of validity period
            - 'valid_to': End date of validity period
            - 'raw_validity_period': Original validity period string
    """
    import re
    
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    result = {
        'carrier_agreement': None,
        'valid_from': None,
        'valid_to': None,
        'raw_validity_period': None,
        'source_file': file_path
    }
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(full_path, data_only=True)
        
        # Check if "General info" sheet exists
        if "General info" not in workbook.sheetnames:
            print(f"   [WARNING] 'General info' sheet not found in {file_path}")
            return result
        
        sheet = workbook["General info"]
        
        # Search for "Agreement number" and "Validity period" in the first column
        for row_idx in range(1, min(50, sheet.max_row + 1)):  # Search first 50 rows
            cell_value = sheet.cell(row=row_idx, column=1).value
            
            if cell_value:
                cell_str = str(cell_value).strip().lower()
                
                # Look for Agreement number
                if 'agreement' in cell_str and 'number' in cell_str:
                    agreement_value = sheet.cell(row=row_idx, column=2).value
                    if agreement_value:
                        result['carrier_agreement'] = str(agreement_value).strip()
                        print(f"   [INFO] Found Agreement number: {result['carrier_agreement']}")
                
                # Look for Validity period
                if 'validity' in cell_str and 'period' in cell_str:
                    validity_value = sheet.cell(row=row_idx, column=2).value
                    if validity_value:
                        result['raw_validity_period'] = str(validity_value).strip()
                        print(f"   [INFO] Found Validity period: {result['raw_validity_period']}")
                        
                        # Parse the validity period (format: "DD.MM.YYYY - DD.MM.YYYY")
                        validity_str = str(validity_value).strip()
                        
                        # Try different separators: " - ", "-", " – ", "–"
                        separators = [' - ', ' – ', ' — ', '-', '–', '—']
                        dates = None
                        
                        for sep in separators:
                            if sep in validity_str:
                                parts = validity_str.split(sep)
                                if len(parts) == 2:
                                    dates = [parts[0].strip(), parts[1].strip()]
                                    break
                        
                        if dates and len(dates) == 2:
                            result['valid_from'] = dates[0]
                            result['valid_to'] = dates[1]
                            print(f"   [INFO] Parsed dates - Valid from: {result['valid_from']}, Valid to: {result['valid_to']}")
                        else:
                            print(f"   [WARNING] Could not parse validity period: {validity_str}")
        
        workbook.close()
        
    except Exception as e:
        print(f"   [ERROR] Failed to extract general info from {file_path}: {str(e)}")
    
    return result


def get_mandatory_columns(file_path):
    """
    Get the list of mandatory (black font) columns from a rate card.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder
    
    Returns:
        list: Sorted list of mandatory column names
    """
    _, column_names, _ = process_rate_card(file_path)
    return sorted(column_names)


def validate_mandatory_columns(file_paths):
    """
    Validate that all rate cards have the same mandatory columns.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
    
    Returns:
        tuple: (is_valid, reference_columns, differences)
            - is_valid: True if all files have matching mandatory columns
            - reference_columns: The mandatory columns from the first file
            - differences: Dictionary of {file_path: {'missing': [...], 'extra': [...]}}
    
    Raises:
        ValueError: If mandatory columns don't match across files
    """
    if not file_paths:
        raise ValueError("No file paths provided")
    
    print(f"\n{'='*60}")
    print(f"[VALIDATION] Checking mandatory columns across {len(file_paths)} rate cards")
    print(f"{'='*60}")
    
    # Get columns from the first file as reference
    reference_file = file_paths[0]
    reference_columns = set(get_mandatory_columns(reference_file))
    
    print(f"   Reference file: {reference_file}")
    print(f"   Reference columns ({len(reference_columns)}): {sorted(reference_columns)}")
    
    differences = {}
    is_valid = True
    
    # Compare with other files
    for file_path in file_paths[1:]:
        current_columns = set(get_mandatory_columns(file_path))
        
        missing = reference_columns - current_columns
        extra = current_columns - reference_columns
        
        if missing or extra:
            is_valid = False
            differences[file_path] = {
                'missing': sorted(missing),
                'extra': sorted(extra)
            }
            print(f"\n   [MISMATCH] {file_path}:")
            if missing:
                print(f"      Missing columns: {sorted(missing)}")
            if extra:
                print(f"      Extra columns: {sorted(extra)}")
        else:
            print(f"   [OK] {file_path}: Columns match")
    
    return is_valid, sorted(reference_columns), differences


def combine_business_rules(file_paths):
    """
    Combine business rules from multiple rate card files.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
    
    Returns:
        dict: Combined business rules with source file tracking
    """
    print(f"\n{'='*60}")
    print(f"[COMBINING] Business rules from {len(file_paths)} rate cards")
    print(f"{'='*60}")
    
    combined_rules = {
        'postal_code_zones': [],
        'country_regions': [],
        'no_data_added': [],
        'raw_rules': [],
        'source_files': {}
    }
    
    for file_path in file_paths:
        print(f"\n   Processing business rules from: {file_path}")
        
        rules = process_business_rules(file_path)
        
        # Track source file for each rule
        for rule in rules.get('raw_rules', []):
            rule['source_file'] = file_path
            combined_rules['raw_rules'].append(rule)
        
        for zone in rules.get('postal_code_zones', []):
            zone['source_file'] = file_path
            combined_rules['postal_code_zones'].append(zone)
        
        for region in rules.get('country_regions', []):
            region['source_file'] = file_path
            combined_rules['country_regions'].append(region)
        
        for entry in rules.get('no_data_added', []):
            entry['source_file'] = file_path
            combined_rules['no_data_added'].append(entry)
        
        combined_rules['source_files'][file_path] = len(rules.get('raw_rules', []))
    
    print(f"\n   [SUMMARY] Combined business rules:")
    print(f"      - Total postal code zones: {len(combined_rules['postal_code_zones'])}")
    print(f"      - Total country regions: {len(combined_rules['country_regions'])}")
    print(f"      - Total no data added: {len(combined_rules['no_data_added'])}")
    print(f"      - Total raw rules: {len(combined_rules['raw_rules'])}")
    
    return combined_rules


def combine_conditions(file_paths):
    """
    Combine conditions from multiple rate card files.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
    
    Returns:
        dict: Combined conditions dictionary with source tracking
    """
    print(f"\n{'='*60}")
    print(f"[COMBINING] Conditions from {len(file_paths)} rate cards")
    print(f"{'='*60}")
    
    combined_conditions = {}
    
    for file_path in file_paths:
        print(f"\n   Processing conditions from: {file_path}")
        
        _, _, conditions = process_rate_card(file_path)
        
        for col_name, condition in conditions.items():
            if col_name not in combined_conditions:
                combined_conditions[col_name] = {
                    'condition': condition,
                    'source_files': [file_path]
                }
            else:
                # If condition already exists, check if it's the same
                existing_condition = combined_conditions[col_name]['condition']
                if existing_condition != condition:
                    print(f"      [WARNING] Different condition for '{col_name}' in {file_path}")
                    # Keep both conditions by appending source
                    combined_conditions[col_name]['source_files'].append(file_path)
                    # Optionally merge conditions
                    combined_conditions[col_name]['condition'] += f"\n[From {file_path}]: {condition}"
                else:
                    combined_conditions[col_name]['source_files'].append(file_path)
        
        print(f"      Found {len(conditions)} conditions")
    
    # Convert back to simple format for compatibility
    simple_conditions = {col: data['condition'] for col, data in combined_conditions.items()}
    
    print(f"\n   [SUMMARY] Total unique conditions: {len(simple_conditions)}")
    
    return simple_conditions


def process_multiple_rate_cards(file_paths, validate_columns=True):
    """
    Process multiple rate card files and combine them into a single dataset.
    
    This function:
    1. Validates that mandatory columns match across all files (optional)
    2. Reads all business rules and conditions and combines them
    3. Processes each rate card and extracts data
    4. Adds metadata from "General info" tab (Carrier agreement, Valid from, Valid to)
    5. Combines all rate cards into one dataframe
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
        validate_columns (bool): Whether to validate mandatory columns match (default: True)
    
    Returns:
        tuple: (combined_dataframe, column_names, combined_conditions, combined_business_rules)
            - combined_dataframe: Merged pandas DataFrame with all lanes and metadata columns
            - column_names: List of column names in the combined dataframe
            - combined_conditions: Dictionary of all conditions
            - combined_business_rules: Dictionary of all business rules
    
    Raises:
        ValueError: If mandatory columns don't match and validate_columns is True
    """
    if not file_paths:
        raise ValueError("No file paths provided")
    
    if isinstance(file_paths, str):
        file_paths = [file_paths]
    
    print(f"\n{'='*60}")
    print(f"[PROCESSING] Multiple Rate Cards")
    print(f"{'='*60}")
    print(f"   Files to process: {len(file_paths)}")
    for fp in file_paths:
        print(f"      - {fp}")
    
    # Step 1: Validate mandatory columns if required
    if validate_columns and len(file_paths) > 1:
        is_valid, reference_columns, differences = validate_mandatory_columns(file_paths)
        
        if not is_valid:
            error_msg = "Mandatory columns do not match across rate cards:\n"
            for file_path, diff in differences.items():
                error_msg += f"\n  {file_path}:"
                if diff['missing']:
                    error_msg += f"\n    Missing: {diff['missing']}"
                if diff['extra']:
                    error_msg += f"\n    Extra: {diff['extra']}"
            raise ValueError(error_msg)
        
        print(f"\n   [OK] All rate cards have matching mandatory columns")
    
    # Step 2: Combine business rules from all files
    combined_business_rules = combine_business_rules(file_paths)
    
    # Step 3: Combine conditions from all files
    combined_conditions = combine_conditions(file_paths)
    
    # Step 4: Process each rate card and add metadata
    print(f"\n{'='*60}")
    print(f"[PROCESSING] Individual Rate Cards with Metadata")
    print(f"{'='*60}")
    
    all_dataframes = []
    
    for file_path in file_paths:
        print(f"\n   Processing: {file_path}")
        
        # Extract general info metadata
        general_info = extract_general_info(file_path)
        
        # Process the rate card
        df, columns, _ = process_rate_card(file_path)
        
        # Add metadata columns
        df['Carrier agreement'] = general_info.get('carrier_agreement', '')
        df['Valid from'] = general_info.get('valid_from', '')
        df['Valid to'] = general_info.get('valid_to', '')
        df['Source file'] = file_path
        
        print(f"      Rows: {len(df)}, Columns: {len(df.columns)}")
        print(f"      Carrier agreement: {general_info.get('carrier_agreement', 'N/A')}")
        print(f"      Valid from: {general_info.get('valid_from', 'N/A')}")
        print(f"      Valid to: {general_info.get('valid_to', 'N/A')}")
        
        all_dataframes.append(df)
    
    # Step 5: Combine all dataframes
    print(f"\n{'='*60}")
    print(f"[COMBINING] All Rate Cards into One DataFrame")
    print(f"{'='*60}")
    
    if len(all_dataframes) == 1:
        combined_df = all_dataframes[0]
    else:
        combined_df = pd.concat(all_dataframes, ignore_index=True)
    
    # Get final column names
    column_names = combined_df.columns.tolist()
    
    print(f"   Combined DataFrame:")
    print(f"      - Total rows: {len(combined_df)}")
    print(f"      - Total columns: {len(column_names)}")
    print(f"      - Source files: {len(file_paths)}")
    
    # Show sample of data per source file
    if 'Source file' in combined_df.columns:
        print(f"\n   Rows per source file:")
        for sf in combined_df['Source file'].unique():
            count = len(combined_df[combined_df['Source file'] == sf])
            print(f"      - {sf}: {count} rows")
    
    return combined_df, column_names, combined_conditions, combined_business_rules


def save_combined_rate_cards(file_paths, output_path=None, validate_columns=True):
    """
    Process multiple rate cards, combine them, and save to Excel.
    
    This is a convenience function that wraps process_multiple_rate_cards
    and saves the result to an Excel file.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
        output_path (str): Optional output path. If None, saves to partly_df folder
        validate_columns (bool): Whether to validate mandatory columns match
    
    Returns:
        str: Path to the saved combined Excel file
    """
    # Process all rate cards
    combined_df, column_names, combined_conditions, combined_business_rules = \
        process_multiple_rate_cards(file_paths, validate_columns)
    
    # Set output path
    if output_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        partly_df_folder = os.path.join(script_dir, "partly_df")
        if not os.path.exists(partly_df_folder):
            os.makedirs(partly_df_folder)
        output_path = os.path.join(partly_df_folder, "Combined_Rate_Cards.xlsx")
    
    # Transform business rules conditions for saving
    business_rules_conditions = transform_business_rules_to_conditions(combined_business_rules)
    
    # Create conditions DataFrame
    conditions_data = []
    for col_name in column_names:
        if col_name in ['Carrier agreement', 'Valid from', 'Valid to', 'Source file']:
            continue  # Skip metadata columns
        raw_condition = combined_conditions.get(col_name, "")
        cleaned_condition = clean_condition_text(raw_condition) if raw_condition else ""
        conditions_data.append({
            'Column': col_name,
            'Has Condition': 'Yes' if col_name in combined_conditions else 'No',
            'Condition Rule': cleaned_condition
        })
    
    df_conditions = pd.DataFrame(conditions_data)
    
    # Save to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Combined Rate Card Data
        combined_df.to_excel(writer, sheet_name='Rate Card Data', index=False)
        
        # Sheet 2: Conditions
        df_conditions.to_excel(writer, sheet_name='Conditions', index=False)
        
        # Sheet 3: Business Rules
        business_rules_data = []
        for rule_name, condition in business_rules_conditions.items():
            business_rules_data.append({
                'Rule Name': rule_name,
                'Section': condition.get('section', '').replace('_', ' ').title(),
                'Country': condition.get('country', ''),
                'Postal Codes': condition.get('raw_postal_code', ''),
                'Exclude': 'Yes' if condition.get('exclude') else 'No',
                'Source File': condition.get('source_file', '')
            })
        
        df_business_rules = pd.DataFrame(business_rules_data)
        if not df_business_rules.empty:
            df_business_rules.to_excel(writer, sheet_name='Business Rules', index=False)
        
        # Sheet 4: Summary
        summary_data = {
            'Metric': [
                'Total Rows',
                'Total Columns',
                'Source Files',
                'Columns with Conditions',
                'Total Business Rules',
                'Postal Code Zones',
                'Country Regions'
            ],
            'Value': [
                len(combined_df),
                len(column_names),
                ', '.join(file_paths),
                len(combined_conditions),
                len(combined_business_rules.get('raw_rules', [])),
                len(combined_business_rules.get('postal_code_zones', [])),
                len(combined_business_rules.get('country_regions', []))
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    print(f"\n✅ Combined Rate Cards saved to: {output_path}")
    print(f"   - Sheet 'Rate Card Data': {len(combined_df)} rows x {len(column_names)} columns")
    print(f"   - Sheet 'Conditions': {len(combined_conditions)} conditions")
    print(f"   - Sheet 'Business Rules': {len(business_rules_conditions)} rules")
    print(f"   - Sheet 'Summary': Overview statistics")
    
    return output_path


def process_rate_card_from_combined(combined_file_path):
    """
    Process a previously combined rate card file.
    
    This function is designed to be compatible with process_rate_card output format,
    allowing downstream code to work with combined rate cards seamlessly.
    
    Args:
        combined_file_path (str): Path to the combined rate card Excel file
    
    Returns:
        tuple: (dataframe, list of column names, conditions dictionary)
            Same format as process_rate_card for compatibility
    """
    print(f"\n[INFO] Loading combined rate card from: {combined_file_path}")
    
    # Read the data sheet
    df = pd.read_excel(combined_file_path, sheet_name='Rate Card Data')
    
    # Read conditions
    try:
        df_conditions = pd.read_excel(combined_file_path, sheet_name='Conditions')
        conditions = {}
        for _, row in df_conditions.iterrows():
            if row.get('Has Condition') == 'Yes' and row.get('Condition Rule'):
                conditions[row['Column']] = row['Condition Rule']
    except Exception as e:
        print(f"   [WARNING] Could not read conditions: {e}")
        conditions = {}
    
    column_names = df.columns.tolist()
    
    print(f"   Loaded {len(df)} rows, {len(column_names)} columns, {len(conditions)} conditions")
    
    return df, column_names, conditions


def process_rate_card_extended(file_paths, validate_columns=True):
    """
    Extended rate card processor that handles both single and multiple files.
    
    This function maintains backward compatibility with process_rate_card while
    adding support for multiple rate cards. It can be used as a drop-in replacement
    for process_rate_card in places that need multi-file support.
    
    Args:
        file_paths (str or list): Single file path or list of file paths relative to "input/" folder
        validate_columns (bool): Whether to validate mandatory columns match (for multiple files)
    
    Returns:
        tuple: (dataframe, list of column names, conditions dictionary)
            Same format as process_rate_card for full compatibility
    
    Usage:
        # Single file (same as process_rate_card)
        df, cols, conditions = process_rate_card_extended("rate_card.xlsx")
        
        # Multiple files (combined with metadata)
        df, cols, conditions = process_rate_card_extended(["rate_card_1.xlsx", "rate_card_2.xlsx"])
    """
    # Handle single file - delegate to original process_rate_card
    if isinstance(file_paths, str):
        print(f"\n[INFO] Processing single rate card: {file_paths}")
        return process_rate_card(file_paths)
    
    # Handle list with single file
    if isinstance(file_paths, list) and len(file_paths) == 1:
        print(f"\n[INFO] Processing single rate card: {file_paths[0]}")
        return process_rate_card(file_paths[0])
    
    # Handle multiple files
    print(f"\n[INFO] Processing {len(file_paths)} rate cards...")
    combined_df, column_names, combined_conditions, _ = process_multiple_rate_cards(
        file_paths, validate_columns
    )
    
    return combined_df, column_names, combined_conditions


# Global variable to store the combined rate card path (for use with process_rate_card wrapper)
_COMBINED_RATE_CARD_PATH = None


def set_combined_rate_card(file_paths, validate_columns=True):
    """
    Combine multiple rate cards and set up for use with process_rate_card.
    
    This function processes multiple rate cards, combines them, saves to a temporary
    file, and sets up the environment so that subsequent calls to process_rate_card
    (using special path "_combined_") will return the combined data.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
        validate_columns (bool): Whether to validate mandatory columns match
    
    Returns:
        str: Path to the combined rate card file
    
    Usage:
        # Set up combined rate cards
        combined_path = set_combined_rate_card(["rc1.xlsx", "rc2.xlsx", "rc3.xlsx"])
        
        # Now use the combined path with existing code
        df, cols, conditions = process_rate_card(combined_path)  # Not directly compatible
        
        # Or use the extended function
        df, cols, conditions = process_rate_card_extended(["rc1.xlsx", "rc2.xlsx"])
    """
    global _COMBINED_RATE_CARD_PATH
    
    # Save combined rate cards
    output_path = save_combined_rate_cards(file_paths, validate_columns=validate_columns)
    _COMBINED_RATE_CARD_PATH = output_path
    
    print(f"\n[INFO] Combined rate card saved and set: {output_path}")
    print(f"[INFO] Use process_rate_card_from_combined('{output_path}') to load the combined data")
    
    return output_path


# Alias for backward compatibility - result.py uses this name
process_rate_cards = process_rate_card_extended


#if __name__ == "__main__":
    # =========================================================================
    # Multiple Rate Cards Processing
    # =========================================================================
    
    #MULTIPLE_FILES = ["rate_iff_2.xlsx", "rate_iff_1.xlsx"]
    
    #print("\n" + "="*60)
    #print("TESTING: process_multiple_rate_cards")
    #print("="*60)
    #combined_df, columns, conditions, business_rules = process_multiple_rate_cards(MULTIPLE_FILES)
    #print(f"\nCombined result: {len(combined_df)} rows, {len(columns)} columns")
    #print(f"Sample of metadata columns:")
    #print(combined_df[['Carrier agreement', 'Valid from', 'Valid to', 'Source file']].head())
    
    # Save combined rate cards to Excel
    #combined_path = save_combined_rate_cards(MULTIPLE_FILES)
    #print(f"\nCombined file saved to: {combined_path}")
    
    # Load the combined file back
    #df_loaded, cols_loaded, conds_loaded = process_rate_card_from_combined(combined_path)
    #print(f"\nLoaded back: {len(df_loaded)} rows")
    
    # =========================================================================
    # Show DataFrame sample
    # =========================================================================
    #print("\n" + "="*60)
    #print("DATAFRAME SAMPLE")
    #print("="*60)
    #print(f"\nColumns: {columns}")
    #print(f"\nFirst 5 rows:")
    #print(combined_df.head())
