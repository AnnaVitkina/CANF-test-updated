"""
Multiple Rate Cards Processing Module

This module provides functions to process, validate, and combine multiple rate cards.

Usage in Google Colab:
    from multiple_rates import upload_and_merge_rate_cards
    upload_and_merge_rate_cards()

Usage programmatically:
    from multiple_rates import save_combined_rate_cards
    save_combined_rate_cards(["rate1.xlsx", "rate2.xlsx"])
"""

import pandas as pd
import openpyxl
import os
import shutil
import sys
import io
import warnings
from contextlib import contextmanager

from part4_rate_card_processing import (
    process_rate_card as _process_rate_card_orig,
    process_business_rules as _process_business_rules_orig,
    transform_business_rules_to_conditions,
    clean_condition_text
)


@contextmanager
def _suppress_output():
    """Suppress stdout and warnings."""
    old_stdout = sys.stdout
    old_stderr = sys.stderr
    sys.stdout = io.StringIO()
    sys.stderr = io.StringIO()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            yield
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr


def process_rate_card(file_path):
    """Process rate card with suppressed output."""
    with _suppress_output():
        return _process_rate_card_orig(file_path)


def process_business_rules(file_path):
    """Process business rules with suppressed output."""
    with _suppress_output():
        return _process_business_rules_orig(file_path)


def extract_general_info(file_path):
    """Extract metadata from the 'General info' tab of a rate card."""
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    result = {
        'carrier_agreement': None,
        'valid_from': None,
        'valid_to': None,
        'raw_validity_period': None,
        'source_file': file_path
    }
    
    try:
        workbook = openpyxl.load_workbook(full_path, data_only=True)
        
        if "General info" not in workbook.sheetnames:
            return result
        
        sheet = workbook["General info"]
        
        for row_idx in range(1, min(50, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row_idx, column=1).value
            
            if cell_value:
                cell_str = str(cell_value).strip().lower()
                
                if 'agreement' in cell_str and 'number' in cell_str:
                    agreement_value = sheet.cell(row=row_idx, column=2).value
                    if agreement_value:
                        result['carrier_agreement'] = str(agreement_value).strip()
                
                if 'validity' in cell_str and 'period' in cell_str:
                    validity_value = sheet.cell(row=row_idx, column=2).value
                    if validity_value:
                        result['raw_validity_period'] = str(validity_value).strip()
                        validity_str = str(validity_value).strip()
                        
                        separators = [' - ', ' – ', ' — ', '-', '–', '—']
                        for sep in separators:
                            if sep in validity_str:
                                parts = validity_str.split(sep)
                                if len(parts) == 2:
                                    result['valid_from'] = parts[0].strip()
                                    result['valid_to'] = parts[1].strip()
                                    break
        
        workbook.close()
    except Exception:
        pass
    
    return result


def get_mandatory_columns(file_path):
    """Get the list of mandatory (black font) columns from a rate card."""
    _, column_names, _ = process_rate_card(file_path)
    return sorted(column_names)


def validate_mandatory_columns(file_paths):
    """Validate that all rate cards have the same mandatory columns."""
    if not file_paths:
        raise ValueError("No file paths provided")
    
    reference_columns = set(get_mandatory_columns(file_paths[0]))
    differences = {}
    is_valid = True
    
    for file_path in file_paths[1:]:
        current_columns = set(get_mandatory_columns(file_path))
        missing = reference_columns - current_columns
        extra = current_columns - reference_columns
        
        if missing or extra:
            is_valid = False
            differences[file_path] = {'missing': sorted(missing), 'extra': sorted(extra)}
    
    return is_valid, sorted(reference_columns), differences


def combine_business_rules(file_paths):
    """Combine business rules from multiple rate card files."""
    combined_rules = {
        'postal_code_zones': [], 'country_regions': [], 
        'no_data_added': [], 'raw_rules': [], 'source_files': {}
    }
    
    for file_path in file_paths:
        rules = process_business_rules(file_path)
        
        for rule in rules.get('raw_rules', []):
            rule['source_file'] = file_path
            combined_rules['raw_rules'].append(rule)
        
        for zone in rules.get('postal_code_zones', []):
            zone['source_file'] = file_path
            combined_rules['postal_code_zones'].append(zone)
        
        for region in rules.get('country_regions', []):
            region['source_file'] = file_path
            combined_rules['country_regions'].append(region)
        
        for entry in rules.get('no_data_added', []):
            entry['source_file'] = file_path
            combined_rules['no_data_added'].append(entry)
        
        combined_rules['source_files'][file_path] = len(rules.get('raw_rules', []))
    
    return combined_rules


def combine_conditions(file_paths):
    """Combine conditions from multiple rate card files."""
    combined_conditions = {}
    
    for file_path in file_paths:
        _, _, conditions = process_rate_card(file_path)
        
        for col_name, condition in conditions.items():
            if col_name not in combined_conditions:
                combined_conditions[col_name] = {'condition': condition, 'source_files': [file_path]}
            else:
                existing = combined_conditions[col_name]['condition']
                if existing != condition:
                    combined_conditions[col_name]['source_files'].append(file_path)
                    combined_conditions[col_name]['condition'] += f"\n[From {file_path}]: {condition}"
                else:
                    combined_conditions[col_name]['source_files'].append(file_path)
    
    return {col: data['condition'] for col, data in combined_conditions.items()}


def process_multiple_rate_cards(file_paths, validate_columns=True):
    """Process multiple rate card files and combine them into a single dataset."""
    if not file_paths:
        raise ValueError("No file paths provided")
    
    if isinstance(file_paths, str):
        file_paths = [file_paths]
    
    # Validate columns
    if validate_columns and len(file_paths) > 1:
        is_valid, _, differences = validate_mandatory_columns(file_paths)
        if not is_valid:
            error_msg = "Column mismatch:\n"
            for fp, diff in differences.items():
                error_msg += f"  {fp}: Missing={diff['missing']}, Extra={diff['extra']}\n"
            raise ValueError(error_msg)
    
    # Combine rules and conditions
    combined_business_rules = combine_business_rules(file_paths)
    combined_conditions = combine_conditions(file_paths)
    
    # Process each rate card
    all_dataframes = []
    for file_path in file_paths:
        general_info = extract_general_info(file_path)
        df, _, _ = process_rate_card(file_path)
        
        df['Carrier agreement'] = general_info.get('carrier_agreement', '')
        df['Valid from'] = general_info.get('valid_from', '')
        df['Valid to'] = general_info.get('valid_to', '')
        df['Source file'] = file_path
        
        all_dataframes.append(df)
    
    # Combine dataframes
    combined_df = pd.concat(all_dataframes, ignore_index=True) if len(all_dataframes) > 1 else all_dataframes[0]
    
    return combined_df, combined_df.columns.tolist(), combined_conditions, combined_business_rules


def save_combined_rate_cards(file_paths, output_path=None, validate_columns=True):
    """Process multiple rate cards, combine them, and save to Excel."""
    combined_df, column_names, combined_conditions, combined_business_rules = \
        process_multiple_rate_cards(file_paths, validate_columns)
    
    if output_path is None:
        output_path = os.path.join("input", "rate_card_modified.xlsx")
    
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    
    business_rules_conditions = transform_business_rules_to_conditions(combined_business_rules)
    
    # Create conditions DataFrame
    conditions_data = []
    for col_name in column_names:
        if col_name in ['Carrier agreement', 'Valid from', 'Valid to', 'Source file']:
            continue
        raw_condition = combined_conditions.get(col_name, "")
        cleaned_condition = clean_condition_text(raw_condition) if raw_condition else ""
        conditions_data.append({
            'Column': col_name,
            'Has Condition': 'Yes' if col_name in combined_conditions else 'No',
            'Condition Rule': cleaned_condition
        })
    
    # Save to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name='Rate Card Data', index=False)
        pd.DataFrame(conditions_data).to_excel(writer, sheet_name='Conditions', index=False)
        
        business_rules_data = [{
            'Rule Name': name,
            'Section': cond.get('section', '').replace('_', ' ').title(),
            'Country': cond.get('country', ''),
            'Postal Codes': cond.get('raw_postal_code', ''),
            'Exclude': 'Yes' if cond.get('exclude') else 'No'
        } for name, cond in business_rules_conditions.items()]
        
        if business_rules_data:
            pd.DataFrame(business_rules_data).to_excel(writer, sheet_name='Business Rules', index=False)
        
        pd.DataFrame({
            'Metric': ['Total Rows', 'Total Columns', 'Source Files', 'Conditions', 'Business Rules'],
            'Value': [len(combined_df), len(column_names), ', '.join(file_paths),
                     len(combined_conditions), len(combined_business_rules.get('raw_rules', []))]
        }).to_excel(writer, sheet_name='Summary', index=False)
    
    return output_path


def process_rate_card_from_combined(combined_file_path):
    """Process a previously combined rate card file."""
    df = pd.read_excel(combined_file_path, sheet_name='Rate Card Data')
    
    conditions = {}
    try:
        df_conditions = pd.read_excel(combined_file_path, sheet_name='Conditions')
        for _, row in df_conditions.iterrows():
            if row.get('Has Condition') == 'Yes' and row.get('Condition Rule'):
                conditions[row['Column']] = row['Condition Rule']
    except Exception:
        pass
    
    return df, df.columns.tolist(), conditions


def process_rate_card_extended(file_paths, validate_columns=True):
    """Extended rate card processor that handles both single and multiple files."""
    if isinstance(file_paths, str):
        return process_rate_card(file_paths)
    
    if isinstance(file_paths, list) and len(file_paths) == 1:
        return process_rate_card(file_paths[0])
    
    combined_df, column_names, combined_conditions, _ = process_multiple_rate_cards(file_paths, validate_columns)
    return combined_df, column_names, combined_conditions


# Alias for backward compatibility
process_rate_cards = process_rate_card_extended


# ============================================================================
# Google Colab File Upload Functions
# ============================================================================

def upload_and_merge_rate_cards(cleanup=True):
    """
    Upload and merge multiple rate card files in Google Colab.
    Opens file upload dialog and processes the uploaded files.
    
    Args:
        cleanup: If True, removes uploaded source files after merging (keeps only merged file)
    """
    try:
        from google.colab import files
    except ImportError:
        print("❌ This function is for Google Colab. Use save_combined_rate_cards() locally.")
        return None
    
    os.makedirs("input", exist_ok=True)
    
    print("📤 Upload rate card files (select multiple with Ctrl/Cmd)...")
    uploaded = files.upload()
    
    if not uploaded:
        print("❌ No files uploaded.")
        return None
    
    # Track files for cleanup
    files_to_cleanup = []
    
    # Copy to input folder
    file_paths = []
    for filename, content in uploaded.items():
        if filename.endswith(('.xlsx', '.xls')):
            input_file_path = os.path.join("input", filename)
            with open(input_file_path, 'wb') as f:
                f.write(content)
            file_paths.append(filename)
            
            # Track for cleanup: file in /content (uploaded by Colab) and in /content/input
            files_to_cleanup.append(os.path.join("/content", filename))
            files_to_cleanup.append(input_file_path)
    
    if not file_paths:
        print("❌ No valid Excel files.")
        return None
    
    if len(file_paths) == 1:
        print(f"✓ 1 file uploaded: {file_paths[0]}")
        return file_paths[0]
    
    # Merge
    try:
        output_path = save_combined_rate_cards(file_paths)
        print(f"✅ Merged {len(file_paths)} files → {output_path}")
        
        # Cleanup uploaded source files
        if cleanup:
            cleaned = 0
            for filepath in files_to_cleanup:
                try:
                    if os.path.exists(filepath) and 'rate_card_modified' not in filepath:
                        os.remove(filepath)
                        cleaned += 1
                except Exception:
                    pass
            if cleaned > 0:
                print(f"🧹 Cleaned up {cleaned} source file(s)")
        
        return output_path
    except ValueError as e:
        print(f"❌ Merge failed: {e}")
        return None


def merge_rate_cards_from_folder(folder_path="input", pattern="*.xlsx"):
    """Merge all rate card files from a folder."""
    import glob
    
    files = [f for f in glob.glob(os.path.join(folder_path, pattern)) 
             if 'rate_card_modified' not in f.lower()]
    
    if not files:
        print(f"❌ No files found in '{folder_path}'")
        return None
    
    file_names = [os.path.basename(f) for f in files]
    
    if len(file_names) == 1:
        print(f"✓ 1 file found: {file_names[0]}")
        return file_names[0]
    
    try:
        output_path = save_combined_rate_cards(file_names)
        print(f"✅ Merged {len(file_names)} files → {output_path}")
        return output_path
    except ValueError as e:
        print(f"❌ Merge failed: {e}")
        return None


if __name__ == "__main__":
    try:
        from google.colab import files
        upload_and_merge_rate_cards()
    except ImportError:
        # Local: try to merge from input folder
        result = merge_rate_cards_from_folder("input")
        if not result:
            print("\nUsage:")
            print("  save_combined_rate_cards(['file1.xlsx', 'file2.xlsx'])")
