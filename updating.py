"""
Update Google Drive Data Collection file with Pivot Data from matching.py output.

This script:
1. Reads Pivot Data from local Output folder (Matched_Shipments_with.xlsx, sheet "Pivot Data")
2. Opens Data Collection.xlsx file on Google Drive
3. For each row from pivot:
   - If Shipper + Carrier + Cause combination exists: adds Amount to existing
   - If not exists: adds new row with the data
"""

import pandas as pd
import os


def update_data_collection(
    google_drive_path: str,
    local_output_folder: str = None,
    pivot_file_name: str = "Result.xlsx",
    pivot_sheet_name: str = "Pivot Data",
    collection_file_name: str = "Data Collection.xlsx"
):
    """
    Update Google Drive Data Collection file with Pivot Data.
    
    Args:
        google_drive_path: Path to folder on Google Drive (e.g., "My Drive/CANF/Data")
        local_output_folder: Path to local output folder. If None, uses script_dir/output
        pivot_file_name: Name of the file containing pivot data (default: "Matched_Shipments_with.xlsx")
        pivot_sheet_name: Sheet name containing pivot data (default: "Pivot Data")
        collection_file_name: Name of the Data Collection file on Google Drive (default: "Data Collection.xlsx")
    
    Returns:
        bool: True if successful, False otherwise
    """
    
    # Check if running in Google Colab
    try:
        from google.colab import drive
        in_colab = True
        #print("📁 Running in Google Colab - mounting Google Drive...")
        drive.mount('/content/drive')
        drive_base = "/content/drive"
    except ImportError:
        in_colab = False
        #print("⚠️ Not running in Google Colab.")
        #print("   For local execution, please ensure Google Drive is synced locally.")
        #print("   Or run this script in Google Colab.")
        # For local execution with Google Drive Desktop app
        # Common paths: 
        # Windows: "G:/My Drive" or "C:/Users/USERNAME/Google Drive"
        # Mac: "/Users/USERNAME/Google Drive"
        drive_base = ""  # User needs to provide full path
    
    # Step 1: Determine local output folder
    if local_output_folder is None:
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            script_dir = os.getcwd()
        local_output_folder = os.path.join(script_dir, "output")
    
    # Step 2: Read Pivot Data from local output
    pivot_file_path = os.path.join(local_output_folder, pivot_file_name)
    
    if not os.path.exists(pivot_file_path):
        print(f"❌ Error: Pivot file not found at: {pivot_file_path}")
        #print(f"   Please run matching.py first to generate the output file.")
        return False
    
    print(f"📄 Reading pivot data from: {pivot_file_path}")
    
    try:
        # Check if the sheet exists
        excel_file = pd.ExcelFile(pivot_file_path)
        if pivot_sheet_name not in excel_file.sheet_names:
            print(f"❌ Error: Sheet '{pivot_sheet_name}' not found in {pivot_file_name}")
            print(f"   Available sheets: {excel_file.sheet_names}")
            return False
        
        df_pivot = pd.read_excel(pivot_file_path, sheet_name=pivot_sheet_name)
        print(f"   Loaded {len(df_pivot)} rows from Pivot Data")
        print(f"   Columns: {list(df_pivot.columns)}")
        
        if df_pivot.empty:
            print("⚠️ Warning: Pivot Data is empty. Nothing to update.")
            return True
            
    except Exception as e:
        print(f"❌ Error reading pivot file: {e}")
        return False
    
    # Step 3: Construct Google Drive file path
    if in_colab:
        collection_path = os.path.join(google_drive_path, collection_file_name)
    
    print(f"📁 Google Drive file path: {collection_path}")
    
    # Step 4: Read or create Data Collection file
    if os.path.exists(collection_path):
        print(f"📄 Reading existing Data Collection file...")
        try:
            df_collection = pd.read_excel(collection_path)
            print(f"   Loaded {len(df_collection)} existing rows")
        except Exception as e:
            print(f"❌ Error reading Data Collection file: {e}")
            return False
    else:
        print(f"📄 Data Collection file not found. Creating new file...")
        # Create empty DataFrame with expected columns
        df_collection = pd.DataFrame(columns=['Shipper Value', 'Carrier', 'Cause of CANF', 'Amount'])
    
    # Step 5: Ensure required columns exist in both dataframes
    required_columns = ['Shipper Value', 'Carrier', 'Cause of CANF', 'Amount']
    
    # Check pivot columns (handle variations)
    pivot_column_mapping = {}
    for req_col in required_columns:
        found = False
        for col in df_pivot.columns:
            if col.lower().replace(' ', '').replace('_', '') == req_col.lower().replace(' ', '').replace('_', ''):
                pivot_column_mapping[req_col] = col
                found = True
                break
        if not found:
            print(f"❌ Error: Required column '{req_col}' not found in Pivot Data")
            print(f"   Available columns: {list(df_pivot.columns)}")
            return False
    
    # Rename pivot columns to standard names
    df_pivot = df_pivot.rename(columns={v: k for k, v in pivot_column_mapping.items()})
    
    # Ensure collection has all required columns
    for col in required_columns:
        if col not in df_collection.columns:
            df_collection[col] = None
    
    # Step 6: Update Data Collection with Pivot Data
    print(f"\n🔄 Updating Data Collection...")
    
    rows_updated = 0
    rows_added = 0
    
    for idx, pivot_row in df_pivot.iterrows():
        shipper = pivot_row['Shipper Value']
        carrier = pivot_row['Carrier']
        cause = pivot_row['Cause of CANF']
        amount = pivot_row['Amount']
        
        # Skip rows with missing key values
        if pd.isna(shipper) or pd.isna(carrier) or pd.isna(cause):
            print(f"   ⚠️ Skipping row {idx}: Missing key values (Shipper/Carrier/Cause)")
            continue
        
        # Convert amount to numeric
        try:
            amount = float(amount) if pd.notna(amount) else 0.0
        except (ValueError, TypeError):
            amount = 0.0
        
        # Find matching row in collection
        mask = (
            (df_collection['Shipper Value'].astype(str) == str(shipper)) &
            (df_collection['Carrier'].astype(str) == str(carrier)) &
            (df_collection['Cause of CANF'].astype(str) == str(cause))
        )
        
        matching_rows = df_collection[mask]
        
        if len(matching_rows) > 0:
            # Update existing row - add amount
            existing_idx = matching_rows.index[0]
            existing_amount = df_collection.loc[existing_idx, 'Amount']
            try:
                existing_amount = float(existing_amount) if pd.notna(existing_amount) else 0.0
            except (ValueError, TypeError):
                existing_amount = 0.0
            
            new_amount = existing_amount + amount
            df_collection.loc[existing_idx, 'Amount'] = new_amount
            rows_updated += 1
            #print(f"   ✓ Updated: {shipper} | {carrier} | {cause} | {existing_amount} + {amount} = {new_amount}")
        else:
            # Add new row
            new_row = pd.DataFrame([{
                'Shipper Value': shipper,
                'Carrier': carrier,
                'Cause of CANF': cause,
                'Amount': amount
            }])
            df_collection = pd.concat([df_collection, new_row], ignore_index=True)
            rows_added += 1
            #print(f"   + Added: {shipper} | {carrier} | {cause} | {amount}")
    
    # Step 7: Save updated Data Collection with formatting
    print(f"\n💾 Saving updated Data Collection with formatting...")
    
    try:
        # Save to Excel first
        df_collection.to_excel(collection_path, index=False, engine='openpyxl')
        
        # Apply visual formatting
        apply_excel_formatting(collection_path)
        
        print(f"   ✅ Saved to: {collection_path}")
        print(f"\n📊 Summary:")
        print(f"   Rows updated: {rows_updated}")
        print(f"   Rows added: {rows_added}")
        print(f"   Total rows in collection: {len(df_collection)}")
        return True
    except Exception as e:
        print(f"❌ Error saving Data Collection: {e}")
        return False


def apply_excel_formatting(file_path: str):
    """
    Apply visual formatting to the Data Collection Excel file.
    
    Features:
    - Header styling (bold, dark blue background, white text)
    - Column width auto-adjustment
    - Number formatting for Amount column
    - Alternating row colors
    - Borders
    - Freeze header row
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
    
    print("   🎨 Applying visual formatting...")
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")  # White, bold
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Alternating row colors
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Header border (thicker bottom)
    header_border = Border(
        left=Side(style='thin', color='1F4E79'),
        right=Side(style='thin', color='1F4E79'),
        top=Side(style='thin', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79')
    )
    
    # Data cell alignment
    data_alignment = Alignment(horizontal="left", vertical="center")
    amount_alignment = Alignment(horizontal="right", vertical="center")
    
    # Format header row (row 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Format data rows with alternating colors
    for row_num in range(2, ws.max_row + 1):
        # Alternating row colors
        if row_num % 2 == 0:
            row_fill = light_gray_fill
        else:
            row_fill = white_fill
        
        for cell in ws[row_num]:
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=11)
        
        # Set row height
        ws.row_dimensions[row_num].height = 22
    
    # Auto-adjust column widths
    column_widths = {
        'A': 20,  # Shipper Value
        'B': 18,  # Carrier
        'C': 35,  # Cause of CANF
        'D': 15,  # Amount
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Also check actual content and adjust if needed
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Use max of defined width or content width (with some padding)
        defined_width = column_widths.get(column_letter, 15)
        adjusted_width = max(defined_width, min(max_length + 3, 50))
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add filter to header row
    ws.auto_filter.ref = ws.dimensions
    
    # Save the formatted workbook
    wb.save(file_path)
    #print("   ✓ Header styling (dark blue with white text)")
    #print("   ✓ Alternating row colors")
    #print("   ✓ Number formatting for Amount column")
    #print("   ✓ Auto-adjusted column widths")
    #print("   ✓ Frozen header row")
    #print("   ✓ Auto-filter enabled")


def update_from_colab(google_drive_folder_path: str):
    """
    Convenience function for Google Colab usage.
    
    Args:
        google_drive_folder_path: Path to folder containing Data Collection.xlsx
                                  (e.g., "My Drive/CANF/Data" or just "CANF/Data")
    
    Example:
        update_from_colab("My Drive/CANF Reports/2024")
    """
    return update_data_collection(google_drive_path=google_drive_folder_path)


# Example usage
if __name__ == "__main__":
    # For Google Colab - provide the path to your Google Drive folder
    # Example: "My Drive/CANF/Reports" or "CANF/Reports"
    
    # Uncomment and modify the path below:
    update_from_colab("/content/drive/Shareddrives/FA Operations Europe/AT/Generic_AA/CANF Analyzer")
    
