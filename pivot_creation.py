import pandas as pd
import os
import glob
import re
from collections import Counter

def clean_comment_line(line):
    """
    Clean a single comment line by normalizing to patterns (removing specific values).
    
    Examples:
        "Destination Postal Code: Shipment value '12230' needs to be changed to '43300'"
        -> "Destination Postal Code: Shipment value needs to be changed"
        
        "Origin Country: Shipment value 'US' needs to be changed to 'USA'"
        -> "Origin Country: Shipment value needs to be changed"
        
        "Date '20241201' is outside valid date range..."
        -> "Date is outside valid date range..."
    """
    if pd.isna(line) or line == '':
        return None
    line_stripped = str(line).strip()
    
    # Remove 'Discrepancies for Match' lines
    if line_stripped.startswith('Discrepancies for Match'):
        return None
    
    # Skip lines about "possible rate lanes"
    if 'possible rate lanes' in line_stripped.lower():
        return None
    
    # Pattern 1: "Field: Shipment value 'X' needs to be changed to 'Y'"
    # -> "Field: Shipment value needs to be changed"
    match = re.match(r"^(.+?):\s*Shipment value\s*'[^']*'\s*needs to be changed to\s*'[^']*'\.?$", line_stripped)
    if match:
        field_name = match.group(1).strip()
        return f"{field_name}: Shipment value needs to be changed"
    
    # Pattern 2: "Field: Rate Card value 'X' - Shipment has 'Y'"
    # -> "Field: Rate Card value differs from Shipment"
    match = re.match(r"^(.+?):\s*Rate Card value\s*'[^']*'\s*-\s*Shipment has\s*'[^']*'\.?$", line_stripped)
    if match:
        field_name = match.group(1).strip()
        return f"{field_name}: Rate Card value differs from Shipment"
    
    # Pattern 3: "Field: needs to be changed from 'X' to 'Y'"
    # -> "Field: needs to be changed"
    match = re.match(r"^(.+?):\s*needs to be changed from\s*'[^']*'\s*to\s*'[^']*'\.?$", line_stripped)
    if match:
        field_name = match.group(1).strip()
        return f"{field_name}: needs to be changed"
    
    # Pattern 4: Normalize date comments
    # "Date 'YYYYMMDD' is outside valid date range..." -> "Date is outside valid date range..."
    if "Date '" in line_stripped and "is outside valid date range" in line_stripped:
        if "for all matching rate card entries" in line_stripped:
            return "Date is outside valid date range for all matching rate card entries"
        else:
            cleaned_line = re.sub(r"Date '[^']+'", "Date", line_stripped)
            return cleaned_line
    
    # Pattern 5: Generic - remove any quoted values from the line
    # This catches other patterns we might have missed
    cleaned = re.sub(r"'[^']*'", "", line_stripped)
    # Clean up extra spaces
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    # Clean up phrases like "needs to be changed to" -> "needs to be changed"
    cleaned = re.sub(r'needs to be changed to\s*$', 'needs to be changed', cleaned)
    cleaned = re.sub(r'needs to be changed to\s*\.', 'needs to be changed.', cleaned)
    
    return cleaned


def extract_discrepancy_field(comment):
    """
    Extract the field/column name from a discrepancy comment.
    Examples:
        "Origin Postal Code: Shipment value '19454' needs to be changed to '19454413'" -> "Origin Postal Code"
        "Destination Country: Shipment value 'US' needs to be changed to 'USA'" -> "Destination Country"
        "Please recheck the shipment details" -> "General"
    """
    if pd.isna(comment) or comment == '':
        return None
    
    comment_str = str(comment).strip()
    
    # Skip pattern analysis lines
    if comment_str.startswith('Discrepancies for Match'):
        return None
    if 'possible rate lanes' in comment_str.lower():
        return None
    
    # Pattern 1: "Field Name: Shipment value..." or "Field Name: needs to be changed..."
    match = re.match(r'^([^:]+):\s*(Shipment value|needs to be changed|Rate Card value)', comment_str)
    if match:
        return match.group(1).strip()
    
    # Pattern 2: "Field Name: value..." (generic colon pattern)
    if ':' in comment_str:
        field = comment_str.split(':')[0].strip()
        # Only return if it looks like a field name (not too long, no sentences)
        if len(field) < 50 and not any(word in field.lower() for word in ['please', 'recheck', 'shipment', 'too many']):
            return field
    
    # Pattern 3: Check for known generic messages
    generic_patterns = [
        'Please recheck the shipment details',
        'Too many shipment details to update',
        'Date is outside valid date range',
        'No matching rate card'
    ]
    for pattern in generic_patterns:
        if pattern.lower() in comment_str.lower():
            return pattern
    
    return "Other"


def analyze_discrepancy_patterns(comments_list):
    """
    Analyze a list of comments to find the most common discrepancy patterns.
    
    Returns:
        dict with pattern analysis results:
        - field_counts: Counter of field occurrences
        - dominant_field: Most common field (if any)
        - dominant_percentage: Percentage of dominant field
        - pattern_summary: List of (field, count, percentage) tuples
    """
    field_counts = Counter()
    
    for comment in comments_list:
        if pd.isna(comment) or comment == '':
            continue
        
        # Split multi-line comments
        lines = str(comment).split('\n')
        for line in lines:
            field = extract_discrepancy_field(line)
            if field:
                field_counts[field] += 1
    
    total = sum(field_counts.values())
    if total == 0:
        return {
            'field_counts': field_counts,
            'dominant_field': None,
            'dominant_percentage': 0,
            'pattern_summary': []
        }
    
    # Calculate percentages and sort
    pattern_summary = []
    for field, count in field_counts.most_common():
        percentage = (count / total) * 100
        pattern_summary.append((field, count, percentage))
    
    # Determine dominant field (>30% of all discrepancies)
    dominant_field = None
    dominant_percentage = 0
    if pattern_summary:
        top_field, top_count, top_pct = pattern_summary[0]
        if top_pct >= 30:  # Considered dominant if >= 30%
            dominant_field = top_field
            dominant_percentage = top_pct
    
    return {
        'field_counts': field_counts,
        'dominant_field': dominant_field,
        'dominant_percentage': dominant_percentage,
        'pattern_summary': pattern_summary
    }

def update_canf_file(matching_output_file=None,
                     shipper_value=None):
    """
    Process matching.py output file and add pivot data tab to the original file.

    Args:
        matching_output_file: Path to the matching.py output Excel file (Matched_Shipments_with.xlsx)
        shipper_value: Shipper value to add to the new tab
    """
    try:
        # Find matching.py output file
        if matching_output_file is None:
            # Try to find Matched_Shipments_with.xlsx in common locations
            script_dir = os.path.dirname(os.path.abspath(__file__))
            possible_locations = [
                os.path.join(script_dir, "Matched_Shipments_with.xlsx"),
                os.path.join(os.path.dirname(script_dir), "test folder", "Matched_Shipments_with.xlsx"),
                "Matched_Shipments_with.xlsx"
            ]
            
            for loc in possible_locations:
                if os.path.exists(loc):
                    matching_output_file = loc
                    print(f"Found matching output file: {matching_output_file}")
                    break
            
            if matching_output_file is None:
                print(f"Error: Could not find Matched_Shipments_with.xlsx in any of these locations:")
                for loc in possible_locations:
                    print(f"  - {loc}")
                return False
        elif not os.path.exists(matching_output_file):
            print(f"Error: Matching output file not found: {matching_output_file}")
            return False

        print(f"Reading matching output file: {matching_output_file}")

        # Read the "Matched Shipments" sheet from matching.py output
        try:
            df_etofs = pd.read_excel(matching_output_file, sheet_name='Matched Shipments')
            print(f"Loaded {len(df_etofs)} rows from 'Matched Shipments' sheet")
        except Exception as e:
            print(f"Error reading 'Matched Shipments' sheet: {e}")
            # Try reading the first sheet as fallback
            df_etofs = pd.read_excel(matching_output_file)
            print(f"Loaded {len(df_etofs)} rows from first sheet")

        # Prepare data for Google Sheets: Carrier, Cause of CANF, Amount
        # Check for 'comment' column (matching.py uses 'comment', not 'Comments')
        comment_col = None
        if 'comment' in df_etofs.columns:
            comment_col = 'comment'
        elif 'Comments' in df_etofs.columns:
            comment_col = 'Comments'
        
        if 'Carrier' in df_etofs.columns and comment_col:
            # Create cross-product of Carrier and cleaned Comments
            # First, merge Carrier with cleaned comments
            carrier_cause_df = df_etofs[['Carrier', comment_col]].copy()
            carrier_cause_df.rename(columns={comment_col: 'Comments'}, inplace=True)
            
            # Expand comments: split multi-line comments into separate rows
            # Each line becomes a separate row for proper grouping
            expanded_rows = []
            for idx, row in carrier_cause_df.iterrows():
                carrier = row['Carrier']
                comment = row['Comments']
                
                if pd.isna(comment) or comment == '':
                    continue
                
                # Split comment into lines and clean each line
                comment_lines = str(comment).split('\n')
                for line in comment_lines:
                    # Clean the line (normalize date comments, remove "Discrepancies for Match")
                    cleaned_line = clean_comment_line(line)
                    
                    # Only add if cleaned_line is not None and not empty
                    if cleaned_line and cleaned_line.strip() != '':
                        expanded_rows.append({
                            'Carrier': carrier,
                            'Cause of CANF': cleaned_line
                        })
            
            # Create dataframe from expanded rows
            if expanded_rows:
                expanded_df = pd.DataFrame(expanded_rows)
                
                # Count occurrences of each Carrier + Cause combination
                google_sheets_data = expanded_df.groupby(['Carrier', 'Cause of CANF']).size().reset_index(name='Amount')
                
                # Add shipper value column to the pivot data
                if shipper_value:
                    google_sheets_data['Shipper Value'] = shipper_value
                else:
                    google_sheets_data['Shipper Value'] = 'Not provided'
                
                # Reorder columns: Shipper Value, Carrier, Cause of CANF, Amount
                google_sheets_data = google_sheets_data[['Shipper Value', 'Carrier', 'Cause of CANF', 'Amount']]
                google_sheets_data = google_sheets_data.sort_values(['Carrier', 'Cause of CANF'])
                
                # ========== PATTERN ANALYSIS ==========
                # Analyze overall discrepancy patterns
                all_comments = df_etofs[comment_col].dropna().tolist()
                overall_patterns = analyze_discrepancy_patterns(all_comments)
                
                # Create Pattern Summary DataFrame
                pattern_rows = []
                for field, count, percentage in overall_patterns['pattern_summary']:
                    pattern_rows.append({
                        'Discrepancy Field': field,
                        'Count': count,
                        'Percentage': f"{percentage:.1f}%",
                        'Is Dominant': 'YES' if field == overall_patterns['dominant_field'] else ''
                    })
                pattern_summary_df = pd.DataFrame(pattern_rows)
                
                # Analyze patterns PER CARRIER
                carrier_pattern_rows = []
                for carrier in df_etofs['Carrier'].dropna().unique():
                    carrier_comments = df_etofs[df_etofs['Carrier'] == carrier][comment_col].dropna().tolist()
                    carrier_patterns = analyze_discrepancy_patterns(carrier_comments)
                    
                    for field, count, percentage in carrier_patterns['pattern_summary'][:5]:  # Top 5 per carrier
                        carrier_pattern_rows.append({
                            'Carrier': carrier,
                            'Discrepancy Field': field,
                            'Count': count,
                            'Percentage': f"{percentage:.1f}%",
                            'Is Dominant': 'YES' if field == carrier_patterns['dominant_field'] else ''
                        })
                
                carrier_patterns_df = pd.DataFrame(carrier_pattern_rows)
                
                # Print pattern analysis summary
                print(f"\n{'='*60}")
                print("DISCREPANCY PATTERN ANALYSIS")
                print('='*60)
                if overall_patterns['dominant_field']:
                    print(f"  DOMINANT ISSUE: {overall_patterns['dominant_field']} ({overall_patterns['dominant_percentage']:.1f}%)")
                else:
                    print("  No single dominant issue found (discrepancies are diverse)")
                print(f"\n  Top discrepancy fields:")
                for field, count, percentage in overall_patterns['pattern_summary'][:10]:
                    marker = " <<<" if field == overall_patterns['dominant_field'] else ""
                    print(f"    - {field}: {count} ({percentage:.1f}%){marker}")
                print('='*60)
                
            else:
                google_sheets_data = pd.DataFrame(columns=['Shipper Value', 'Carrier', 'Cause of CANF', 'Amount'])
                pattern_summary_df = pd.DataFrame(columns=['Discrepancy Field', 'Count', 'Percentage', 'Is Dominant'])
                carrier_patterns_df = pd.DataFrame(columns=['Carrier', 'Discrepancy Field', 'Count', 'Percentage', 'Is Dominant'])

            print(f"\nPrepared {len(google_sheets_data)} Carrier-Cause combinations.")

            # Update the original matching.py output file by adding new sheets
            try:
                # Read all existing sheets from the matching output file
                excel_file = pd.ExcelFile(matching_output_file)
                existing_sheets = {}
                
                for sheet_name in excel_file.sheet_names:
                    existing_sheets[sheet_name] = pd.read_excel(matching_output_file, sheet_name=sheet_name)
                    print(f"  Preserved existing sheet: '{sheet_name}' ({len(existing_sheets[sheet_name])} rows)")
                
                # Save all sheets back to the original file (existing sheets + new sheets)
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter
                    FORMATTING_AVAILABLE = True
                except ImportError:
                    FORMATTING_AVAILABLE = False
                
                with pd.ExcelWriter(matching_output_file, engine='openpyxl') as writer:
                    # Write all existing sheets
                    for sheet_name, sheet_df in existing_sheets.items():
                        sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Add new Pivot Data sheet
                    google_sheets_data = google_sheets_data.sort_values(['Carrier', 'Cause of CANF']).reset_index(drop=True)
                    google_sheets_data.to_excel(writer, sheet_name='Pivot Data', index=False)
                    
                    # Add Pattern Summary sheet (overall patterns)
                    pattern_summary_df.to_excel(writer, sheet_name='Pattern Summary', index=False)
                    
                    # Add Carrier Patterns sheet (patterns per carrier)
                    if not carrier_patterns_df.empty:
                        carrier_patterns_df.to_excel(writer, sheet_name='Carrier Patterns', index=False)
                    
                    # Apply formatting if available
                    if FORMATTING_AVAILABLE:
                        workbook = writer.book
                        
                        # Format all sheets
                        for sheet_name in workbook.sheetnames:
                            ws = workbook[sheet_name]
                            
                            # Determine header color based on sheet name
                            if sheet_name == 'Matched Shipments':
                                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            elif sheet_name == 'Rate Card Reference':
                                header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                            elif sheet_name == 'Pivot Data':
                                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            elif sheet_name == 'Pattern Summary':
                                header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")  # Orange
                            elif sheet_name == 'Carrier Patterns':
                                header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # Purple
                            else:
                                # Default header color for other sheets
                                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            
                            header_font = Font(bold=True, color="FFFFFF", size=11)
                            
                            # Style header row
                            for cell in ws[1]:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            
                            # Auto-adjust column widths
                            for column in ws.columns:
                                max_length = 0
                                column_letter = get_column_letter(column[0].column)
                                
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                
                                # Set width with some padding, but cap at 50
                                adjusted_width = min(max_length + 2, 50)
                                ws.column_dimensions[column_letter].width = adjusted_width
                            
                            # Freeze header row
                            ws.freeze_panes = 'A2'
                            
                            # Special formatting for specific sheets
                            if sheet_name == 'Pivot Data':
                                # Make "Cause of CANF" column wider for better readability
                                if 'Cause of CANF' in google_sheets_data.columns:
                                    cause_col_idx = list(google_sheets_data.columns).index('Cause of CANF') + 1
                                    cause_col_letter = get_column_letter(cause_col_idx)
                                    ws.column_dimensions[cause_col_letter].width = 60
                                    
                                    # Wrap text in Cause of CANF column
                                    for row in ws.iter_rows(min_row=2, min_col=cause_col_idx, max_col=cause_col_idx):
                                        for cell in row:
                                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                            
                            elif sheet_name == 'Matched Shipments':
                                # Make comment column wider and wrap text
                                if 'comment' in existing_sheets.get('Matched Shipments', pd.DataFrame()).columns:
                                    comment_col_idx = list(existing_sheets['Matched Shipments'].columns).index('comment') + 1
                                    comment_col_letter = get_column_letter(comment_col_idx)
                                    ws.column_dimensions[comment_col_letter].width = 60
                                    
                                    # Wrap text in comment column
                                    for row in ws.iter_rows(min_row=2, min_col=comment_col_idx, max_col=comment_col_idx):
                                        for cell in row:
                                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                            
                            elif sheet_name == 'Pattern Summary':
                                # Highlight dominant rows
                                dominant_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Light yellow
                                for row in ws.iter_rows(min_row=2):
                                    # Check if "Is Dominant" column has "YES"
                                    if len(row) >= 4 and row[3].value == 'YES':
                                        for cell in row:
                                            cell.fill = dominant_fill
                                            cell.font = Font(bold=True)
                                
                                # Set column widths
                                ws.column_dimensions['A'].width = 40  # Discrepancy Field
                                ws.column_dimensions['B'].width = 12  # Count
                                ws.column_dimensions['C'].width = 12  # Percentage
                                ws.column_dimensions['D'].width = 14  # Is Dominant
                            
                            elif sheet_name == 'Carrier Patterns':
                                # Highlight dominant rows
                                dominant_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
                                for row in ws.iter_rows(min_row=2):
                                    # Check if "Is Dominant" column has "YES"
                                    if len(row) >= 5 and row[4].value == 'YES':
                                        for cell in row:
                                            cell.fill = dominant_fill
                                            cell.font = Font(bold=True)
                                
                                # Set column widths
                                ws.column_dimensions['A'].width = 25  # Carrier
                                ws.column_dimensions['B'].width = 40  # Discrepancy Field
                                ws.column_dimensions['C'].width = 12  # Count
                                ws.column_dimensions['D'].width = 12  # Percentage
                                ws.column_dimensions['E'].width = 14  # Is Dominant
                
                print(f"\nSuccessfully updated file '{matching_output_file}'!")
                print(f"  - Preserved {len(existing_sheets)} existing sheet(s)")
                print(f"  - Added 'Pivot Data' sheet with {len(google_sheets_data)} rows")
                print(f"  - Added 'Pattern Summary' sheet with {len(pattern_summary_df)} patterns")
                print(f"  - Added 'Carrier Patterns' sheet with {len(carrier_patterns_df)} carrier-specific patterns")

            except Exception as e:
                print(f"Error updating file: {str(e)}")
                import traceback
                traceback.print_exc()
                return False
        else:
            print(f"Carrier or Comments column not found. Available columns: {list(df_etofs.columns)}")
            print("Skipping Excel file update.")
            return False

        return True

    except Exception as e:
        print(f"Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

# Example usage:
#if __name__ == "__main__":
    # USER INPUT: Provide shipper value here
#    SHIPPER_VALUE = "Your Shipper Value Here"  # Change this to your shipper value
    
#    update_canf_file(
 #       matching_output_file=None,  # Will auto-detect Matched_Shipments_with.xlsx
 #       shipper_value=SHIPPER_VALUE
 #   )
